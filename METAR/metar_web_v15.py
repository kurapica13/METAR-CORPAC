"""
METAR DIGITAL - VERSIÓN PROFESIONAL CORPAC PERÚ
Aeropuerto Internacional Jorge Chávez (SPJC)
Características:
✅ RVR - Formato SPJC: R32/0400, R12R/1700, R10/M0050, R14L/P2000 
✅ Sin validaciones arbitrarias - El operador copia lo que ve en el equipo
✅ Almacenamiento mensual automático (SPJC_METAR_YYYY_MM.xlsx)
✅ Viento con reglas circulares (340V080) - CORREGIDO: No muestra variación <60°
✅ Visibilidad mínima con cuadrantes (N, NE, E, SE, S, SW, W, NW)
✅ Estándar oficial nubes CORPAC (30m/1000m)
✅ Fenómenos especiales (PRFG, VCFG, BCFG, MIFG)
✅ Excel con formato profesional mensual
✅ Sin duplicados - Reemplaza reportes con misma fecha/hora
✅ Persistencia de datos entre sesiones
✅ Exportación simplificada con formato YYYY_MM
✅ Columna METAR con ancho automático en Excel
✅ Temperatura y Rocío - CORREGIDO: Ambos se redondean igual
✅ Sistema de autenticación - Solo personal autorizado
✅ Nubes - CORREGIDO: Respeta texto claro (7 ST 300M → OVC010)
✅ Visibilidad Vertical - Soporta "vis ver 600M" → VV020
"""

import streamlit as st
from datetime import datetime, timezone
import pandas as pd
from pathlib import Path
import re
import os
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
import hmac

# ============================================
# CONFIGURACIÓN DE PÁGINA
# ============================================
st.set_page_config(
    page_title="REGISTRO DE OBSERVACIONES ORDINARIAS Y ESPECIALES SPJC",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ============================================
# SISTEMA DE AUTENTICACIÓN
# ============================================
def verificar_autenticacion():
    """Sistema de login mejorado - SIN FORM BLANCO"""
    
    if 'autenticado' not in st.session_state:
        st.session_state.autenticado = False
        st.session_state.usuario = None
    
    with st.sidebar:
        if st.session_state.autenticado:
            st.markdown(f"👤 **Usuario:** {st.session_state.usuario}")
            if st.button("🚪 Cerrar Sesión", use_container_width=True):
                st.session_state.autenticado = False
                st.session_state.usuario = None
                st.rerun()
            st.markdown("---")
    
    if st.session_state.autenticado:
        return True
    
    # CSS limpio sin form
    st.markdown("""
    <style>
        [data-testid="stAppViewContainer"] {
            background: linear-gradient(135deg, #e8f0fe 0%, #d3e3fd 100%) !important;
        }
        [data-testid="stHeader"], footer, header { display: none !important; }
        
        .block-container {
            max-width: 420px !important;
            padding: 8vh 20px 40px 20px !important;
        }
        
        .login-card {
            background: white;
            border-radius: 24px;
            padding: 48px 40px 40px 40px;
            box-shadow: 0 12px 48px rgba(11,61,145,0.12);
            text-align: center;
        }
        
        .logo-corpac {
            width: 140px;
            height: auto;
            margin: 0 auto 24px auto;
            display: block;
        }
        
        h1.login-title {
            color: #0b3d91 !important;
            font-size: 1.6rem !important;
            font-weight: 700 !important;
            margin: 0 0 8px 0 !important;
        }
        
        .login-subtitle {
            color: #666;
            font-size: 0.92rem;
            margin-bottom: 32px;
        }
        
        /* Inputs mejorados */
        .stTextInput > div > div > input {
            border-radius: 12px !important;
            border: 2px solid #e0e7ef !important;
            padding: 14px 18px !important;
            font-size: 0.95rem !important;
            transition: all 0.2s !important;
        }
        
        .stTextInput > div > div > input:focus {
            border-color: #0b3d91 !important;
            box-shadow: 0 0 0 3px rgba(11,61,145,0.08) !important;
        }
        
        /* Botón mejorado */
        .stButton > button {
            background: linear-gradient(135deg, #0b3d91 0%, #1a4fa0 100%) !important;
            color: white !important;
            border: none !important;
            border-radius: 12px !important;
            padding: 14px 24px !important;
            font-size: 1rem !important;
            font-weight: 600 !important;
            width: auto !important;
            min-width: 180px !important;
            margin-top: 20px !important;
            transition: all 0.3s !important;
            display: block !important;
            margin-left: auto !important;
            margin-right: auto !important;
        }
        
        .stButton > button:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 8px 20px rgba(11,61,145,0.25) !important;
        }
        
        /* Ocultar labels */
        .stTextInput > label { display: none !important; }
        
        /* Centrar el contenedor del botón */
        .stButton {
            display: flex !important;
            justify-content: center !important;
        }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown("<div class='login-card'>", unsafe_allow_html=True)
    st.markdown("<h1 class='login-title'>METAR Digital · SPJC</h1>", unsafe_allow_html=True)
    st.markdown("<div class='login-subtitle'>Solo personal autorizado</div>", unsafe_allow_html=True)
    
    # Inputs (sin form!)
    usuario = st.text_input("Usuario", placeholder="Usuario", key='login_usr')
    password = st.text_input("Contraseña", type="password", placeholder="Contraseña", key='login_pwd')
    
    # Botón
    if st.button("🔐 INGRESAR", key='btn_login_submit'):
        try:
            passwords = st.secrets.get("passwords", {})
            if not passwords:
                passwords = {"admin": "corpac2024", "metar": "spjc2024"}
        except:
            passwords = {"admin": "corpac2024", "metar": "spjc2024"}
        
        if usuario in passwords and hmac.compare_digest(password, passwords[usuario]):
            st.session_state.autenticado = True
            st.session_state.usuario = usuario
            st.rerun()
        else:
            st.error("❌ Credenciales incorrectas")
    
    st.markdown("</div>", unsafe_allow_html=True)
    st.stop()



# ============================================
# VERIFICAR AUTENTICACIÓN
# ============================================
verificar_autenticacion()

# ============================================
# CONSTANTES Y CONFIGURACIÓN
# ============================================
DIRECTORIO_DATOS = Path("datos_metar")
DIRECTORIO_DATOS.mkdir(exist_ok=True)

# ============================================
# FUNCIÓN DE REDONDEO CORREGIDA
# ============================================
def redondear_metar(valor):
    """
    Redondeo tradicional para METAR:
    - 14.5 → 15
    - 14.4 → 14
    - 15.5 → 16
    """
    try:
        d = Decimal(str(valor))
        return int(d.quantize(Decimal('1'), rounding=ROUND_HALF_UP))
    except:
        return int(round(float(valor)))

# ============================================
# ESTILOS CSS PERSONALIZADOS
# ============================================
st.markdown("""
<style>
    .section-title {
        color: #0b3d91 !important;
        font-weight: 700 !important;
        font-size: 1rem !important;
        margin-bottom: 0.6rem !important;
        border-bottom: 2px solid #0b3d91 !important;
        padding-bottom: 0.3rem !important;
    }
    .stTextInput label, .stSelectbox label, .stNumberInput label {
        color: #0b3d91 !important;
        font-weight: 600 !important;
        font-size: 0.85rem !important;
    }
    @media (prefers-color-scheme: dark) {
        .stTextInput label, .stSelectbox label, .stNumberInput label { color: #6ab0ff !important; }
        .section-title { color: #6ab0ff !important; border-bottom-color: #6ab0ff !important; }
    }
    .metar-box {
        background: #1e1e1e;
        color: #00ff00;
        padding: 1rem;
        border-radius: 5px;
        font-family: 'Courier New', monospace;
        font-size: 1.05rem;
        border-left: 5px solid #0b3d91;
        word-break: break-all;
    }
    .historial-item {
        background: #f8f9fa;
        padding: 0.6rem 0.8rem;
        margin-bottom: 0.4rem;
        border-radius: 4px;
        font-family: 'Courier New', monospace;
        font-size: 11px;
        border-left: 3px solid #0b3d91;
    }
    .capa-nube-box {
        background: rgba(11,61,145,0.06);
        border: 1px solid rgba(11,61,145,0.20);
        border-radius: 8px;
        padding: 10px 12px 6px 12px;
        margin-bottom: 8px;
    }
    .fenomeno-tag {
        display: inline-block;
        background: #e8f0fe;
        color: #0b3d91;
        border-radius: 12px;
        padding: 2px 10px;
        font-size: 0.82rem;
        font-weight: 600;
        margin: 2px 3px;
        border: 1px solid #0b3d91;
    }
    div[data-testid="stButton"] button[kind="primary"] {
        background-color: #0b3d91 !important;
        color: white !important;
    }
</style>
""", unsafe_allow_html=True)

# ============================================
# FUNCIONES DE GESTIÓN DE ARCHIVOS
# ============================================
def obtener_nombre_archivo_mensual():
    """Genera el nombre del archivo mensual basado en la fecha actual UTC"""
    ahora = datetime.now(timezone.utc)
    return f"SPJC_METAR_{ahora.strftime('%Y_%m')}.xlsx"

def cargar_registros_mes():
    """Carga todos los registros del mes actual desde el archivo Excel"""
    archivo = DIRECTORIO_DATOS / obtener_nombre_archivo_mensual()
    
    if archivo.exists():
        try:
            df = pd.read_excel(archivo, sheet_name='METAR SPJC')
            
            registros = []
            for _, row in df.iterrows():
                registro = row.to_dict()
                
                if 'DIA' in registro:
                    registro['Día'] = str(registro['DIA']).zfill(2)
                if 'HORA' in registro:
                    registro['Hora'] = str(registro['HORA']).zfill(4)
                if 'TIPO' in registro:
                    registro['Tipo'] = registro['TIPO']
                if 'METAR' in registro:
                    registro['METAR_Completo'] = registro['METAR']
                
                registros.append(registro)
            
            return registros
        except Exception as e:
            st.error(f"Error al cargar archivo mensual: {e}")
            return []
    return []

def guardar_registros_mes(registros):
    """Guarda todos los registros en el archivo mensual"""
    if not registros:
        return False, "No hay registros para guardar"
    
    try:
        archivo = DIRECTORIO_DATOS / obtener_nombre_archivo_mensual()
        
        df = pd.DataFrame(registros)
        
        df = df.rename(columns={
            'Día': 'DIA',
            'Hora': 'HORA',
            'Tipo': 'TIPO',
            'Dirección_Viento': 'DIR VIENTO',
            'Intensidad_Viento': 'INTENSIDAD',
            'Variación_Viento': 'VARIACION',
            'Visibilidad_Original': 'VIS (ORIGINAL)',
            'Visibilidad_Metros': 'VIS (CODIGO)',
            'Visibilidad_Mínima': 'VIS MIN',
            'RVR': 'RVR',
            'Fenómeno_Texto': 'FENOMENO',
            'Fenómeno_Código': 'WX',
            'Nubes_Texto': 'NUBOSIDAD',
            'Nubes_Código': 'CLD',
            'Temperatura': 'TEMP °C',
            'Punto_Rocío': 'ROCÍO °C',
            'Humedad_Relativa_%': 'HR %',
            'QNH': 'QNH',
            'Presión_Estación': 'PRESION',
            'Info_Suplementaria': 'RMK',
            'METAR_Completo': 'METAR'
        })
        
        columnas_excel = [
            'DIA', 'HORA', 'TIPO', 'DIR VIENTO', 'INTENSIDAD', 'VARIACION',
            'VIS (ORIGINAL)', 'VIS (CODIGO)', 'VIS MIN', 'RVR',
            'FENOMENO', 'WX', 'NUBOSIDAD', 'CLD',
            'TEMP °C', 'ROCÍO °C', 'HR %', 'QNH', 'PRESION', 'RMK', 'METAR'
        ]
        
        columnas_disponibles = [col for col in columnas_excel if col in df.columns]
        df = df[columnas_disponibles]
        
        df['DIA'] = df['DIA'].astype(str).str.zfill(2)
        df['HORA'] = df['HORA'].astype(str).str.zfill(4)
        df = df.sort_values(['DIA', 'HORA'], ascending=[True, True])
        
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='METAR SPJC', index=False)
            workbook = writer.book
            worksheet = writer.sheets['METAR SPJC']

            for col in range(1, len(columnas_disponibles) + 1):
                column_letter = get_column_letter(col)
                header_cell = worksheet.cell(row=1, column=col)
                max_length = len(str(header_cell.value)) if header_cell.value else 0
                for row in range(2, min(len(df) + 2, 102)):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                worksheet.column_dimensions[column_letter].width = max(min(max_length + 2, 70), 8)

            header_font  = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill  = PatternFill(start_color='0B3D91', end_color='0B3D91', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for col in range(1, len(columnas_disponibles) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font, cell.fill, cell.alignment = header_font, header_fill, header_align
            worksheet.row_dimensions[1].height = 30

            border = Border(
                left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC')
            )
            for row in range(2, len(df) + 2):
                tipo_reporte = worksheet.cell(row=row, column=3).value
                for col in range(1, len(columnas_disponibles) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border    = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if tipo_reporte == 'SPECI':
                        cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                        cell.font = Font(name='Calibri', size=10, bold=True)
                    else:
                        cell.font = Font(name='Calibri', size=10)
            worksheet.freeze_panes = 'A2'

        output.seek(0)
        
        with open(archivo, 'wb') as f:
            f.write(output.getvalue())
        
        return True, f"✅ {len(registros)} registros guardados en {archivo.name}"
        
    except Exception as e:
        return False, f"Error al guardar: {str(e)}"

# ============================================
# INICIALIZAR ESTADO DE SESIÓN
# ============================================
if 'registros' not in st.session_state:
    st.session_state.registros = cargar_registros_mes()

if 'historial' not in st.session_state:
    st.session_state.historial = []

if 'contador' not in st.session_state:
    st.session_state.contador = len(st.session_state.registros)

# ============================================
# FUNCIÓN PARA LIMPIAR / INICIALIZAR CAMPOS
# ============================================
CAMPOS_DEFAULT = {
    'dia_val': lambda: datetime.now(timezone.utc).strftime("%d"),
    'hora_val': '',
    'tipo_val': 'METAR',
    'dir_viento_val': '', 'int_viento_val': '', 'var_viento_val': '',
    'vis_val': '', 'vis_min_val': '', 'rvr_val': '',
    'temp_val': '', 'rocio_val': '', 'hr_val': '', 'qnh_val': '', 'presion_val': '',
    'suplementaria_val': '',
}

def inicializar_campos():
    """Inicializa todos los campos del formulario a sus valores por defecto."""
    for campo, valor in CAMPOS_DEFAULT.items():
        st.session_state[campo] = valor() if callable(valor) else valor
    # Fenómenos y nubes estructurados
    st.session_state.fenomenos_lista = []
    st.session_state.nubes_lista    = []

def limpiar_campos():
    inicializar_campos()

if 'campos_inicializados' not in st.session_state:
    inicializar_campos()
    st.session_state.campos_inicializados = True


# ============================================
# FUNCIÓN PARA ACTUALIZAR O INSERTAR REGISTRO
# ============================================
def actualizar_o_insertar_registro(registros, nuevo_registro):
    dia_nuevo = str(nuevo_registro.get('Día', '')).zfill(2)
    hora_nueva = str(nuevo_registro.get('Hora', '')).zfill(4)
    
    dia_hora_clave = f"{dia_nuevo}_{hora_nueva}"
    accion = "insertado"
    
    for i, registro in enumerate(registros):
        dia_existente = str(registro.get('Día', '')).zfill(2)
        hora_existente = str(registro.get('Hora', '')).zfill(4)
        
        clave_existente = f"{dia_existente}_{hora_existente}"
        
        if clave_existente == dia_hora_clave:
            registros[i] = nuevo_registro
            accion = "actualizado"
            break
    else:
        registros.insert(0, nuevo_registro)
    
    guardar_registros_mes(registros)
    
    return accion

# ============================================
# FUNCIONES DE PROCESAMIENTO - VIENTO
# ============================================
def procesar_viento(direccion, intensidad, variacion):
    """
    PROCESAMIENTO DE VIENTO - REGLAS CORPAC PERÚ
    Caso especial: Dirección 000 e intensidad 00 → 00000KT (ignorar variación)
    """
    dir_int = int(direccion)
    intensidad_str = str(intensidad).upper().strip()
    
    # Caso especial: Viento en calma
    if dir_int == 0 and intensidad_str == "00":
        return "00000KT"
    
    # Procesar ráfagas
    if 'G' in intensidad_str:
        if 'G' in intensidad_str and not ' ' in intensidad_str.replace('G', ''):
            base_int, gust_int = intensidad_str.split('G')
            int_base = int(base_int)
            int_gust = int(gust_int)
            intensidad_metar = f"{int_base:02d}G{int_gust:02d}"
        else:
            parts = intensidad_str.replace('G', ' ').split()
            int_base = int(parts[0])
            int_gust = int(parts[1])
            intensidad_metar = f"{int_base:02d}G{int_gust:02d}"
    else:
        int_base = int(intensidad_str)
        intensidad_metar = f"{int_base:02d}"
    
    # Si NO hay variación
    if not variacion:
        return f"{dir_int:03d}{intensidad_metar}KT"
    
    try:
        variacion = variacion.upper().replace(' ', '')
        if 'V' not in variacion:
            return f"{dir_int:03d}{intensidad_metar}KT"
        
        desde, hasta = map(int, variacion.split('V'))
        
        diff1 = abs(hasta - desde)
        diff2 = 360 - diff1
        
        if desde > hasta:
            diferencia = diff2
        else:
            diferencia = diff1
        
        if diferencia < 60:
            return f"{dir_int:03d}{intensidad_metar}KT"
        
        if diferencia >= 180:
            return f"VRB{intensidad_metar}KT"
        
        if diferencia >= 60:
            if int_base < 3:
                return f"VRB{intensidad_metar}KT"
            else:
                if desde < hasta:
                    return f"{dir_int:03d}{intensidad_metar}KT {desde:03d}V{hasta:03d}"
                else:
                    return f"{dir_int:03d}{intensidad_metar}KT {hasta:03d}V{desde:03d}"
        
        return f"{dir_int:03d}{intensidad_metar}KT"
        
    except Exception as e:
        return f"{dir_int:03d}{intensidad_metar}KT"

# ============================================
# FUNCIONES DE PROCESAMIENTO - VISIBILIDAD
# ============================================
def convertir_visibilidad(vis_texto):
    vis_texto = vis_texto.strip().upper()
    if not vis_texto:
        raise ValueError("Visibilidad es obligatoria")
    
    try:
        if vis_texto.endswith("KM"):
            km = float(vis_texto[:-2])
            return 9999 if km >= 10 else int(km * 1000)
        elif vis_texto.endswith("M"):
            return int(vis_texto[:-1])
        else:
            metros = int(vis_texto)
            return 9999 if metros >= 10000 else metros
    except:
        raise ValueError("Formato de visibilidad inválido")

def procesar_visibilidad_minima(vis_min_texto, vis_m):
    if not vis_min_texto:
        return "", ""
    
    vis_min_texto = vis_min_texto.strip().upper()
    
    if vis_min_texto.endswith('NW'):
        valor = vis_min_texto[:-2]
        cuadrante = 'NW'
    elif vis_min_texto.endswith('NE'):
        valor = vis_min_texto[:-2]
        cuadrante = 'NE'
    elif vis_min_texto.endswith('SW'):
        valor = vis_min_texto[:-2]
        cuadrante = 'SW'
    elif vis_min_texto.endswith('SE'):
        valor = vis_min_texto[:-2]
        cuadrante = 'SE'
    elif vis_min_texto.endswith('N'):
        valor = vis_min_texto[:-1]
        cuadrante = 'N'
    elif vis_min_texto.endswith('S'):
        valor = vis_min_texto[:-1]
        cuadrante = 'S'
    elif vis_min_texto.endswith('E'):
        valor = vis_min_texto[:-1]
        cuadrante = 'E'
    elif vis_min_texto.endswith('W'):
        valor = vis_min_texto[:-1]
        cuadrante = 'W'
    else:
        valor = vis_min_texto
        cuadrante = ''
    
    try:
        if valor.endswith("KM"):
            km = float(valor[:-2])
            vis_min_m = 9999 if km >= 10 else int(km * 1000)
        elif valor.endswith("M"):
            vis_min_m = int(valor[:-1])
        else:
            vis_min_m = int(valor)
            vis_min_m = 9999 if vis_min_m >= 10000 else vis_min_m
        
        es_valida = False
        if vis_min_m < 1500:
            es_valida = True
        if vis_min_m < (vis_m * 0.5) and vis_min_m < 5000:
            es_valida = True
        
        if not es_valida:
            return "", "⚠️ No cumple reglas de visibilidad mínima"
        
        if cuadrante:
            return f"{vis_min_m:04d}{cuadrante}", ""
        else:
            return f"{vis_min_m:04d}", ""
        
    except:
        return "", "❌ Formato inválido"

# ============================================
# FUNCIÓN RVR
# ============================================
def procesar_rvr(rvr_texto):
    if not rvr_texto:
        return ""
    return rvr_texto.strip()

# ============================================
# FUNCIÓN DE FENÓMENOS
# ============================================
def codificar_fenomenos(texto, visibilidad_metros):
    if not texto:
        return ""
    
    # ══════════════════════════════════════════════════════════════════
    # Si el texto contiene códigos METAR directos (ej: "FG BR -RA"),
    # simplemente devolverlos tal cual
    # ══════════════════════════════════════════════════════════════════
    codigos_metar = ["FG", "BR", "RA", "DZ", "SN", "GR", "GS", "TS", 
                     "SH", "FZ", "PRFG", "VCFG", "BCFG", "MIFG",
                     "DU", "SA", "FU", "HZ"]
    palabras = texto.strip().split()
    # Si todas las palabras son códigos conocidos o tienen prefijos de intensidad
    todas_son_codigos = all(
        any(cod in p.upper() for cod in codigos_metar) or 
        p.upper() in ['+', '-', 'M'] or
        p.upper().startswith(('+', '-'))
        for p in palabras if p
    )
    if todas_son_codigos:
        return texto.strip()
    # ══════════════════════════════════════════════════════════════════
    
    texto_lower = texto.lower().strip()
    precipitaciones = []
    oscurecimiento = []
    especiales = []
    
    # Fenómenos especiales de niebla
    if any(x in texto_lower for x in ["niebla parcial", "prfg", "pr fg", "parcial"]):
        especiales.append("PRFG")
        for palabra in ["niebla parcial", "prfg", "pr fg", "parcial"]:
            texto_lower = texto_lower.replace(palabra, "")
    
    if any(x in texto_lower for x in ["niebla en la vecindad", "vcfg", "vc fg", "vecindad"]):
        especiales.append("VCFG")
        for palabra in ["niebla en la vecindad", "vcfg", "vc fg", "vecindad"]:
            texto_lower = texto_lower.replace(palabra, "")
    
    if any(x in texto_lower for x in ["niebla en bancos", "bcfg", "bc fg", "bancos"]):
        especiales.append("BCFG")
        for palabra in ["niebla en bancos", "bcfg", "bc fg", "bancos"]:
            texto_lower = texto_lower.replace(palabra, "")
    
    if any(x in texto_lower for x in ["niebla baja", "mifg", "mi fg", "baja"]):
        especiales.append("MIFG")
        for palabra in ["niebla baja", "mifg", "mi fg", "baja"]:
            texto_lower = texto_lower.replace(palabra, "")
    
    # Mapas de fenómenos
    intensidades = {
        "ligera": "-", "ligero": "-", "leve": "-", "débil": "-", 
        "moderada": "", "moderado": "",
        "fuerte": "+", "intensa": "+", "intenso": "+", "severa": "+"
    }
    
    descriptores = {
        "sh": "SH", "chubasco": "SH", "chubascos": "SH",
        "ts": "TS", "tormenta": "TS", "tormentas": "TS",
        "fz": "FZ", "helada": "FZ", "congelante": "FZ", "congelando": "FZ"
    }
    
    precipitacion_map = {
        "lluvia": "RA", "llovizna": "DZ", "nieve": "SN", 
        "granizo": "GR", "cellisca": "GS"
    }
    
    # Función para determinar código de oscurecimiento según visibilidad
    def get_oscurecimiento_codigo(texto_parte, vis_m):
        if "neblina" in texto_parte:
            if vis_m < 1000:
                return "FG"
            elif vis_m <= 5000:
                return "BR"
            else:
                return None
        elif "niebla" in texto_parte and "parcial" not in texto_parte and "baja" not in texto_parte:
            if vis_m < 1000:
                return "FG"
            else:
                return None
        return None
    
    # Procesar texto
    texto_lower = re.sub(r'\s+y\s+', ', ', texto_lower)
    partes = re.split(r'[,;]', texto_lower)
    
    for parte in partes:
        parte = parte.strip()
        if not parte:
            continue
        
        codigo_base = None
        descriptor = ""
        intensidad = ""
        tipo = None
        parte_procesada = parte
        
        for d_texto, d_codigo in descriptores.items():
            if d_texto in parte_procesada:
                descriptor = d_codigo
                parte_procesada = parte_procesada.replace(d_texto, "").strip()
                break
        
        for i_texto, i_codigo in intensidades.items():
            if i_texto in parte_procesada:
                intensidad = i_codigo
                parte_procesada = parte_procesada.replace(i_texto, "").strip()
                break
        
        for f_texto, f_codigo in precipitacion_map.items():
            if f_texto in parte_procesada:
                codigo_base = f_codigo
                tipo = 'precipitacion'
                break
        
        if not codigo_base:
            codigo_osc = get_oscurecimiento_codigo(parte_procesada, visibilidad_metros)
            if codigo_osc:
                codigo_base = codigo_osc
                tipo = 'oscurecimiento'
        
        if codigo_base:
            codigo_final = codigo_base
            if descriptor and codigo_base not in ["FG", "BR"]:
                codigo_final = descriptor + codigo_final
            if intensidad and codigo_base not in ["FG", "BR"]:
                codigo_final = intensidad + codigo_final
            
            if tipo == 'precipitacion' and codigo_final not in precipitaciones:
                precipitaciones.append(codigo_final)
            elif tipo == 'oscurecimiento' and codigo_final not in oscurecimiento:
                oscurecimiento.append(codigo_final)
    
    resultados = precipitaciones + especiales + oscurecimiento
    return " ".join(resultados[:3]) if resultados else ""

# ============================================
# FUNCIONES DE PROCESAMIENTO - NUBES
# ============================================
def interpretar_nubes(texto, vis_m, fenomeno):
    """
    CODIFICADOR DE NUBES - ESTÁNDAR CORPAC
    Acepta texto libre O lista estructurada [{'octas':int,'tipo':str,'altura_m':int}, ...]
    - Soporta Visibilidad Vertical (VV): metros ÷ 30 = cientos de pies
    """
    # ── Entrada como lista estructurada (nueva UI) ──────────────────────────
    if isinstance(texto, list):
        if not texto:
            if vis_m >= 9999 and not fenomeno.strip():
                return "CAVOK"
            return "NSC"
        codigos = []
        for capa in texto[:4]:
            octas      = int(capa.get('octas', 0))
            tipo_nube  = capa.get('tipo', 'SC').upper()
            altura_m   = int(capa.get('altura_m', 300))
            # Octas → cobertura
            if octas <= 2:   cod_cant = "FEW"
            elif octas <= 4: cod_cant = "SCT"
            elif octas <= 7: cod_cant = "BKN"
            else:            cod_cant = "OVC"
            # Altura metros → cientos de pies (÷30)
            altura_100ft = max(1, min(round(altura_m / 30), 999))
            codigo = f"{cod_cant}{altura_100ft:03d}"
            if tipo_nube in ("CB", "TCU"):
                codigo += tipo_nube
            if codigo not in codigos:
                codigos.append(codigo)
        return " ".join(codigos) if codigos else "NSC"

    # ── Entrada como texto libre (compatibilidad) ────────────────────────────
    if not texto:
        texto = ""
    texto = texto.strip().upper()

    # Visibilidad Vertical
    if any(x in texto for x in ["VIS VER", "VV", "VIS VERT", "VISIBILIDAD VERTICAL"]):
        if "///" in texto or "//" in texto:
            return "VV///"
        numeros = re.findall(r'\d+', texto)
        if numeros:
            altura_cientos = min(max(round(int(numeros[0]) / 30), 0), 999)
            return f"VV{altura_cientos:03d}"

    # CAVOK
    if vis_m >= 9999 and not fenomeno.strip():
        if not texto or texto in ("NSC", "SKC", "CLR", "DESPEJADO", "SIN NUBES", "NO NUBES"):
            return "CAVOK"

    if not texto or texto in ("NSC", "SKC", "CLR", "DESPEJADO", "SIN NUBES", "NO NUBES"):
        return "NSC"

    tipos_nubes = {"CU","SC","ST","CB","TCU","AC","AS","NS","CI","FS"}
    capas = texto.split(",")
    codigos_nubes = []

    for capa in capas[:4]:
        capa = capa.strip()
        if not capa:
            continue
        patron = r'(\d+)\s+([A-Z]{2,4})\s+(\d+)(?:M)?'
        match = re.search(patron, capa)
        if match:
            cantidad   = int(match.group(1))
            tipo       = match.group(2)
            altura     = int(match.group(3))
            tipo_nube  = tipo if tipo in tipos_nubes else tipo
            if altura <= 3000:
                if altura % 30 != 0:
                    altura = (altura // 30) * 30
                altura_cientos = altura // 30
            else:
                if altura % 1000 != 0:
                    altura = (altura // 1000) * 1000
                altura_cientos = (altura // 1000) * 32
            altura_cientos = min(max(altura_cientos, 1), 999)
            if cantidad <= 2:   cod_cant = "FEW"
            elif cantidad <= 4: cod_cant = "SCT"
            elif cantidad <= 7: cod_cant = "BKN"
            else:               cod_cant = "OVC"
            codigo = f"{cod_cant}{altura_cientos:03d}"
            if tipo_nube in ("CB", "TCU"):
                codigo += tipo_nube
            if codigo not in codigos_nubes:
                codigos_nubes.append(codigo)

    return " ".join(codigos_nubes[:4]) if codigos_nubes else "NSC"

# ============================================
# VALIDAR INFO SUPLEMENTARIA (TEMP MIN/MAX + PRECIP)
# ============================================
def validar_info_suplementaria(hora, suplementaria_texto):
    """
    Valida que la info suplementaria contenga los datos obligatorios:
    - 1200Z: debe tener TN (temperatura mínima)
    - 2200Z: debe tener TX (temperatura máxima)
    - Siempre: debe tener PP seguido de números (precipitación)
    
    Retorna: (es_valido, mensaje_error)
    """
    if not suplementaria_texto or not suplementaria_texto.strip():
        return False, "⚠️ Falta información suplementaria obligatoria: precipitación PPxxx"
    
    texto_upper = suplementaria_texto.strip().upper()
    partes = texto_upper.split()
    
    # Verificar precipitación PP seguido de dígitos
    tiene_precip = any(
        p.startswith('PP') and len(p) >= 4 and p[2:5].replace('T','').replace('R','').replace('Z','').isdigit()
        for p in partes
    )
    
    if not tiene_precip:
        return False, "⚠️ Falta precipitación: debe incluir PPxxx (ej: PP000, PP001, PPTRZ)"
    
    # Validar temperatura según hora
    if hora and hora.isdigit() and len(hora) == 4:
        hora_int = int(hora)
        
        if hora_int == 1200:
            tiene_temp_min = any(p.startswith('TN') for p in partes)
            if not tiene_temp_min:
                return False, "⚠️ Observación de las 12Z requiere temperatura mínima (TNxxx)"
        
        elif hora_int == 2200:
            tiene_temp_max = any(p.startswith('TX') for p in partes)
            if not tiene_temp_max:
                return False, "⚠️ Observación de las 22Z requiere temperatura máxima (TXxxx)"
    
    return True, ""


# ============================================
# FUNCIONES DE VALIDACIÓN
# ============================================
def validar_hora(hora_str):
    if not hora_str:
        raise ValueError("Hora requerida - Formato HHMM")
    if len(hora_str) != 4 or not hora_str.isdigit():
        raise ValueError("Hora debe ser HHMM (4 dígitos)")
    h = int(hora_str[:2])
    m = int(hora_str[2:])
    if h > 23 or m > 59:
        raise ValueError("Hora inválida")
    return hora_str

def validar_intensidad_viento(intensidad_str):
    intensidad_str = str(intensidad_str).strip().upper()
    if not intensidad_str:
        raise ValueError("Intensidad de viento requerida")
    
    intensidad_str = intensidad_str.replace(' G ', 'G').replace(' G', 'G').replace('G ', 'G')
    
    if 'G' in intensidad_str:
        partes = intensidad_str.split('G')
        if len(partes) != 2:
            raise ValueError("Formato de ráfagas inválido. Use: 15G25")
        base = int(partes[0])
        rafaga = int(partes[1])
        if base < 0 or base > 100:
            raise ValueError("Intensidad base fuera de rango")
        if rafaga < base:
            raise ValueError("Ráfaga debe ser mayor o igual a intensidad base")
        return intensidad_str
    else:
        intensidad = int(intensidad_str)
        if intensidad < 0 or intensidad > 100:
            raise ValueError("Intensidad fuera de rango")
        return intensidad_str

def validar_numero(valor, min_val, max_val, nombre):
    if not valor:
        raise ValueError(f"{nombre} es obligatorio")
    try:
        num = float(valor)
        if not (min_val <= num <= max_val):
            raise ValueError(f"{nombre} fuera de rango")
        return num
    except:
        raise ValueError(f"{nombre} inválido")

def validar_temp_rocio(temp, rocio):
    if float(rocio) > float(temp):
        raise ValueError(f"Rocío ({rocio}°C) no puede ser > Temperatura ({temp}°C)")
    return True

# ============================================
# FUNCIÓN PRINCIPAL DE GENERACIÓN
# ============================================
def generar_metar(datos):
    try:
        if not datos['dir_viento'] or not datos['int_viento']:
            raise ValueError("Dirección e intensidad del viento son obligatorias")
        if not datos['vis']:
            raise ValueError("Visibilidad es obligatoria")
        if not datos['temp'] or not datos['rocio'] or not datos['qnh']:
            raise ValueError("Temperatura, Rocío y QNH son obligatorios")
        
        hora = validar_hora(datos['hora'])
        int_viento = validar_intensidad_viento(datos['int_viento'])
        viento = procesar_viento(datos['dir_viento'], int_viento, datos['var_viento'])
        vis_m = convertir_visibilidad(datos['vis'])
        
        vis_min_codigo = ""
        if datos['vis_min']:
            vis_min_codigo, vis_min_error = procesar_visibilidad_minima(datos['vis_min'], vis_m)
            if vis_min_error:
                raise ValueError(vis_min_error)
        
        rvr_codigo = procesar_rvr(datos['rvr'])
        
        # Fenómenos: si viene lista de códigos, usarla directamente
        if isinstance(datos['fenomeno'], list):
            fenomeno = " ".join(datos['fenomeno'][:3])  # máximo 3 fenómenos
        else:
            fenomeno = codificar_fenomenos(datos['fenomeno'], vis_m)
        
        nubes    = interpretar_nubes(datos['nubes'], vis_m, fenomeno)
        
        temp = validar_numero(datos['temp'], -10, 40, "Temperatura")
        rocio = validar_numero(datos['rocio'], -10, 40, "Punto de rocío")
        validar_temp_rocio(temp, rocio)
        qnh = validar_numero(datos['qnh'], 850, 1100, "QNH")
        
        temp_metar = redondear_metar(temp)
        rocio_metar = redondear_metar(rocio)
        qnh_metar = int(qnh)
        
        # Validar info suplementaria (precipitación obligatoria + temp según hora)
        es_valida, error_msg = validar_info_suplementaria(hora, datos['suplementaria'])
        if not es_valida:
            raise ValueError(error_msg)
        
        info_sup = datos['suplementaria'].strip().upper() if datos['suplementaria'] else ""
        
        # Construir METAR
        metar_parts = [f"{datos['tipo']} SPJC {datos['dia']}{hora}Z {viento}"]
        
        # Si nubes es CAVOK, usar CAVOK
        if nubes == "CAVOK":
            metar_parts.append("CAVOK")
        else:
            metar_parts.append(f"{vis_m:04d}")
            if vis_min_codigo:
                metar_parts.append(vis_min_codigo)
            if rvr_codigo:
                metar_parts.append(rvr_codigo)
            if fenomeno:
                metar_parts.append(fenomeno)
            metar_parts.append(nubes)
        
        metar_parts.append(f"{temp_metar:02d}/{rocio_metar:02d} Q{qnh_metar}")
        
        if info_sup:
            metar_parts.append(info_sup)
        
        metar_completo = " ".join(metar_parts) + "="
        
        registro = {
            'Día': str(datos['dia']).zfill(2),
            'Hora': hora,
            'Tipo': datos['tipo'],
            'Dirección_Viento': datos['dir_viento'],
            'Intensidad_Viento': datos['int_viento'],
            'Variación_Viento': datos['var_viento'],
            'Visibilidad_Original': datos['vis'],
            'Visibilidad_Metros': vis_m,
            'Visibilidad_Mínima': vis_min_codigo,
            'RVR': rvr_codigo,
            'Fenómeno_Texto': datos['fenomeno'],
            'Fenómeno_Código': fenomeno,
            'Nubes_Texto': datos['nubes'],
            'Nubes_Código': "CAVOK" if nubes == "CAVOK" else nubes,
            'Temperatura': temp,
            'Punto_Rocío': rocio,
            'Humedad_Relativa_%': datos['hr'] if datos['hr'] else "",
            'QNH': qnh,
            'Presión_Estación': datos['presion'],
            'Info_Suplementaria': datos['suplementaria'],
            'METAR_Completo': metar_completo
        }
        
        return {'success': True, 'metar': metar_completo, 'registro': registro}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}

# ============================================
# FUNCIÓN PARA EXPORTAR EXCEL
# ============================================
def exportar_a_excel(registros):
    """Genera un archivo Excel - VERSIÓN CORREGIDA"""
    if not registros:
        return None, "No hay registros para exportar"
    
    try:
        # Crear nuevo DataFrame manualmente
        datos_exportar = []
        for r in registros:
            dia = str(r.get('Día', '')).zfill(2)
            hora = str(r.get('Hora', '')).zfill(4)
            
            datos_exportar.append({
                'DIA': dia,
                'HORA': hora,
                'TIPO': r.get('Tipo', ''),
                'DIR VIENTO': r.get('Dirección_Viento', ''),
                'INTENSIDAD': r.get('Intensidad_Viento', ''),
                'VARIACION': r.get('Variación_Viento', ''),
                'VIS (ORIGINAL)': r.get('Visibilidad_Original', ''),
                'VIS (CODIGO)': r.get('Visibilidad_Metros', ''),
                'VIS MIN': r.get('Visibilidad_Mínima', ''),
                'RVR': r.get('RVR', ''),
                'FENOMENO': r.get('Fenómeno_Texto', ''),
                'WX': r.get('Fenómeno_Código', ''),
                'NUBOSIDAD': r.get('Nubes_Texto', ''),
                'CLD': r.get('Nubes_Código', ''),
                'TEMP °C': r.get('Temperatura', ''),
                'ROCÍO °C': r.get('Punto_Rocío', ''),
                'HR %': r.get('Humedad_Relativa_%', ''),
                'QNH': r.get('QNH', ''),
                'PRESION': r.get('Presión_Estación', ''),
                'RMK': r.get('Info_Suplementaria', ''),
                'METAR': r.get('METAR_Completo', '')
            })
        
        df = pd.DataFrame(datos_exportar)
        df = df.sort_values(['DIA', 'HORA'], ascending=[True, True])
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='METAR SPJC', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['METAR SPJC']
            
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            for col in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col)
                col_name = df.columns[col-1]
                
                max_length = len(str(col_name))
                for row in range(2, min(len(df) + 2, 100)):
                    cell_value = df.iloc[row-2, col-1]
                    if cell_value:
                        cell_length = len(str(cell_value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                if col_name == 'METAR':
                    adjusted_width = min(max_length + 5, 120)
                elif col_name in ['TEMP °C', 'ROCÍO °C', 'HR %', 'QNH']:
                    adjusted_width = 12
                else:
                    adjusted_width = min(max_length + 2, 50)
                
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 8)
            
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='0B3D91', end_color='0B3D91', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            worksheet.row_dimensions[1].height = 30
            
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            for row in range(2, len(df) + 2):
                tipo_reporte = worksheet.cell(row=row, column=3).value
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    
                    if tipo_reporte == 'SPECI':
                        cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                        cell.font = Font(name='Calibri', size=10, bold=True)
                    else:
                        cell.font = Font(name='Calibri', size=10)
                    
                    col_name = df.columns[col-1]
                    if col_name == 'METAR':
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            worksheet.freeze_panes = 'A2'
        
        output.seek(0)
        return output, f"✅ {len(registros)} registros exportados"
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return None, f"Error al exportar: {str(e)}"

# ============================================
# HEADER
# ============================================
col_header1, col_header2 = st.columns([3, 1])

with col_header1:
    st.markdown("<h1 style='color: #0b3d91;'>REGISTRO DE OBSERVACIONES ORDINARIAS Y ESPECIALES</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color: #666;'>Aeropuerto Internacional Jorge Chávez - CORPAC Perú</p>", unsafe_allow_html=True)

with col_header2:
    hoy = datetime.now(timezone.utc).strftime('%d/%m/%Y')
    st.markdown(f"""
    <div style='text-align: right;'>
        <p style='color: #666; font-size: 14px;'>{hoy}</p>
        <p style='color: #0b3d91; font-size: 12px; font-weight: bold;'>{obtener_nombre_archivo_mensual()}</p>
        <p style='color: #999; font-size: 11px;'>CORPAC PERU</p>
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")

# ============================================
# INTERFAZ DE USUARIO
# ============================================
col_izq, col_der = st.columns([2, 1])

with col_izq:
    with st.form(key='metar_form'):
        st.markdown("<div class='section-title'>DATOS DEL REPORTE</div>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            _tipo_idx = ["METAR", "SPECI"].index(st.session_state.get('tipo_val', 'METAR'))
            tipo = st.selectbox("Tipo", ["METAR", "SPECI"], index=_tipo_idx)
        with col2:
            dia = st.text_input("Dia", value=st.session_state.get('dia_val', ''), help="Formato: DD (01-31)")
        with col3:
            hora = st.text_input("Hora UTC", value=st.session_state.get('hora_val', ''), help="Formato HHMM (ej: 1230)")
            
        st.markdown("---")
        
        st.markdown("<div class='section-title'>VIENTO</div>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            dir_viento = st.text_input("Direccion", value=st.session_state.get('dir_viento_val', ''), help="Grados (0-360)")
        with col2:
            int_viento = st.text_input("Intensidad (KT)", value=st.session_state.get('int_viento_val', ''), help="Nudos. Rafagas: 15G25")
        with col3:
            var_viento = st.text_input("Variacion", value=st.session_state.get('var_viento_val', ''), help="Formato: 340V080")
        
        st.markdown("---")
        
        st.markdown("<div class='section-title'>VISIBILIDAD</div>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            vis = st.text_input("Visibilidad", value=st.session_state.get('vis_val', ''), help="Ej: 10km, 5000m, 9999")
        with col2:
            vis_min = st.text_input("Visibilidad Minima", value=st.session_state.get('vis_min_val', ''), help="Ej: 1200SW, 0800NE, 1500SE")
        with col3:
            rvr = st.text_input("RVR", value=st.session_state.get('rvr_val', ''), help="R32/0400, R12R/1700, R10/M0050")
        
        st.markdown("---")
        
        st.markdown("<div class='section-title'>🌦️ FENÓMENOS</div>", unsafe_allow_html=True)

        FENOMENOS_OPCIONES = {
            "Lluvia ligera (-RA)": "-RA",
            "Lluvia moderada (RA)": "RA",
            "Lluvia fuerte (+RA)": "+RA",
            "Llovizna ligera (-DZ)": "-DZ",
            "Llovizna moderada (DZ)": "DZ",
            "Lluvia con tormenta (TSRA)": "TSRA",
            "Chubasco de lluvia (SHRA)": "SHRA",
            "Nieve ligera (-SN)": "-SN",
            "Nieve moderada (SN)": "SN",
            "Granizo (GR)": "GR",
            "Niebla (FG)": "FG",
            "Neblina (BR)": "BR",
            "Niebla parcial (PRFG)": "PRFG",
            "Niebla en bancos (BCFG)": "BCFG",
            "Niebla baja (MIFG)": "MIFG",
            "Niebla en vecindad (VCFG)": "VCFG",
            "Tormenta eléctrica (TS)": "TS",
            "Polvo en suspensión (DU)": "DU",
            "Arena (SA)": "SA",
            "Humo (FU)": "FU",
        }

        # Mostrar fenómenos ya agregados
        fenomeno_a_eliminar = None
        for idx, codigo_fx in enumerate(st.session_state.fenomenos_lista):
            c1, c2 = st.columns([5, 1])
            with c1:
                st.markdown(
                    f"<div style='background:#e8f0fe; border:1px solid #0b3d91; border-radius:8px; "
                    f"padding:6px 12px; margin:4px 0; display:inline-block;'>"
                    f"<span style='color:#0b3d91; font-weight:600; font-size:0.9rem;'>{codigo_fx}</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )
            with c2:
                if st.form_submit_button("✕", key=f"del_fx_{idx}"):
                    fenomeno_a_eliminar = idx
        
        if fenomeno_a_eliminar is not None:
            st.session_state.fenomenos_lista.pop(fenomeno_a_eliminar)

        # Selector para agregar nuevo fenómeno
        if len(st.session_state.fenomenos_lista) < 3:
            col_fx1, col_fx2 = st.columns([4, 1])
            with col_fx1:
                sel_fenomeno = st.selectbox(
                    "Agregar fenómeno",
                    list(FENOMENOS_OPCIONES.keys()),
                    key='sel_fenomeno_dd',
                    label_visibility="collapsed"
                )
            with col_fx2:
                if st.form_submit_button("➕", key='btn_agregar_fenomeno', use_container_width=True):
                    codigo_fx = FENOMENOS_OPCIONES.get(sel_fenomeno)
                    if codigo_fx and codigo_fx not in st.session_state.fenomenos_lista:
                        st.session_state.fenomenos_lista.append(codigo_fx)
        else:
            st.caption("Máximo 3 fenómenos")
        
        if not st.session_state.fenomenos_lista:
            st.caption("Sin fenómenos — NSC o CAVOK si aplica")

        st.markdown("---")

        # ── NUBES ESTRUCTURADAS ───────────────────────────────────────────
        st.markdown("<div class='section-title'>☁️ NUBOSIDAD</div>", unsafe_allow_html=True)

        TIPOS_NUBE = ["ST", "SC", "CU", "TCU", "CB", "AC", "AS", "NS", "CI", "FS"]
        OCTAS_LABELS = {
            1: "1 octa (FEW)", 2: "2 octas (FEW)",
            3: "3 octas (SCT)", 4: "4 octas (SCT)",
            5: "5 octas (BKN)", 6: "6 octas (BKN)", 7: "7 octas (BKN)",
            8: "8 octas (OVC)"
        }

        # Mostrar capas existentes
        capas_a_eliminar = None
        for idx, capa in enumerate(st.session_state.nubes_lista):
            octas_v   = capa['octas']
            tipo_v    = capa['tipo']
            altura_v  = capa['altura_m']
            cod_cob   = "FEW" if octas_v<=2 else "SCT" if octas_v<=4 else "BKN" if octas_v<=7 else "OVC"
            altura_ft = round(altura_v / 30)
            preview   = f"{cod_cob}{altura_ft:03d}" + (tipo_v if tipo_v in ('CB','TCU') else '')
            st.markdown(f"<div class='capa-nube-box'>", unsafe_allow_html=True)
            c1, c2 = st.columns([4, 1])
            with c1:
                st.markdown(
                    f"**Capa {idx+1}:** {octas_v} octas · {tipo_v} · {altura_v} m "
                    f"→ <code>{preview}</code>",
                    unsafe_allow_html=True
                )
            with c2:
                if st.form_submit_button(f"✕", key=f"del_capa_{idx}"):
                    capas_a_eliminar = idx
            st.markdown("</div>", unsafe_allow_html=True)

        if capas_a_eliminar is not None:
            st.session_state.nubes_lista.pop(capas_a_eliminar)

        # Formulario para nueva capa
        if len(st.session_state.nubes_lista) < 4:
            st.markdown("**Nueva capa:**")
            cn1, cn2, cn3, cn4 = st.columns([2, 2, 2, 1])
            with cn1:
                nueva_octas = st.selectbox(
                    "Octas", list(OCTAS_LABELS.keys()),
                    format_func=lambda x: OCTAS_LABELS[x],
                    key='nueva_octas'
                )
            with cn2:
                nuevo_tipo = st.selectbox("Tipo nube", TIPOS_NUBE, key='nuevo_tipo_nube')
            with cn3:
                nueva_altura = st.number_input(
                    "Altura (m)", min_value=30, max_value=15000,
                    value=300, step=30, key='nueva_altura_m'
                )
            with cn4:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.form_submit_button("➕", key='btn_agregar_capa_nube', use_container_width=True):
                    st.session_state.nubes_lista.append({
                        'octas': nueva_octas,
                        'tipo': nuevo_tipo,
                        'altura_m': nueva_altura
                    })
        else:
            st.caption("Máximo 4 capas de nubes")

        st.markdown("---")
        
        st.markdown("<div class='section-title'>TEMPERATURA Y PRESION</div>", unsafe_allow_html=True)
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            temp = st.text_input("Temp °C", value=st.session_state.get('temp_val', ''), help="-10 a 40°C")
        with col2:
            rocio = st.text_input("Rocío °C", value=st.session_state.get('rocio_val', ''), help="≤ Temperatura")
        with col3:
            hr = st.text_input("HR %", value=st.session_state.get('hr_val', ''), help="0-100%")
        with col4:
            qnh = st.text_input("QNH hPa", value=st.session_state.get('qnh_val', ''), help="850-1100 hPa")
        with col5:
            presion = st.text_input("Presión Est.", value=st.session_state.get('presion_val', ''), help="Opcional")
        
        st.markdown("---")
        
        st.markdown("<div class='section-title'>INFORMACION SUPLEMENTARIA</div>", unsafe_allow_html=True)
        st.caption("⚠️ **Obligatorio:** Precipitación PPxxx (ej: PP000, PP005, PPTRZ)")
        st.caption("📋 12Z requiere TN (temp mín) · 22Z requiere TX (temp máx)")
        suplementaria = st.text_input(
            "Info suplementaria", 
            value=st.session_state.get('suplementaria_val', ''),
            placeholder="NOSIG RMK PP000",
            help="Siempre incluir PPxxx. A las 12Z agregar TNxxx, a las 22Z agregar TXxxx",
            label_visibility="collapsed"
        )
        
        st.markdown("---")
        
        generar = st.form_submit_button("✅ GENERAR METAR / SPECI", use_container_width=True, type="primary")
        
        if generar:
            datos = {
                'tipo': tipo, 'dia': dia, 'hora': hora,
                'dir_viento': dir_viento, 'int_viento': int_viento, 'var_viento': var_viento,
                'vis': vis, 'vis_min': vis_min, 'rvr': rvr,
                'fenomeno': st.session_state.fenomenos_lista,  # ← lista directa
                'nubes': st.session_state.nubes_lista,
                'temp': temp, 'rocio': rocio, 'hr': hr,
                'qnh': qnh, 'presion': presion, 'suplementaria': suplementaria
            }
            
            resultado = generar_metar(datos)
            
            if resultado['success']:
                # Textos legibles para el registro Excel
                fenomeno_texto_legible = " ".join(st.session_state.fenomenos_lista)
                nubes_texto_legible = ", ".join(
                    f"{c['octas']} {c['tipo']} {c['altura_m']}M"
                    for c in st.session_state.nubes_lista
                ) if st.session_state.nubes_lista else ""
                resultado['registro']['Fenómeno_Texto'] = fenomeno_texto_legible
                resultado['registro']['Nubes_Texto'] = nubes_texto_legible
                accion = actualizar_o_insertar_registro(st.session_state.registros, resultado['registro'])
                
                dia_hora_clave = f"{str(resultado['registro']['Día']).zfill(2)}_{resultado['registro']['Hora']}"
                nuevo_historial = []
                
                for metar in st.session_state.historial:
                    match = re.search(r'SPJC (\d{2})(\d{4})Z', metar)
                    if match:
                        dia_hist = match.group(1)
                        hora_hist = match.group(2)
                        if f"{dia_hist}_{hora_hist}" != dia_hora_clave:
                            nuevo_historial.append(metar)
                    else:
                        nuevo_historial.append(metar)
                
                nuevo_historial.insert(0, resultado['metar'])
                st.session_state.historial = nuevo_historial[:20]
                st.session_state.contador = len(st.session_state.registros)
                
                if accion == "actualizado":
                    st.warning("METAR ACTUALIZADO - Reemplazo reporte existente")
                else:
                    st.success("METAR generado correctamente")
                
                st.session_state.ultimo_metar = resultado['metar']
                st.session_state.ultimo_tipo = tipo
                st.session_state.ultimo_registro = resultado['registro']
                
                # Limpiar todos los campos después de generar
                limpiar_campos()
                st.rerun()
            else:
                st.error(f"ERROR: {resultado['error']}")

with col_der:
    st.markdown("<div class='section-title'>📋 ÚLTIMO REPORTE</div>", unsafe_allow_html=True)
    if 'ultimo_metar' in st.session_state:
        tipo_ultimo = st.session_state.get('ultimo_tipo', 'METAR')
        if tipo_ultimo == "SPECI":
            st.markdown(f"<div style='background: #FFE699; padding: 15px; border-radius: 5px; font-family: monospace; border-left: 5px solid #FFC000;'>{st.session_state.ultimo_metar}</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div class='metar-box'>{st.session_state.ultimo_metar}</div>", unsafe_allow_html=True)
    else:
        st.info("---")
    
    st.markdown("---")
    st.markdown("<div class='section-title'>EXPORTAR</div>", unsafe_allow_html=True)
    
    if st.button("📥 Exportar METAR", use_container_width=True, type="primary"):
        if st.session_state.registros:
            excel_file, mensaje = exportar_a_excel(st.session_state.registros)
            if excel_file:
                nombre_archivo = obtener_nombre_archivo_mensual()
                st.download_button(
                    label="✅ Confirmar Descarga",
                    data=excel_file,
                    file_name=nombre_archivo,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.success(mensaje)
            else:
                st.warning(mensaje)
        else:
            st.warning("No hay registros")
    
    if st.button("🗑️ Limpiar Memoria", use_container_width=True):
        st.session_state.registros = []
        st.session_state.historial = []
        st.session_state.contador = 0
        st.success("Memoria limpiada")
    
    st.markdown("---")
    st.metric("REGISTROS EN MEMORIA", st.session_state.contador)
    
    st.markdown("---")
    st.markdown("<div class='section-title'>HISTORIAL</div>", unsafe_allow_html=True)
    if st.session_state.historial:
        for metar in st.session_state.historial[:10]:
            if "SPECI" in metar:
                st.markdown(f"<div style='background: #FFE699; padding: 8px; margin-bottom: 5px; border-radius: 3px; font-family: monospace; font-size: 12px; border-left: 3px solid #FFC000;'>{metar}</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='background: #f0f0f0; padding: 8px; margin-bottom: 5px; border-radius: 3px; font-family: monospace; font-size: 12px; border-left: 3px solid #0b3d91;'>{metar}</div>", unsafe_allow_html=True)
    else:
        st.info("No hay METARs en el historial")

# ============================================
# FOOTER
# ============================================
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666; padding: 20px;'>
    <p>METAR Digital - CORPAC Peru | Aeropuerto Internacional Jorge Chavez (SPJC)</p>
    <p style='font-size: 0.8rem;'>Sistema de Registro de Observaciones - Versión 11.0</p>
</div>
""", unsafe_allow_html=True)

# ============================================
# GUARDADO AUTOMATICO
# ============================================
if st.session_state.registros:
    guardar_registros_mes(st.session_state.registros)
