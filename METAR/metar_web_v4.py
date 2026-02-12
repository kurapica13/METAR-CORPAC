"""
METAR DIGITAL WEB - V CORPAC PERÚ
Características:
✅ SIN DUPLICADOS: Reemplaza reportes con misma fecha/hora
✅ VIENTO: Reglas circulares (340V080) corregidas
✅ VISIBILIDAD MÍNIMA: Cuadrantes SE, NW, etc. CORREGIDO
✅ RVR: Runway Visual Range
✅ NUBES: Estándar oficial CORPAC (30m/1000m)
✅ EXCEL: Encabezados personalizados en español CORREGIDO
"""

import streamlit as st
from datetime import datetime, timezone
import pandas as pd
from pathlib import Path
import re
import time
import os
import base64
from io import BytesIO

# ============================================
# CONFIGURACIÓN DE PÁGINA - VERSIÓN 3.0
# ============================================
st.set_page_config(
    page_title="METAR Digital - CORPAC Perú",
    page_icon="🔄",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Limpiar cache al inicio
st.cache_data.clear()

# ============================================
# ESTILOS CSS PERSONALIZADOS
# ============================================
st.markdown("""
<style>
    .stApp {
        background-color: #f0f8ff;
    }
    
    .metar-header {
        background: linear-gradient(90deg, #0b3d91 0%, #1a4fa0 100%);
        padding: 1.5rem;
        border-radius: 10px;
        color: white;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    
    .panel {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        margin-bottom: 1rem;
    }
    
    .section-title {
        color: #0b3d91;
        font-weight: 600;
        margin-bottom: 1rem;
        border-bottom: 2px solid #e8eef7;
        padding-bottom: 0.5rem;
    }
    
    .metar-box {
        background: #1e1e1e;
        color: #00ff00;
        padding: 1rem;
        border-radius: 5px;
        font-family: 'Courier New', monospace;
        font-size: 1.1rem;
        border-left: 5px solid #0b3d91;
    }
    
    .historial-item {
        background: #f8f9fa;
        padding: 0.8rem;
        border-radius: 5px;
        margin-bottom: 0.5rem;
        font-family: 'Courier New', monospace;
        border-left: 3px solid #0b3d91;
    }
    
    .historial-item-speci {
        background: #FFE699;
        border-left: 3px solid #FFC000;
    }
    
    .badge-v3 {
        background: #27ae60;
        color: white;
        padding: 4px 12px;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: bold;
        display: inline-block;
        margin-left: 10px;
    }
    
    .stButton button {
        width: 100%;
        border-radius: 5px;
        font-weight: 600;
    }
    
    .stTextInput input, .stSelectbox select {
        border-radius: 5px;
        border: 1px solid #ddd;
    }
</style>
""", unsafe_allow_html=True)

# ============================================
# INICIALIZAR ESTADO DE SESIÓN
# ============================================
if 'registros_v3' not in st.session_state:
    st.session_state.registros_v3 = []
if 'historial_v3' not in st.session_state:
    st.session_state.historial_v3 = []
if 'contador_v3' not in st.session_state:
    st.session_state.contador_v3 = 0
if 'campos_inicializados_v3' not in st.session_state:
    st.session_state.campos_inicializados_v3 = False

# ============================================
# FUNCIÓN PARA LIMPIAR CAMPOS
# ============================================
def limpiar_campos_v3():
    """Limpia todos los campos del formulario"""
    st.session_state.dia_v3 = datetime.now(timezone.utc).strftime("%d")
    st.session_state.hora_v3 = datetime.now(timezone.utc).strftime("%H%M")
    st.session_state.tipo_v3 = "METAR"
    st.session_state.dir_viento_v3 = ""
    st.session_state.int_viento_v3 = ""
    st.session_state.var_viento_v3 = ""
    st.session_state.vis_v3 = ""
    st.session_state.vis_min_v3 = ""
    st.session_state.rvr_v3 = ""
    st.session_state.fenomeno_v3 = ""
    st.session_state.nubes_v3 = ""
    st.session_state.temp_v3 = ""
    st.session_state.rocio_v3 = ""
    st.session_state.hr_v3 = ""
    st.session_state.qnh_v3 = ""
    st.session_state.presion_v3 = ""
    st.session_state.suplementaria_v3 = ""
    st.session_state.campos_inicializados_v3 = True

# ============================================
# INICIALIZAR CAMPOS VACÍOS
# ============================================
if not st.session_state.campos_inicializados_v3:
    limpiar_campos_v3()

# ============================================
# FUNCIÓN PARA ACTUALIZAR O INSERTAR REGISTRO (EVITAR DUPLICADOS)
# ============================================
def actualizar_o_insertar_registro_v3(registros, nuevo_registro):
    """Busca un registro existente con el mismo día y hora. Si existe, lo reemplaza."""
    dia_hora_clave = f"{nuevo_registro['Día']}_{nuevo_registro['Hora']}"
    
    # Buscar índice del registro con mismo día y hora
    indice_existente = -1
    for i, registro in enumerate(registros):
        clave_existente = f"{registro.get('Día', '')}_{registro.get('Hora', '')}"
        if clave_existente == dia_hora_clave:
            indice_existente = i
            break
    
    # Si existe, reemplazarlo; si no, insertar al inicio
    if indice_existente != -1:
        registros[indice_existente] = nuevo_registro
        return "actualizado"
    else:
        registros.insert(0, nuevo_registro)
        return "insertado"

# ============================================
# FUNCIONES DE PROCESAMIENTO - VIENTO
# ============================================
def procesar_viento(direccion, intensidad, variacion):
    """
    PROCESAMIENTO DE VIENTO - REGLAS CORPAC PERÚ
    Caso 1: Variación ≥60° y <180° con viento <3kt → VRBxxKT
    Caso 2: Variación ≥60° y <180° con viento ≥3kt → dddffKT bbbVnnn
    Caso 3: Variación ≥180° → VRBxxKT (sin importar velocidad)
    """
    dir_int = int(direccion)
    intensidad_str = str(intensidad).upper().strip()
    
    # Procesar ráfagas
    if 'G' in intensidad_str:
        if 'G' in intensidad_str and not ' ' in intensidad_str.replace('G', ''):
            base_int, gust_int = intensidad_str.split('G')
            int_base = int(base_int)
            int_gust = int(gust_int)
            intensidad_metar = f"{int_base:02d}G{int_gust:02d}"
        else:
            parts = intensidad_str.replace('G', ' ').split()
            int_base = int(parts[0])
            int_gust = int(parts[1])
            intensidad_metar = f"{int_base:02d}G{int_gust:02d}"
    else:
        int_base = int(intensidad_str)
        intensidad_metar = f"{int_base:02d}"
    
    # Si NO hay variación
    if not variacion:
        return f"{dir_int:03d}{intensidad_metar}KT"
    
    try:
        # Extraer valores de variación (formato: bbbVnnn)
        variacion = variacion.upper().replace(' ', '')
        if 'V' not in variacion:
            return f"{dir_int:03d}{intensidad_metar}KT"
        
        desde, hasta = map(int, variacion.split('V'))
        
        # Calcular diferencia CIRCULAR (importante para 340V080)
        diff1 = abs(hasta - desde)
        diff2 = 360 - diff1
        diferencia = min(diff1, diff2)
        
        # CASO 3: Variación ≥ 180°
        if diferencia >= 180:
            return f"VRB{intensidad_metar}KT"
        
        # CASO 1 y 2: Variación ≥ 60° y < 180°
        if diferencia >= 60:
            if int_base < 3:
                # CASO 1: Viento < 3kt
                return f"VRB{intensidad_metar}KT"
            else:
                # CASO 2: Viento ≥ 3kt
                if diff1 <= 180:
                    return f"{dir_int:03d}{intensidad_metar}KT {desde:03d}V{hasta:03d}"
                else:
                    return f"{dir_int:03d}{intensidad_metar}KT {hasta:03d}V{desde:03d}"
        
        # Variación < 60°
        return f"{dir_int:03d}{intensidad_metar}KT {variacion}"
        
    except Exception as e:
        return f"{dir_int:03d}{intensidad_metar}KT"

# ============================================
# FUNCIONES DE PROCESAMIENTO - VISIBILIDAD
# ============================================
def convertir_visibilidad(vis_texto):
    """Convierte visibilidad a metros"""
    vis_texto = vis_texto.strip().upper()
    if not vis_texto:
        raise ValueError("Visibilidad es obligatoria")
    
    try:
        if vis_texto.endswith("KM"):
            km = float(vis_texto[:-2])
            return 9999 if km >= 10 else int(km * 1000)
        elif vis_texto.endswith("M"):
            return int(vis_texto[:-1])
        else:
            metros = int(vis_texto)
            return 9999 if metros >= 10000 else metros
    except:
        raise ValueError("Formato de visibilidad inválido")

# ============================================
# FUNCIÓN CORREGIDA PARA VISIBILIDAD MÍNIMA - ¡ESTA ES LA BUENA!
# ============================================
def procesar_visibilidad_minima_v3(vis_min_texto, vis_m):
    """
    Procesa visibilidad mínima con cuadrantes - VERSIÓN CORREGIDA
    Cuadrantes: N, NE, E, SE, S, SW, W, NW
    Reglas: 1) <1500m 2) <50% vis reinante y <5000m
    """
    if not vis_min_texto:
        return "", ""
    
    vis_min_texto = vis_min_texto.strip().upper()
    # ORDEN CRÍTICO: primero cuadrantes de 2 letras, luego de 1 letra
    cuadrantes = ['NW', 'NE', 'SW', 'SE', 'N', 'S', 'E', 'W']
    
    valor = ""
    cuadrante = ""
    
    # Buscar cuadrante (prioridad a los de 2 letras)
    for c in cuadrantes:
        if vis_min_texto.endswith(c):
            valor = vis_min_texto[:-len(c)]
            cuadrante = c
            break
    
    if not cuadrante:
        valor = vis_min_texto
    
    try:
        # Convertir a metros
        if valor.endswith("KM"):
            km = float(valor[:-2])
            vis_min_m = 9999 if km >= 10 else int(km * 1000)
        elif valor.endswith("M"):
            vis_min_m = int(valor[:-1])
        else:
            vis_min_m = int(valor)
            vis_min_m = 9999 if vis_min_m >= 10000 else vis_min_m
        
        # Validar según reglas CORPAC
        es_valida = False
        if vis_min_m < 1500:
            es_valida = True
        if vis_min_m < (vis_m * 0.5) and vis_min_m < 5000:
            es_valida = True
        
        if not es_valida:
            return "", "⚠️ No cumple reglas de visibilidad mínima"
        
        # Formatear salida
        if cuadrante:
            return f"{vis_min_m:04d}{cuadrante}", ""
        else:
            return f"{vis_min_m:04d}", ""
        
    except ValueError:
        return "", "❌ Formato inválido - Use números + N, NE, E, SE, S, SW, W, NW"
    except:
        return "", "❌ Formato inválido"

def procesar_rvr(rvr_texto):
    """Procesa RVR (Runway Visual Range)"""
    if not rvr_texto:
        return ""
    
    rvr_texto = rvr_texto.strip().upper().replace('M', '').replace('RVR', '')
    
    try:
        rvr_valor = int(rvr_texto)
        if rvr_valor < 50 or rvr_valor > 2000:
            return ""
        return f"RVR{rvr_valor:04d}"
    except:
        return ""

# ============================================
# FUNCIONES DE PROCESAMIENTO - FENÓMENOS
# ============================================
def codificar_fenomenos(texto):
    """CODIFICADOR COMPLETO DE FENÓMENOS - CORPAC"""
    if not texto:
        return ""
    
    texto_lower = texto.lower().strip()
    
    if any(x in texto_lower for x in ["niebla parcial", "prfg", "pr fg", "parcial"]):
        return "PRFG"
    if any(x in texto_lower for x in ["niebla en la vecindad", "vcfg", "vc fg", "vecindad"]):
        return "VCFG"
    if any(x in texto_lower for x in ["niebla en bancos", "bcfg", "bc fg", "bancos"]):
        return "BCFG"
    if any(x in texto_lower for x in ["niebla baja", "mifg", "mi fg", "baja"]):
        return "MIFG"
    
    intensidades = {"ligera": "-", "ligero": "-", "leve": "-", "débil": "-",
                   "fuerte": "+", "intensa": "+", "intenso": "+", "severa": "+"}
    
    descriptores = {"sh": "SH", "chubasco": "SH", "ts": "TS", "tormenta": "TS",
                   "fz": "FZ", "helada": "FZ", "congelante": "FZ"}
    
    fenomenos = {
        "lluvia": "RA", "llovizna": "DZ", "niebla": "FG", "neblina": "BR",
        "nieve": "SN", "granizo": "GR", "cellisca": "GS", "tormenta": "TS",
        "polvo": "DU", "arena": "SA", "humo": "FU", "ceniza": "VA", "calima": "HZ"
    }
    
    partes = [p.strip() for p in texto.split(",")][:3]
    resultados = []
    
    for parte in partes:
        if not parte:
            continue
        parte_lower = parte.lower()
        codigo = ""
        descriptor = ""
        intensidad = ""
        
        for d_texto, d_codigo in descriptores.items():
            if d_texto in parte_lower:
                descriptor = d_codigo
                parte_lower = parte_lower.replace(d_texto, "").strip()
                break
        
        for i_texto, i_codigo in intensidades.items():
            if i_texto in parte_lower:
                intensidad = i_codigo
                parte_lower = parte_lower.replace(i_texto, "").strip()
                break
        
        for f_texto, f_codigo in fenomenos.items():
            if f_texto in parte_lower:
                codigo = f_codigo
                break
        
        if codigo:
            if descriptor:
                codigo = descriptor + codigo
            if intensidad:
                codigo = intensidad + codigo
            resultados.append(codigo)
    
    return " ".join(resultados) if resultados else ""

# ============================================
# FUNCIONES DE PROCESAMIENTO - NUBES
# ============================================
def interpretar_nubes(texto, vis_m, fenomeno):
    """CODIFICADOR DE NUBES - ESTÁNDAR CORPAC PERÚ"""
    texto = texto.strip().upper()
    
    if texto in ["DESPEJADO", "SKC", "CLR", "", "NSC", "SIN NUBES", "NO NUBES"]:
        return "NSC"
    
    if vis_m >= 9999 and not fenomeno.strip() and texto in ["NSC", "SKC", "CLR", "DESPEJADO"]:
        return "CAVOK"
    
    tipos_nubes = {
        "CU": "CU", "SC": "SC", "ST": "ST", "CB": "CB", "TCU": "TCU",
        "AC": "AC", "AS": "AS", "NS": "NS", "CI": "CI"
    }
    
    capas = texto.split(",")
    codigos_nubes = []
    
    for capa in capas[:4]:
        capa = capa.strip()
        if not capa:
            continue
        
        patron = r'(\d+)\s+([A-Z]{2,4})\s+(\d+)(?:M)?'
        match = re.search(patron, capa)
        
        if match:
            cantidad = int(match.group(1))
            tipo = match.group(2)
            altura = int(match.group(3))
            
            tipo_nube = tipos_nubes.get(tipo, tipo)
            
            if altura <= 3000:
                if altura % 30 != 0:
                    altura = (altura // 30) * 30
                altura_cientos = altura // 30
            else:
                if altura % 1000 != 0:
                    altura = (altura // 1000) * 1000
                altura_cientos = (altura // 1000) * 32
            
            altura_cientos = min(max(altura_cientos, 1), 999)
            
            if cantidad <= 2:
                cod_cant = "FEW"
            elif cantidad <= 4:
                cod_cant = "SCT"
            elif cantidad <= 7:
                cod_cant = "BKN"
            else:
                cod_cant = "OVC"
            
            codigo = f"{cod_cant}{altura_cientos:03d}"
            if tipo_nube in ["CB", "TCU"]:
                codigo += tipo_nube
            codigos_nubes.append(codigo)
        
        else:
            cb_match = re.search(r'CB\s+(\d+)(?:M)?', capa)
            if cb_match:
                altura = int(cb_match.group(1))
                if altura <= 3000:
                    altura = (altura // 30) * 30
                    codigos_nubes.append(f"BKN{(altura//30):03d}CB")
                else:
                    altura = (altura // 1000) * 1000
                    codigos_nubes.append(f"BKN{(altura//1000)*32:03d}CB")
            
            tcu_match = re.search(r'TCU\s+(\d+)(?:M)?', capa)
            if tcu_match:
                altura = int(tcu_match.group(1))
                if altura <= 3000:
                    altura = (altura // 30) * 30
                    codigos_nubes.append(f"BKN{(altura//30):03d}TCU")
                else:
                    altura = (altura // 1000) * 1000
                    codigos_nubes.append(f"BKN{(altura//1000)*32:03d}TCU")
    
    return " ".join(codigos_nubes[:4]) if codigos_nubes else "NSC"

def verificar_cavok(vis_m, fenomeno, nubes):
    """Verifica condiciones para CAVOK"""
    return (vis_m >= 9999 and not fenomeno.strip() and nubes in ["NSC", "SKC", "CLR"])

# ============================================
# FUNCIONES DE VALIDACIÓN
# ============================================
def validar_hora(hora_str):
    """Valida formato de hora HHMM"""
    if len(hora_str) != 4 or not hora_str.isdigit():
        raise ValueError("Hora debe ser HHMM (4 dígitos)")
    h = int(hora_str[:2])
    m = int(hora_str[2:])
    if h > 23 or m > 59:
        raise ValueError("Hora inválida")
    return hora_str

def validar_intensidad_viento(intensidad_str):
    """Valida formato de intensidad de viento"""
    intensidad_str = str(intensidad_str).strip().upper()
    if not intensidad_str:
        raise ValueError("Intensidad de viento requerida")
    
    intensidad_str = intensidad_str.replace(' G ', 'G').replace(' G', 'G').replace('G ', 'G')
    
    if 'G' in intensidad_str:
        partes = intensidad_str.split('G')
        if len(partes) != 2:
            raise ValueError("Formato de ráfagas inválido. Use: 15G25")
        base = int(partes[0])
        rafaga = int(partes[1])
        if base < 0 or base > 100:
            raise ValueError("Intensidad base fuera de rango (0-100)")
        if rafaga < base:
            raise ValueError("Ráfaga debe ser mayor o igual a intensidad base")
        if rafaga > 150:
            raise ValueError("Ráfaga excede límite (150 KT)")
        diferencia = rafaga - base
        if diferencia < 10:
            raise ValueError(f"Ráfaga requiere diferencia ≥10 KT (actual: {diferencia} KT)")
        return intensidad_str
    else:
        intensidad = int(intensidad_str)
        if intensidad < 0 or intensidad > 100:
            raise ValueError("Intensidad fuera de rango (0-100)")
        return intensidad_str

def validar_numero(valor, min_val, max_val, nombre):
    """Valida un número dentro de un rango"""
    if not valor:
        raise ValueError(f"{nombre} es obligatorio")
    try:
        num = float(valor)
        if not (min_val <= num <= max_val):
            raise ValueError(f"{nombre} fuera de rango ({min_val}-{max_val})")
        return num
    except ValueError as e:
        raise ValueError(f"{nombre} inválido: {str(e)}")

def validar_temp_rocio(temp, rocio):
    """Valida que rocío ≤ temperatura"""
    if float(rocio) > float(temp):
        raise ValueError(f"Rocío ({rocio}°C) no puede ser > Temperatura ({temp}°C)")
    return True

def validar_humedad(hr):
    """Valida humedad relativa"""
    if not hr:
        return ""
    try:
        num = float(hr)
        if num < 0 or num > 100:
            raise ValueError("Humedad fuera de rango (0-100%)")
        return num
    except:
        raise ValueError("Humedad inválida")

# ============================================
# FUNCIÓN PRINCIPAL DE GENERACIÓN - CORREGIDA
# ============================================
def generar_metar_v3(datos):
    """Genera código METAR desde los datos del formulario - VERSIÓN CORREGIDA"""
    try:
        # Validar campos obligatorios
        if not datos['dir_viento'] or not datos['int_viento']:
            raise ValueError("Dirección e intensidad del viento son obligatorias")
        if not datos['vis']:
            raise ValueError("Visibilidad es obligatoria")
        if not datos['temp'] or not datos['rocio'] or not datos['qnh']:
            raise ValueError("Temperatura, Rocío y QNH son obligatorios")
        
        # Validar hora
        hora = validar_hora(datos['hora'])
        
        # Procesar viento
        int_viento = validar_intensidad_viento(datos['int_viento'])
        viento = procesar_viento(datos['dir_viento'], int_viento, datos['var_viento'])
        
        # Procesar visibilidad
        vis_m = convertir_visibilidad(datos['vis'])
        
        # ===== IMPORTANTE: USAR LA FUNCIÓN CORREGIDA =====
        # Procesar visibilidad mínima con cuadrante - VERSIÓN V3 CORREGIDA
        vis_min_codigo = ""
        if datos['vis_min']:
            vis_min_codigo, vis_min_error = procesar_visibilidad_minima_v3(datos['vis_min'], vis_m)
            if vis_min_error:
                raise ValueError(vis_min_error)
        
        # Procesar RVR
        rvr_codigo = procesar_rvr(datos['rvr'])
        
        # Procesar fenómenos y nubes
        fenomeno = codificar_fenomenos(datos['fenomeno'])
        nubes = interpretar_nubes(datos['nubes'], vis_m, fenomeno)
        
        # Validar temperaturas
        temp = validar_numero(datos['temp'], -10, 40, "Temperatura")
        rocio = validar_numero(datos['rocio'], -10, 40, "Punto de rocío")
        validar_temp_rocio(temp, rocio)
        
        # Validar humedad
        hr = validar_humedad(datos['hr'])
        
        # Validar QNH
        qnh = validar_numero(datos['qnh'], 850, 1100, "QNH")
        
        # Verificar CAVOK
        es_cavok = verificar_cavok(vis_m, fenomeno, nubes)
        
        # Construir METAR
        metar_parts = [f"{datos['tipo']} SPJC {datos['dia']}{hora}Z {viento}"]
        
        if es_cavok:
            metar_parts.append("CAVOK")
        else:
            metar_parts.append(f"{vis_m:04d}")
            if vis_min_codigo:
                metar_parts.append(vis_min_codigo)
            if rvr_codigo:
                metar_parts.append(rvr_codigo)
            if fenomeno:
                metar_parts.append(fenomeno)
            metar_parts.append(nubes)
        
        metar_parts.append(f"{round(temp):02d}/{round(rocio):02d} Q{round(qnh)}")
        
        if datos['suplementaria']:
            metar_parts.append(datos['suplementaria'].upper())
        
        metar_completo = " ".join(metar_parts) + "="
        
        # Crear registro completo
        registro = {
            'Día': datos['dia'],
            'Hora': hora,
            'Tipo': datos['tipo'],
            'Dirección_Viento': datos['dir_viento'],
            'Intensidad_Viento': datos['int_viento'],
            'Variación_Viento': datos['var_viento'],
            'Visibilidad_Original': datos['vis'],
            'Visibilidad_Metros': vis_m,
            'Visibilidad_Mínima': vis_min_codigo,
            'RVR': rvr_codigo,
            'Fenómeno_Texto': datos['fenomeno'],
            'Fenómeno_Código': fenomeno,
            'Nubes_Texto': datos['nubes'],
            'Nubes_Código': "CAVOK" if es_cavok else nubes,
            'Temperatura': temp,
            'Punto_Rocío': rocio,
            'Humedad_Relativa_%': hr if hr else "",
            'QNH': qnh,
            'Presión_Estación': datos['presion'],
            'Info_Suplementaria': datos['suplementaria'],
            'METAR_Completo': metar_completo
        }
        
        return {
            'success': True,
            'metar': metar_completo,
            'registro': registro
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

# ============================================
# FUNCIÓN PARA EXPORTAR EXCEL - VERSIÓN CORREGIDA
# ============================================
def exportar_a_excel_v3(registros):
    """Exporta registros a Excel con formato profesional y encabezados personalizados - VERSIÓN CORREGIDA"""
    if not registros:
        return None, "No hay registros para exportar"
    
    try:
        # Crear DataFrame
        df = pd.DataFrame(registros)
        
        # DICCIONARIO DE RENOMBRADO - ENCABEZADOS PERSONALIZADOS EN ESPAÑOL
        mapeo_columnas = {
            'Día': 'DIA',
            'Hora': 'HORA',
            'Tipo': 'TIPO',
            'Dirección_Viento': 'DIRECCION DEL VIENTO',
            'Intensidad_Viento': 'INTENSIDAD DEL VIENTO',
            'Variación_Viento': 'VARIACION DEL VIENTO',
            'Visibilidad_Original': 'VISIBILIDAD (TEXTO CLARO)',
            'Visibilidad_Metros': 'VISIBILIDAD (CODIGO)',
            'Visibilidad_Mínima': 'VISIBILIDAD MINIMA',
            'RVR': 'RVR',
            'Fenómeno_Texto': 'FEN. METEOROLOGICOS (TEXTO CLARO)',
            'Fenómeno_Código': 'FEN. METEOROLOGICOS (CODIGO)',
            'Nubes_Texto': 'NUBES (TEXTO CLARO)',
            'Nubes_Código': 'NUBES (CODIGO)',
            'Temperatura': 'TEMPERATURA',
            'Punto_Rocío': 'PUNTO DE ROCIO',
            'Humedad_Relativa_%': 'HUMEDAD RELATIVA (%)',
            'QNH': 'QNH',
            'Presión_Estación': 'PRESION DE ESTACION',
            'Info_Suplementaria': 'INFORMACIÓN SUPLEMENTARIA',
            'METAR_Completo': 'METAR COMPLETO (CODIGO)'
        }
        
        # APLICAR EL RENOMBRADO DE COLUMNAS - ¡ESTA ES LA LÍNEA CLAVE!
        df = df.rename(columns=mapeo_columnas)
        
        # Usar los nuevos nombres de columnas
        columnas_nuevas = list(mapeo_columnas.values())
        
        # Asegurar que todas las columnas existen
        for col in columnas_nuevas:
            if col not in df.columns:
                df[col] = ""
        
        # Reordenar columnas
        df = df[columnas_nuevas]
        
        # Formatear DÍA y HORA con ceros a la izquierda
        df['DIA'] = df['DIA'].astype(str).str.zfill(2)
        df['HORA'] = df['HORA'].astype(str).str.zfill(4)
        
        # Exportar a Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='METAR V3', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['METAR V3']
            
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Ancho automático de columnas
            for col in range(1, len(columnas_nuevas) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                header_cell = worksheet.cell(row=1, column=col)
                max_length = len(str(header_cell.value)) if header_cell.value else 0
                
                for row in range(2, min(len(df) + 2, 102)):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = min(max_length + 2, 70)
                adjusted_width = max(adjusted_width, 8)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # FORMATO PROFESIONAL
            COLOR_HEADER = "0B3D91"
            COLOR_HEADER_TEXTO = "FFFFFF"
            COLOR_FILA_IMPAR = "E8EEF7"
            COLOR_BORDE = "CCCCCC"
            COLOR_SPECI = "FFE699"
            
            # Estilo para cabecera
            header_font = Font(name='Calibri', size=11, bold=True, color=COLOR_HEADER_TEXTO)
            header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_border = Border(
                left=Side(style='medium', color=COLOR_HEADER),
                right=Side(style='medium', color=COLOR_HEADER),
                top=Side(style='medium', color=COLOR_HEADER),
                bottom=Side(style='medium', color=COLOR_HEADER)
            )
            
            # Aplicar formato a cabecera
            for col in range(1, len(columnas_nuevas) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = header_border
            
            worksheet.row_dimensions[1].height = 30
            
            # Estilo para datos
            border_datos = Border(
                left=Side(style='thin', color=COLOR_BORDE),
                right=Side(style='thin', color=COLOR_BORDE),
                top=Side(style='thin', color=COLOR_BORDE),
                bottom=Side(style='thin', color=COLOR_BORDE)
            )
            
            fill_impar = PatternFill(start_color=COLOR_FILA_IMPAR, end_color=COLOR_FILA_IMPAR, fill_type="solid")
            font_datos = Font(name='Calibri', size=10)
            alineacion_centrada = Alignment(horizontal='center', vertical='center')
            alineacion_izquierda = Alignment(horizontal='left', vertical='center')
            
            # Estilo para SPECI
            speci_fill = PatternFill(start_color=COLOR_SPECI, end_color=COLOR_SPECI, fill_type="solid")
            speci_font = Font(name='Calibri', size=10, bold=True)
            
            # Aplicar formato a datos
            for row in range(2, len(df) + 2):
                es_impar = (row % 2 == 1)
                tipo_reporte = df.iloc[row-2]['TIPO']
                
                for col in range(1, len(columnas_nuevas) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    if tipo_reporte == "SPECI":
                        cell.fill = speci_fill
                        cell.font = speci_font
                        cell.border = border_datos
                    else:
                        cell.border = border_datos
                        cell.font = font_datos
                        if es_impar:
                            cell.fill = fill_impar
                    
                    # Centrar columnas específicas
                    col_letter = get_column_letter(col)
                    if col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                        cell.alignment = alineacion_centrada
                    else:
                        cell.alignment = alineacion_izquierda
            
            worksheet.freeze_panes = 'A2'
        
        output.seek(0)
        
        # Mensaje de éxito
        speci_count = len([r for r in registros if r.get('Tipo') == 'SPECI'])
        mensaje = f"✅ {len(registros)} registros exportados\n"
        mensaje += f"   📊 METAR: {len(registros) - speci_count}\n"
        mensaje += f"   🟨 SPECI: {speci_count} (resaltados en amarillo)"
        
        return output, mensaje
        
    except Exception as e:
        import traceback
        error_detallado = traceback.format_exc()
        print(error_detallado)
        return None, f"Error al exportar: {str(e)}"

# ============================================
# INTERFAZ DE USUARIO - VERSIÓN 3 CORREGIDA
# ============================================
# Header
col1, col2 = st.columns([3, 1])
with col1:
    st.markdown("""
    <div style='display: flex; align-items: center;'>
        <h1 style='color: #0b3d91; margin-right: 15px;'>✈️ METAR DIGITAL </h1>
        <span style='background: #27ae60; color: white; padding: 8px 20px; border-radius: 30px; font-size: 16px; font-weight: bold;'></span>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("<p style='color: #666; font-size: 1.1rem;'>Aeropuerto Internacional Jorge Chávez (SPJC) - CORPAC Perú</p>", unsafe_allow_html=True)

with col2:
    ahora = datetime.now(timezone.utc).strftime("%H:%M:%S")
    st.markdown(f"<h3 style='color: #0b3d91; text-align: right;'>UTC {ahora}</h3>", unsafe_allow_html=True)
    st.markdown(f"<p style='color: #666; text-align: right;'>{datetime.now().strftime('%d/%m/%Y')}</p>", unsafe_allow_html=True)

st.markdown("---")

# Columnas principales
col_izq, col_der = st.columns([2, 1])

with col_izq:
    with st.form(key='metar_form_v3'):
        # DATOS DEL REPORTE
        st.subheader("📋 DATOS DEL REPORTE")
        col1, col2, col3 = st.columns(3)
        with col1:
            tipo = st.selectbox("Tipo", ["METAR", "SPECI"], key='tipo_v3')
        with col2:
            dia = st.text_input("Día", key='dia_v3')
        with col3:
            hora = st.text_input("Hora UTC", key='hora_v3', help="Formato HHMM")
        
        st.markdown("---")
        
        # VIENTO
        st.subheader("💨 VIENTO")
        col1, col2, col3 = st.columns(3)
        with col1:
            dir_viento = st.text_input("Dirección", key='dir_viento_v3', help="Grados (0-360)")
        with col2:
            int_viento = st.text_input("Intensidad (KT)", key='int_viento_v3', help="Nudos. Ráfagas: 15G25")
        with col3:
            var_viento = st.text_input("Variación", key='var_viento_v3', help="Formato: 340V080")
        
        st.markdown("---")
        
        # VISIBILIDAD
        st.subheader("👁️ VISIBILIDAD")
        col1, col2, col3 = st.columns(3)
        with col1:
            vis = st.text_input("Visibilidad", key='vis_v3', help="Ej: 10km, 5000m, 9999")
        with col2:
            vis_min = st.text_input("Visibilidad Mínima", key='vis_min_v3', help="Ej: 1200SW, 0800NE, 3000N, 1500SE, 2000NW")
        with col3:
            rvr = st.text_input("RVR (m)", key='rvr_v3', help="Runway Visual Range - Ej: 0600, 1200")
        
        st.markdown("---")
        
        # FENÓMENOS Y NUBES
        st.subheader("☁️ FENÓMENOS Y NUBES")
        fenomeno = st.text_input("Fenómeno", key='fenomeno_v3', help="Ej: niebla parcial (PRFG), lluvia ligera (-RA)")
        nubes = st.text_input("Nubes", key='nubes_v3', help="Ej: 8 ST 300M, 5 AC 5000M, CB 1500M")
        
        st.markdown("---")
        
        # TEMPERATURA, HUMEDAD Y PRESIÓN
        st.subheader("🌡️ TEMPERATURA, HUMEDAD Y PRESIÓN")
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            temp = st.text_input("Temp °C", key='temp_v3', help="-10 a 40°C")
        with col2:
            rocio = st.text_input("Rocío °C", key='rocio_v3', help="≤ Temperatura")
        with col3:
            hr = st.text_input("HR %", key='hr_v3', help="Humedad Relativa (0-100%)")
        with col4:
            qnh = st.text_input("QNH hPa", key='qnh_v3', help="850-1100 hPa")
        with col5:
            presion = st.text_input("Presión Est.", key='presion_v3', help="Opcional")
        
        st.markdown("---")
        
        # INFORMACIÓN SUPLEMENTARIA
        st.subheader("📝 INFORMACIÓN SUPLEMENTARIA")
        suplementaria = st.text_input("Suplementaria", key='suplementaria_v3', help="Opcional - Ej: RMK CB AL NE")
        
        st.markdown("---")
        
        # BOTONES
        col1, col2 = st.columns(2)
        with col1:
            generar = st.form_submit_button("🔄 GENERAR METAR V3", use_container_width=True)
        with col2:
            limpiar = st.form_submit_button("🧹 LIMPIAR CAMPOS", use_container_width=True)
        
        if limpiar:
            limpiar_campos_v3()
            st.rerun()
        
        if generar:
            datos = {
                'tipo': tipo,
                'dia': dia,
                'hora': hora,
                'dir_viento': dir_viento,
                'int_viento': int_viento,
                'var_viento': var_viento,
                'vis': vis,
                'vis_min': vis_min,
                'rvr': rvr,
                'fenomeno': fenomeno,
                'nubes': nubes,
                'temp': temp,
                'rocio': rocio,
                'hr': hr,
                'qnh': qnh,
                'presion': presion,
                'suplementaria': suplementaria
            }
            
            resultado = generar_metar_v3(datos)
            
            if resultado['success']:
                # Actualizar o insertar registro (evitar duplicados)
                accion = actualizar_o_insertar_registro_v3(st.session_state.registros_v3, resultado['registro'])
                
                # Actualizar historial - eliminar duplicados
                dia_hora_clave = f"{resultado['registro']['Día']}_{resultado['registro']['Hora']}"
                
                nuevo_historial = []
                for metar in st.session_state.historial_v3:
                    match = re.search(r'SPJC (\d{2})(\d{4})Z', metar)
                    if match:
                        dia_hist = match.group(1)
                        hora_hist = match.group(2)
                        clave_hist = f"{dia_hist}_{hora_hist}"
                        if clave_hist != dia_hora_clave:
                            nuevo_historial.append(metar)
                    else:
                        nuevo_historial.append(metar)
                
                nuevo_historial.insert(0, resultado['metar'])
                st.session_state.historial_v3 = nuevo_historial
                st.session_state.contador_v3 = len(st.session_state.registros_v3)
                
                if accion == "actualizado":
                    st.warning("🔄 METAR ACTUALIZADO - Reemplazó reporte existente con la misma fecha/hora")
                else:
                    st.success("✅ METAR generado correctamente")
                
                st.session_state.ultimo_metar_v3 = resultado['metar']
                st.session_state.ultimo_tipo_v3 = tipo
            else:
                st.error(f"❌ {resultado['error']}")

with col_der:
    # METAR GENERADO
    st.subheader("📊 METAR GENERADO")
    if 'ultimo_metar_v3' in st.session_state:
        tipo_ultimo = st.session_state.get('ultimo_tipo_v3', 'METAR')
        if tipo_ultimo == "SPECI":
            st.markdown(f"<div style='background: #FFE699; padding: 15px; border-radius: 5px; font-family: monospace; border-left: 5px solid #FFC000;'><b>⚠️ SPECI V3</b><br>{st.session_state.ultimo_metar_v3}</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div style='background: #1e1e1e; color: #00ff00; padding: 15px; border-radius: 5px; font-family: monospace; border-left: 5px solid #0b3d91;'>{st.session_state.ultimo_metar_v3}</div>", unsafe_allow_html=True)
    else:
        st.info("---")
    
    st.markdown("---")
    
    # BOTONES DE ACCIÓN
    st.subheader("💾 EXPORTAR")
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("📥 Exportar Excel V3", use_container_width=True):
            if st.session_state.registros_v3:
                excel_file, mensaje = exportar_a_excel_v3(st.session_state.registros_v3)
                if excel_file:
                    st.download_button(
                        label="📥 Descargar Excel V3",
                        data=excel_file,
                        file_name=f"METAR_SPJC_{datetime.now().strftime('%Y_%m_%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    st.success(mensaje)
                else:
                    st.warning(mensaje)
            else:
                st.warning("⚠️ No hay registros para exportar")
    
    with col2:
        if st.button("🗑️ Limpiar Memoria V3", use_container_width=True):
            st.session_state.registros_v3 = []
            st.session_state.historial_v3 = []
            st.session_state.contador_v3 = 0
            st.success("✅ Memoria V3 limpiada")
    
    st.markdown("---")
    
    # CONTADOR DE REGISTROS ÚNICOS
    st.metric("📋 REGISTROS ÚNICOS V3", st.session_state.contador_v3)
    
    st.markdown("---")
    
    # HISTORIAL
    st.subheader("📜 HISTORIAL")
    if st.session_state.historial_v3:
        for i, metar in enumerate(st.session_state.historial_v3[:10]):
            if "SPECI" in metar:
                st.markdown(f"<div style='background: #FFE699; padding: 8px; margin-bottom: 5px; border-radius: 3px; font-family: monospace; font-size: 12px; border-left: 3px solid #FFC000;'>{metar}</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='background: #f0f0f0; padding: 8px; margin-bottom: 5px; border-radius: 3px; font-family: monospace; font-size: 12px; border-left: 3px solid #0b3d91;'>{metar}</div>", unsafe_allow_html=True)
    else:
        st.info("No hay METARs en el historial V3")

# ============================================
# FOOTER - VERSIÓN 3
# ============================================
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666; padding: 20px;'>
    <p style='font-size: 1.1rem;'>✈️ METAR Digital <span style='background: #27ae60; color: white; padding: 3px 15px; border-radius: 20px; font-weight: bold;'>V3.0</span> - CORPAC Perú</p>
</div>
""", unsafe_allow_html=True)
