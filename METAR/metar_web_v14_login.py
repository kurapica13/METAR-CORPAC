"""
METAR DIGITAL - VERSIÓN PROFESIONAL CORPAC PERÚ
Aeropuerto Internacional Jorge Chávez (SPJC)
Versión 14.0 - CON FRAGMENTS

Características:
✅ Sistema de hora inteligente - Bloqueada para METAR, modificable para SPECI
✅ Auditoría de diferencia horaria - Evita "trampas" con la hora
✅ Viento con 4 campos - Dirección, intensidad, variación desde/hasta
✅ Fenómenos con múltiples selecciones - Catálogo completo de Lima con botones +/-
✅ Nubes con validación 1-3-5 - Sistema de capas que valida regla de octas
✅ Visibilidad Vertical (VV) - Soporte completo con VV/// para altura desconocida
✅ PP Obligatorio - Formato PPTRZ (PPTRZ=trazas, PP001=0.1mm...PP010=1.0mm)
✅ TN/TX Automáticos - TN 12Z (mínima), TX 22Z (máxima)
✅ HR Automática - Cálculo informativo desde temp/rocío
✅ RVR - Formato SPJC: R32/0400, R12R/1700, R10/M0050, R14L/P2000
✅ Almacenamiento mensual automático (SPJC_METAR_YYYY_MM.xlsx)
✅ Viento con reglas circulares (340V080) - No muestra variación <60°
✅ Visibilidad mínima con cuadrantes (N, NE, E, SE, S, SW, W, NW)
✅ Estándar oficial nubes CORPAC (30m/1000m)
✅ Fenómenos especiales (PRFG, VCFG, BCFG, MIFG)
✅ Excel con formato profesional mensual
✅ Sin duplicados - Reemplaza reportes con misma fecha/hora
✅ Persistencia de datos entre sesiones
✅ Sistema de autenticación - Solo personal autorizado
✅ Fragments - Componentes interactivos independientes
"""

import streamlit as st
from datetime import datetime, timezone
import pandas as pd
from pathlib import Path
import re
import os
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
import hmac
from enum import Enum

# ============================================
# CONFIGURACIÓN DE PÁGINA
# ============================================
st.set_page_config(
    page_title="REGISTRO DE OBSERVACIONES ORDINARIAS Y ESPECIALES SPJC",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================
# ENUMS PARA CONSTANTES
# ============================================
class TipoReporte(str, Enum):
    METAR = "METAR"
    SPECI = "SPECI"

class Cuadrante(str, Enum):
    N = "N"
    NE = "NE"
    E = "E"
    SE = "SE"
    S = "S"
    SW = "SW"
    W = "W"
    NW = "NW"

# ============================================
# LISTA COMPLETA DE FENÓMENOS PARA LIMA (SPJC)
# ============================================
FENOMENOS_LIMA = {
    "Nieblas": [
        "FG - Niebla",
        "PRFG - Niebla Parcial",
        "BCFG - Niebla en Bancos",
        "MIFG - Niebla Baja",
        "VCFG - Niebla en la Vecindad",
    ],
    "Nieblinas": [
        "BR - Neblina",
    ],
    "Precipitación": [
        "-RA - Lluvia Ligera",
        "RA - Lluvia Moderada",
        "+RA - Lluvia Fuerte",
        "-DZ - Llovizna Ligera",
        "DZ - Llovizna Moderada",
        "+DZ - Llovizna Fuerte",
        "SHRA - Chubascos de Lluvia",
        "-SHRA - Chubascos Ligeros",
        "+SHRA - Chubascos Fuertes"
    ],
    "Tormentas": [
        "TS - Tormenta",
        "-TSRA - Tormenta con Lluvia Ligera",
        "TSRA - Tormenta con Lluvia",
        "+TSRA - Tormenta con Lluvia Fuerte"
    ],
    "Otros": [
        "HZ - Calima",
        "FU - Humo",
        "DU - Polvo",
        "SA - Arena",
        "VA - Ceniza Volcánica"
    ]
}

# Opciones de precipitación en formato PPTRZ
OPCIONES_PP = {
    "PPTRZ": "Trazas (< 0.1 mm)",
    "PP001": "0.1 mm",
    "PP002": "0.2 mm",
    "PP003": "0.3 mm",
    "PP004": "0.4 mm",
    "PP005": "0.5 mm",
    "PP006": "0.6 mm",
    "PP007": "0.7 mm",
    "PP008": "0.8 mm",
    "PP009": "0.9 mm",
    "PP010": "1.0 mm"
}

# Tipos de nubes y octas
TIPOS_NUBES = ["CU", "SC", "ST", "AC", "AS", "NS", "CI", "CB", "TCU"]
OCTAS = ["1", "2", "3", "4", "5", "6", "7", "8"]

# Mapeo de octas a códigos METAR
MAPEO_OCTAS = {
    '1': 'FEW', '2': 'FEW',
    '3': 'SCT', '4': 'SCT',
    '5': 'BKN', '6': 'BKN', '7': 'BKN',
    '8': 'OVC'
}

# ============================================
# SISTEMA DE AUTENTICACIÓN
# ============================================
def verificar_autenticacion():
    """Sistema completo de login/logout"""
    
    if 'autenticado' not in st.session_state:
        st.session_state.autenticado = False
        st.session_state.usuario = None
    
    with st.sidebar:
        if st.session_state.autenticado:
            st.markdown(f"👤 **Usuario:** {st.session_state.usuario}")
            if st.button("🚪 Cerrar Sesión", use_container_width=True):
                st.session_state.autenticado = False
                st.session_state.usuario = None
                st.rerun()
            st.markdown("---")
    
    if st.session_state.autenticado:
        return True
    
    # Pantalla de login
    st.markdown("""
    <style>
        .login-container {
            max-width: 400px;
            margin: 100px auto;
            padding: 30px;
            background: var(--background-color);
            border-radius: 10px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            text-align: center;
            border: 1px solid rgba(128,128,128,0.2);
        }
        .login-header {
            color: #0b3d91;
            margin-bottom: 20px;
        }
        .login-logo {
            font-size: 48px;
            margin-bottom: 10px;
        }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown("<div class='login-container'>", unsafe_allow_html=True)
    st.markdown("<div class='login-logo'>✈️</div>", unsafe_allow_html=True)
    st.markdown("<h2 class='login-header'>Sistema METAR Digital</h2>", unsafe_allow_html=True)
    st.markdown("Aeropuerto Internacional Jorge Chávez")
    st.markdown("CORPAC Perú")
    st.markdown("---")
    
    with st.form("login_form"):
        usuario = st.text_input("Usuario", placeholder="Ingrese su usuario")
        contraseña = st.text_input("Contraseña", type="password", placeholder="Ingrese su contraseña")
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            submit = st.form_submit_button("🔐 INGRESAR", use_container_width=True)
    
    if submit:
        try:
            passwords = st.secrets.get("passwords", {})
            if not passwords:
                passwords = {"admin": "corpac2024", "metar": "spjc2024"}
        except:
            passwords = {"admin": "corpac2024", "metar": "spjc2024"}
        
        if usuario in passwords and hmac.compare_digest(contraseña, passwords[usuario]):
            st.session_state.autenticado = True
            st.session_state.usuario = usuario
            st.rerun()
        else:
            st.error("❌ Usuario o contraseña incorrectos")
    
    st.markdown("---")
    st.markdown("Solo personal autorizado CORPAC")
    st.markdown("</div>", unsafe_allow_html=True)
    
    st.stop()

# ============================================
# VERIFICAR AUTENTICACIÓN
# ============================================
verificar_autenticacion()

# ============================================
# CONSTANTES Y CONFIGURACIÓN
# ============================================
DIRECTORIO_DATOS = Path("datos_metar")
DIRECTORIO_DATOS.mkdir(exist_ok=True)

# ============================================
# FUNCIÓN DE REDONDEO CORREGIDA
# ============================================
def redondear_metar(valor):
    """
    Redondeo tradicional para METAR:
    - 14.5 → 15
    - 14.4 → 14
    - 15.5 → 16
    """
    try:
        d = Decimal(str(valor))
        return int(d.quantize(Decimal('1'), rounding=ROUND_HALF_UP))
    except:
        return int(round(float(valor)))

# ============================================
# FUNCIONES DE UTILIDAD
# ============================================
def calcular_hr_automatica(temp_c, rocio_c):
    """
    Calcula la Humedad Relativa usando la fórmula de Magnus
    SOLO INFORMATIVA - El operador ingresa el valor real del sensor
    """
    try:
        a, b = 17.625, 243.04
        rocio_c = min(rocio_c, temp_c)
        
        es_temp = a * temp_c / (b + temp_c)
        es_rocio = a * rocio_c / (b + rocio_c)
        
        hr_calculada = 100 * (10**(es_rocio - es_temp))
        return round(min(max(hr_calculada, 0), 100))
    except:
        return None

def on_tipo_change():
    """Maneja el cambio de tipo de reporte y actualiza la hora automáticamente"""
    tipo_seleccionado = st.session_state.tipo_selector
    st.session_state.tipo = tipo_seleccionado
    
    hora_actual_utc = datetime.now(timezone.utc).strftime("%H%M")
    
    if tipo_seleccionado == TipoReporte.METAR.value:
        st.session_state.hora = hora_actual_utc
        st.session_state.hora_bloqueada = True
        st.session_state.hora_original = hora_actual_utc
    else:  # SPECI
        st.session_state.hora_bloqueada = False
        if not st.session_state.get('hora'):
            st.session_state.hora = hora_actual_utc
        st.session_state.hora_original = hora_actual_utc

def validar_hora_auditoria(hora_ingresada):
    """
    Valida que la hora no sea muy antigua y registra la diferencia
    """
    hora_actual = datetime.now(timezone.utc).strftime("%H%M")
    
    def a_minutos(hhmm):
        return int(hhmm[:2]) * 60 + int(hhmm[2:])
    
    minutos_ingresados = a_minutos(hora_ingresada)
    minutos_actuales = a_minutos(hora_actual)
    
    if minutos_ingresados > minutos_actuales:
        diferencia = (1440 - minutos_ingresados) + minutos_actuales
    else:
        diferencia = minutos_actuales - minutos_ingresados
    
    st.session_state.ultima_diferencia_hora = diferencia
    
    if st.session_state.tipo == TipoReporte.METAR.value and diferencia > 1:
        return False, f"METAR debe ser a la hora exacta. Diferencia: {diferencia} min"
    elif st.session_state.tipo == TipoReporte.SPECI.value and diferencia > 15:
        return False, f"SPECI no puede tener más de 15 min de retraso. Diferencia: {diferencia} min"
    
    return True, f"Diferencia aceptable: {diferencia} min"

# ============================================
# ESTILOS CSS PERSONALIZADOS
# ============================================
st.markdown("""
<style>
    .stApp {
        background-color: var(--background-color);
    }
    
    .section-title {
        color: #0b3d91 !important;
        font-weight: 700 !important;
        font-size: 1.2rem !important;
        margin-bottom: 1rem !important;
        border-bottom: 2px solid #0b3d91 !important;
        padding-bottom: 0.5rem !important;
        text-shadow: 0 1px 2px rgba(255,255,255,0.5);
    }
    
    .stTextInput label, .stSelectbox label {
        color: #0b3d91 !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
    }
    
    @media (prefers-color-scheme: dark) {
        .stTextInput label, .stSelectbox label {
            color: #6ab0ff !important;
        }
        .section-title {
            color: #6ab0ff !important;
            border-bottom-color: #6ab0ff !important;
        }
    }
    
    .metar-box {
        background: #1e1e1e;
        color: #00ff00;
        padding: 1rem;
        border-radius: 5px;
        font-family: 'Courier New', monospace;
        font-size: 1.1rem;
        border-left: 5px solid #0b3d91;
    }
    
    .historial-item {
        background: #f8f9fa;
        padding: 0.8rem;
        margin-bottom: 0.5rem;
        border-radius: 3px;
        font-family: 'Courier New', monospace;
        font-size: 12px;
        border-left: 3px solid #0b3d91;
    }
    
    .historial-item-speci {
        background: #FFE699;
        border-left: 3px solid #FFC000;
    }
    
    .stButton button {
        width: 100%;
        border-radius: 5px;
        font-weight: 600;
    }
    
    .stButton button[kind="primary"] {
        background-color: #0b3d91;
        color: white;
        border: none;
    }
    
    .stButton button[kind="primary"]:hover {
        background-color: #1a4fa0;
    }
    
    .delete-btn {
        color: #ff4444;
        cursor: pointer;
        font-size: 20px;
    }
    
    .info-box {
        background: #e7f3ff;
        padding: 10px;
        border-radius: 5px;
        border-left: 3px solid #0b3d91;
        margin: 10px 0;
    }
    
    .vv-box {
        background: #f0f0f0;
        padding: 10px;
        border-radius: 5px;
        border-left: 3px solid #ff9900;
        margin: 5px 0;
    }
    
    .fragment-container {
        border: 1px solid #e0e0e0;
        border-radius: 5px;
        padding: 15px;
        margin: 10px 0;
        background: #fafafa;
    }
</style>
""", unsafe_allow_html=True)

# ============================================
# FUNCIONES DE GESTIÓN DE ARCHIVOS
# ============================================
def obtener_nombre_archivo_mensual():
    """Genera el nombre del archivo mensual basado en la fecha actual UTC"""
    ahora = datetime.now(timezone.utc)
    return f"SPJC_METAR_{ahora.strftime('%Y_%m')}.xlsx"

def cargar_registros_mes():
    """Carga todos los registros del mes actual desde el archivo Excel"""
    archivo = DIRECTORIO_DATOS / obtener_nombre_archivo_mensual()
    
    if archivo.exists():
        try:
            df = pd.read_excel(archivo, sheet_name='METAR SPJC')
            
            registros = []
            for _, row in df.iterrows():
                registro = row.to_dict()
                
                if 'DIA' in registro:
                    registro['Día'] = str(registro['DIA']).zfill(2)
                if 'HORA' in registro:
                    registro['Hora'] = str(registro['HORA']).zfill(4)
                if 'TIPO' in registro:
                    registro['Tipo'] = registro['TIPO']
                if 'METAR' in registro:
                    registro['METAR_Completo'] = registro['METAR']
                
                registros.append(registro)
            
            return registros
        except Exception as e:
            st.error(f"Error al cargar archivo mensual: {e}")
            return []
    return []

def guardar_registros_mes(registros):
    """Guarda todos los registros en el archivo mensual"""
    if not registros:
        return False, "No hay registros para guardar"
    
    try:
        archivo = DIRECTORIO_DATOS / obtener_nombre_archivo_mensual()
        
        df = pd.DataFrame(registros)
        
        df = df.rename(columns={
            'Día': 'DIA',
            'Hora': 'HORA',
            'Tipo': 'TIPO',
            'Dirección_Viento': 'DIR VIENTO',
            'Intensidad_Viento': 'INTENSIDAD',
            'Variación_Viento': 'VARIACION',
            'Visibilidad_Original': 'VIS (ORIGINAL)',
            'Visibilidad_Metros': 'VIS (CODIGO)',
            'Visibilidad_Mínima': 'VIS MIN',
            'RVR': 'RVR',
            'Fenómeno_Texto': 'FENOMENO',
            'Fenómeno_Código': 'WX',
            'Nubes_Texto': 'NUBOSIDAD',
            'Nubes_Código': 'CLD',
            'Temperatura': 'TEMP °C',
            'Punto_Rocío': 'ROCÍO °C',
            'Humedad_Relativa_%': 'HR %',
            'QNH': 'QNH',
            'Presión_Estación': 'PRESION',
            'Info_Suplementaria': 'RMK',
            'METAR_Completo': 'METAR'
        })
        
        columnas_excel = [
            'DIA', 'HORA', 'TIPO', 'DIR VIENTO', 'INTENSIDAD', 'VARIACION',
            'VIS (ORIGINAL)', 'VIS (CODIGO)', 'VIS MIN', 'RVR',
            'FENOMENO', 'WX', 'NUBOSIDAD', 'CLD',
            'TEMP °C', 'ROCÍO °C', 'HR %', 'QNH', 'PRESION', 'RMK', 'METAR'
        ]
        
        columnas_disponibles = [col for col in columnas_excel if col in df.columns]
        df = df[columnas_disponibles]
        
        df['DIA'] = df['DIA'].astype(str).str.zfill(2)
        df['HORA'] = df['HORA'].astype(str).str.zfill(4)
        df = df.sort_values(['DIA', 'HORA'], ascending=[True, True])
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='METAR SPJC', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['METAR SPJC']
            
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            for col in range(1, len(columnas_disponibles) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                header_cell = worksheet.cell(row=1, column=col)
                max_length = len(str(header_cell.value)) if header_cell.value else 0
                
                for row in range(2, min(len(df) + 2, 102)):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = min(max_length + 2, 70)
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 8)
            
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='0B3D91', end_color='0B3D91', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for col in range(1, len(columnas_disponibles) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            worksheet.row_dimensions[1].height = 30
            
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            for row in range(2, len(df) + 2):
                tipo_reporte = worksheet.cell(row=row, column=3).value
                for col in range(1, len(columnas_disponibles) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if tipo_reporte == 'SPECI':
                        cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                        cell.font = Font(name='Calibri', size=10, bold=True)
                    else:
                        cell.font = Font(name='Calibri', size=10)
            
            worksheet.freeze_panes = 'A2'
        
        output.seek(0)
        
        with open(archivo, 'wb') as f:
            f.write(output.getvalue())
        
        return True, f"✅ {len(registros)} registros guardados en {archivo.name}"
        
    except Exception as e:
        return False, f"Error al guardar: {str(e)}"

# ============================================
# FRAGMENTS - COMPONENTES INTERACTIVOS
# ============================================

@st.fragment
def fragment_fenomenos():
    """
    Fragment independiente para manejar fenómenos con botones +/-
    """
    st.markdown("**Fenómenos:**")
    
    if 'fenomenos_seleccionados' not in st.session_state:
        st.session_state.fenomenos_seleccionados = []
    
    # Mostrar fenómenos seleccionados con botón eliminar
    for i, fen in enumerate(st.session_state.fenomenos_seleccionados):
        col1, col2 = st.columns([10, 1])
        with col1:
            st.info(f"📌 {fen}")
        with col2:
            if st.button("✖", key=f"del_fen_{i}"):
                st.session_state.fenomenos_seleccionados.pop(i)
                st.rerun()
    
    # Selector para agregar nuevo fenómeno
    col1, col2 = st.columns([4, 1])
    with col1:
        opciones_planas = []
        for categoria, fenomenos in FENOMENOS_LIMA.items():
            opciones_planas.append(f"--- {categoria} ---")
            opciones_planas.extend(fenomenos)
        
        nuevo_fenomeno = st.selectbox(
            "Agregar fenómeno",
            options=[""] + opciones_planas,
            key="selector_fenomeno",
            format_func=lambda x: x if x else "Seleccione un fenómeno...",
            label_visibility="collapsed"
        )
    
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("➕", key="add_fen", use_container_width=True):
            if nuevo_fenomeno and not nuevo_fenomeno.startswith("---"):
                if nuevo_fenomeno not in st.session_state.fenomenos_seleccionados:
                    st.session_state.fenomenos_seleccionados.append(nuevo_fenomeno)
                    st.rerun()

@st.fragment
def fragment_nubes():
    """
    Fragment independiente para manejar nubes con botones +/-
    Incluye soporte para Visibilidad Vertical (VV)
    """
    st.markdown("**Nubosidad:**")
    
    # Inicializar capas de nubes si no existen
    if 'capas_nubes' not in st.session_state:
        st.session_state.capas_nubes = []
    
    # Inicializar VV si no existe
    if 'vv_activo' not in st.session_state:
        st.session_state.vv_activo = False
        st.session_state.vv_valor = ""
    
    # Mostrar VV si está activo
    if st.session_state.vv_activo:
        with st.container():
            col1, col2, col3, col4 = st.columns([1, 3, 3, 1])
            with col1:
                st.markdown(f"**VV**")
            with col2:
                if st.session_state.vv_valor == "///":
                    st.markdown(f"📊 Visibilidad Vertical: DESCONOCIDA")
                else:
                    st.markdown(f"📊 Visibilidad Vertical: {st.session_state.vv_valor}m")
            with col3:
                if st.session_state.vv_valor == "///":
                    st.markdown(f"**Código:** VV///")
                else:
                    try:
                        vv_codigo = f"VV{round(int(st.session_state.vv_valor)/30):03d}"
                        st.markdown(f"**Código:** {vv_codigo}")
                    except:
                        st.markdown(f"**Código:** VV///")
            with col4:
                if st.button("✖", key="del_vv"):
                    st.session_state.vv_activo = False
                    st.session_state.vv_valor = ""
                    st.rerun()
            st.markdown("---")
    
    # Mostrar capas de nubes existentes
    for i, capa in enumerate(st.session_state.capas_nubes):
        with st.container():
            col1, col2, col3, col4, col5 = st.columns([1, 1, 2, 2, 1])
            
            with col1:
                st.markdown(f"**Capa {i+1}**")
            
            with col2:
                octas = capa.get('octas', '')
                st.markdown(f"**{octas}** octas")
            
            with col3:
                tipo = capa.get('tipo', '')
                st.markdown(f"**{tipo}**")
            
            with col4:
                altura = capa.get('altura', '')
                st.markdown(f"**{altura} m**")
            
            with col5:
                if st.button("✖", key=f"del_capa_{i}"):
                    st.session_state.capas_nubes.pop(i)
                    st.rerun()
    
    st.markdown("---")
    
    # Opciones para agregar (Nubes normales o VV)
    if not st.session_state.vv_activo:
        opcion_agregar = st.radio(
            "Tipo de nubosidad a agregar:",
            ["☁️ Capa de nubes", "🌫️ Visibilidad Vertical (VV)"],
            horizontal=True,
            key="tipo_nubosidad"
        )
        
        if opcion_agregar == "☁️ Capa de nubes":
            st.markdown("**Agregar nueva capa de nubes:**")
            
            col1, col2, col3, col4 = st.columns([1, 2, 2, 1])
            
            with col1:
                nueva_octa = st.selectbox(
                    "Octas", 
                    options=[""] + OCTAS, 
                    key="nueva_octa",
                    help="Cantidad de octas (1-8)"
                )
            
            with col2:
                nuevo_tipo = st.selectbox(
                    "Tipo", 
                    options=[""] + TIPOS_NUBES, 
                    key="nuevo_tipo",
                    help="Tipo de nube (CU, SC, ST, etc.)"
                )
            
            with col3:
                nueva_altura = st.text_input(
                    "Altura (m)", 
                    key="nueva_altura", 
                    placeholder="300",
                    help="Altura en metros sobre el nivel de la estación"
                )
            
            with col4:
                st.markdown("<br>", unsafe_allow_html=True)
                agregar = st.button("➕ Agregar Capa", key="add_capa", use_container_width=True)
            
            # Validar y agregar capa
            if agregar:
                if not nueva_octa:
                    st.error("❌ Seleccione la cantidad de octas")
                elif not nuevo_tipo:
                    st.error("❌ Seleccione el tipo de nube")
                elif not nueva_altura:
                    st.error("❌ Ingrese la altura en metros")
                else:
                    try:
                        altura_int = int(nueva_altura)
                        if altura_int < 0 or altura_int > 30000:
                            st.error("❌ Altura fuera de rango (0-30000m)")
                        else:
                            error_validacion = validar_regla_nubes(
                                st.session_state.capas_nubes,
                                int(nueva_octa),
                                nuevo_tipo,
                                nueva_altura
                            )
                            
                            if error_validacion:
                                st.error(error_validacion)
                            else:
                                st.session_state.capas_nubes.append({
                                    'octas': nueva_octa,
                                    'tipo': nuevo_tipo,
                                    'altura': nueva_altura
                                })
                                st.rerun()
                    except ValueError:
                        st.error("❌ La altura debe ser un número válido")
        
        else:  # Visibilidad Vertical
            st.markdown("**Agregar Visibilidad Vertical (VV):**")
            st.caption("La VV se usa cuando hay oscurecimiento y no se pueden distinguir capas de nubes")
            
            if st.session_state.capas_nubes:
                st.warning("⚠️ VV no puede combinarse con capas de nubes normales")
            
            col1, col2, col3 = st.columns([3, 2, 1])
            
            with col1:
                vv_altura = st.text_input(
                    "Altura de visibilidad vertical (m)",
                    key="vv_altura_input",
                    placeholder="600 (o dejar vacío para VV///)",
                    help="Altura hasta donde se puede ver verticalmente. Dejar vacío si es desconocida."
                )
            
            with col2:
                if vv_altura:
                    try:
                        vv_int = int(vv_altura)
                        vv_cientos = round(vv_int / 30)
                        st.markdown(f"<br><h3 style='color: #ff9900;'>VV{vv_cientos:03d}</h3>", unsafe_allow_html=True)
                    except:
                        st.markdown("<br><h3 style='color: #ff9900;'>VV///</h3>", unsafe_allow_html=True)
                else:
                    st.markdown("<br><h3 style='color: #ff9900;'>VV///</h3>", unsafe_allow_html=True)
            
            with col3:
                st.markdown("<br>", unsafe_allow_html=True)
                agregar_vv = st.button("➕ Activar VV", key="add_vv", use_container_width=True)
            
            if agregar_vv:
                if st.session_state.capas_nubes:
                    st.error("❌ No se puede activar VV cuando hay capas de nubes normales")
                elif not vv_altura:
                    st.session_state.vv_activo = True
                    st.session_state.vv_valor = "///"
                    st.session_state.capas_nubes = []
                    st.rerun()
                else:
                    try:
                        vv_int = int(vv_altura)
                        if vv_int < 0 or vv_int > 3000:
                            st.error("❌ Altura fuera de rango (0-3000m)")
                        else:
                            st.session_state.vv_activo = True
                            st.session_state.vv_valor = str(vv_int)
                            st.session_state.capas_nubes = []
                            st.rerun()
                    except ValueError:
                        st.error("❌ Ingrese un número válido")

def validar_regla_nubes(capas_existentes, nuevas_octas, nuevo_tipo, nueva_altura):
    """
    Valida la regla 1-3-5 para capas de nubes
    """
    num_capa = len(capas_existentes) + 1
    
    if num_capa > 3:
        return "❌ Máximo 3 capas de nubes permitidas"
    
    if num_capa == 1 and nuevas_octas < 1:
        return "❌ La primera capa debe tener al menos 1 octa"
    elif num_capa == 2 and nuevas_octas < 3:
        return "❌ La segunda capa debe tener al menos 3 octas"
    elif num_capa == 3 and nuevas_octas < 5:
        return "❌ La tercera capa debe tener al menos 5 octas"
    
    for capa in capas_existentes:
        if capa['tipo'] == nuevo_tipo and capa['altura'] == nueva_altura:
            return f"⚠️ Ya existe una capa de {nuevo_tipo} a {nueva_altura}m"
    
    return None

def convertir_nubes_completo_a_metar():
    """
    Convierte tanto VV como capas de nubes a formato METAR
    """
    # Si hay VV activo, retornar VV
    if st.session_state.get('vv_activo', False):
        if st.session_state.vv_valor == "///":
            return "VV///"
        else:
            try:
                vv_metros = int(st.session_state.vv_valor)
                vv_cientos = round(vv_metros / 30)
                vv_cientos = min(max(vv_cientos, 0), 999)
                return f"VV{vv_cientos:03d}"
            except:
                return "VV///"
    
    # Si no hay capas, retornar NSC
    if not st.session_state.get('capas_nubes', []):
        return "NSC"
    
    # Convertir capas normales
    codigos = []
    
    for capa in st.session_state.capas_nubes:
        octas = capa['octas']
        tipo = capa['tipo']
        altura = int(capa['altura'])
        
        altura_cientos = round(altura / 30)
        altura_cientos = min(max(altura_cientos, 1), 999)
        
        cod_cant = MAPEO_OCTAS.get(octas, 'SCT')
        
        codigo = f"{cod_cant}{altura_cientos:03d}"
        if tipo in ['CB', 'TCU']:
            codigo += tipo
        
        codigos.append(codigo)
    
    return " ".join(codigos[:4])

# ============================================
# FUNCIONES DE PROCESAMIENTO
# ============================================

def procesar_viento(direccion, intensidad, variacion):
    """
    PROCESAMIENTO DE VIENTO - REGLAS CORPAC PERÚ
    """
    dir_int = int(direccion)
    intensidad_str = str(intensidad).upper().strip()
    
    if dir_int == 0 and intensidad_str == "00":
        return "00000KT"
    
    if 'G' in intensidad_str:
        if 'G' in intensidad_str and not ' ' in intensidad_str.replace('G', ''):
            base_int, gust_int = intensidad_str.split('G')
            int_base = int(base_int)
            int_gust = int(gust_int)
            intensidad_metar = f"{int_base:02d}G{int_gust:02d}"
        else:
            parts = intensidad_str.replace('G', ' ').split()
            int_base = int(parts[0])
            int_gust = int(parts[1])
            intensidad_metar = f"{int_base:02d}G{int_gust:02d}"
    else:
        int_base = int(intensidad_str)
        intensidad_metar = f"{int_base:02d}"
    
    if not variacion:
        return f"{dir_int:03d}{intensidad_metar}KT"
    
    try:
        variacion = variacion.upper().replace(' ', '')
        if 'V' not in variacion:
            return f"{dir_int:03d}{intensidad_metar}KT"
        
        desde, hasta = map(int, variacion.split('V'))
        
        diff1 = abs(hasta - desde)
        diff2 = 360 - diff1
        
        if desde > hasta:
            diferencia = diff2
        else:
            diferencia = diff1
        
        if diferencia < 60:
            return f"{dir_int:03d}{intensidad_metar}KT"
        
        if diferencia >= 180:
            return f"VRB{intensidad_metar}KT"
        
        if diferencia >= 60:
            if int_base < 3:
                return f"VRB{intensidad_metar}KT"
            else:
                if desde < hasta:
                    return f"{dir_int:03d}{intensidad_metar}KT {desde:03d}V{hasta:03d}"
                else:
                    return f"{dir_int:03d}{intensidad_metar}KT {hasta:03d}V{desde:03d}"
        
        return f"{dir_int:03d}{intensidad_metar}KT"
        
    except Exception as e:
        return f"{dir_int:03d}{intensidad_metar}KT"

def convertir_visibilidad(vis_texto):
    """Convierte texto de visibilidad a metros"""
    vis_texto = vis_texto.strip().upper()
    if not vis_texto:
        raise ValueError("Visibilidad es obligatoria")
    
    try:
        if vis_texto.endswith("KM"):
            km = float(vis_texto[:-2])
            return 9999 if km >= 10 else int(km * 1000)
        elif vis_texto.endswith("M"):
            return int(vis_texto[:-1])
        else:
            metros = int(vis_texto)
            return 9999 if metros >= 10000 else metros
    except:
        raise ValueError("Formato de visibilidad inválido")

def procesar_visibilidad_minima(vis_min_texto, vis_m):
    """Procesa visibilidad mínima con cuadrantes"""
    if not vis_min_texto:
        return "", ""
    
    vis_min_texto = vis_min_texto.strip().upper()
    
    for cuadrante in [Cuadrante.NW, Cuadrante.NE, Cuadrante.SW, Cuadrante.SE,
                      Cuadrante.N, Cuadrante.S, Cuadrante.E, Cuadrante.W]:
        if vis_min_texto.endswith(cuadrante.value):
            valor = vis_min_texto[:-len(cuadrante.value)]
            break
    else:
        valor = vis_min_texto
        cuadrante = None
    
    try:
        if valor.endswith("KM"):
            km = float(valor[:-2])
            vis_min_m = 9999 if km >= 10 else int(km * 1000)
        elif valor.endswith("M"):
            vis_min_m = int(valor[:-1])
        else:
            vis_min_m = int(valor)
            vis_min_m = 9999 if vis_min_m >= 10000 else vis_min_m
        
        es_valida = False
        if vis_min_m < 1500:
            es_valida = True
        if vis_min_m < (vis_m * 0.5) and vis_min_m < 5000:
            es_valida = True
        
        if not es_valida:
            return "", "⚠️ No cumple reglas de visibilidad mínima"
        
        if cuadrante:
            return f"{vis_min_m:04d}{cuadrante.value}", ""
        else:
            return f"{vis_min_m:04d}", ""
        
    except:
        return "", "❌ Formato inválido"

def procesar_rvr(rvr_texto):
    """Procesa RVR"""
    if not rvr_texto:
        return ""
    return rvr_texto.strip()

def codificar_fenomenos(texto, visibilidad_metros):
    """Codifica fenómenos a formato METAR"""
    if not texto:
        return ""
    
    texto_lower = texto.lower().strip()
    precipitaciones = []
    oscurecimiento = []
    especiales = []
    
    if any(x in texto_lower for x in ["niebla parcial", "prfg", "pr fg", "parcial"]):
        especiales.append("PRFG")
        for palabra in ["niebla parcial", "prfg", "pr fg", "parcial"]:
            texto_lower = texto_lower.replace(palabra, "")
    
    if any(x in texto_lower for x in ["niebla en la vecindad", "vcfg", "vc fg", "vecindad"]):
        especiales.append("VCFG")
        for palabra in ["niebla en la vecindad", "vcfg", "vc fg", "vecindad"]:
            texto_lower = texto_lower.replace(palabra, "")
    
    if any(x in texto_lower for x in ["niebla en bancos", "bcfg", "bc fg", "bancos"]):
        especiales.append("BCFG")
        for palabra in ["niebla en bancos", "bcfg", "bc fg", "bancos"]:
            texto_lower = texto_lower.replace(palabra, "")
    
    if any(x in texto_lower for x in ["niebla baja", "mifg", "mi fg", "baja"]):
        especiales.append("MIFG")
        for palabra in ["niebla baja", "mifg", "mi fg", "baja"]:
            texto_lower = texto_lower.replace(palabra, "")
    
    intensidades = {
        "ligera": "-", "ligero": "-", "leve": "-", "débil": "-", 
        "moderada": "", "moderado": "",
        "fuerte": "+", "intensa": "+", "intenso": "+", "severa": "+"
    }
    
    descriptores = {
        "sh": "SH", "chubasco": "SH", "chubascos": "SH",
        "ts": "TS", "tormenta": "TS", "tormentas": "TS",
        "fz": "FZ", "helada": "FZ", "congelante": "FZ", "congelando": "FZ"
    }
    
    precipitacion_map = {
        "lluvia": "RA", "llovizna": "DZ", "nieve": "SN", 
        "granizo": "GR", "cellisca": "GS"
    }
    
    def get_oscurecimiento_codigo(texto_parte, vis_m):
        if "neblina" in texto_parte:
            if vis_m < 1000:
                return "FG"
            elif vis_m <= 5000:
                return "BR"
            else:
                return None
        elif "niebla" in texto_parte and "parcial" not in texto_parte and "baja" not in texto_parte:
            if vis_m < 1000:
                return "FG"
            else:
                return None
        return None
    
    import re
    texto_lower = re.sub(r'\s+y\s+', ', ', texto_lower)
    partes = re.split(r'[,;]', texto_lower)
    
    for parte in partes:
        parte = parte.strip()
        if not parte:
            continue
        
        codigo_base = None
        descriptor = ""
        intensidad = ""
        tipo = None
        parte_procesada = parte
        
        for d_texto, d_codigo in descriptores.items():
            if d_texto in parte_procesada:
                descriptor = d_codigo
                parte_procesada = parte_procesada.replace(d_texto, "").strip()
                break
        
        for i_texto, i_codigo in intensidades.items():
            if i_texto in parte_procesada:
                intensidad = i_codigo
                parte_procesada = parte_procesada.replace(i_texto, "").strip()
                break
        
        for f_texto, f_codigo in precipitacion_map.items():
            if f_texto in parte_procesada:
                codigo_base = f_codigo
                tipo = 'precipitacion'
                break
        
        if not codigo_base:
            codigo_osc = get_oscurecimiento_codigo(parte_procesada, visibilidad_metros)
            if codigo_osc:
                codigo_base = codigo_osc
                tipo = 'oscurecimiento'
        
        if codigo_base:
            codigo_final = codigo_base
            if descriptor and codigo_base not in ["FG", "BR"]:
                codigo_final = descriptor + codigo_final
            if intensidad and codigo_base not in ["FG", "BR"]:
                codigo_final = intensidad + codigo_final
            
            if tipo == 'precipitacion' and codigo_final not in precipitaciones:
                precipitaciones.append(codigo_final)
            elif tipo == 'oscurecimiento' and codigo_final not in oscurecimiento:
                oscurecimiento.append(codigo_final)
    
    resultados = precipitaciones + especiales + oscurecimiento
    return " ".join(resultados[:3]) if resultados else ""

def interpretar_nubes(texto, vis_m, fenomeno):
    """
    CODIFICADOR DE NUBES - ESTÁNDAR CORPAC
    """
    if not texto:
        texto = ""
    
    texto = texto.strip().upper()
    
    if any(x in texto for x in ["VIS VER", "VV", "VIS VERT", "VISIBILIDAD VERTICAL"]):
        if "///" in texto or "//" in texto:
            return "VV///"
        
        import re
        numeros = re.findall(r'\d+', texto)
        if numeros:
            altura_metros = int(numeros[0])
            altura_cientos = round(altura_metros / 30)
            altura_cientos = min(max(altura_cientos, 0), 999)
            return f"VV{altura_cientos:03d}"
    
    if vis_m >= 9999 and not fenomeno.strip():
        if not texto or texto in ["NSC", "SKC", "CLR", "DESPEJADO", "SIN NUBES", "NO NUBES"]:
            return "CAVOK"
    
    if not texto or texto in ["NSC", "SKC", "CLR", "DESPEJADO", "SIN NUBES", "NO NUBES"]:
        return "NSC"
    
    tipos_nubes = {
        "CU": "CU", "SC": "SC", "ST": "ST", "CB": "CB", "TCU": "TCU",
        "AC": "AC", "AS": "AS", "NS": "NS", "CI": "CI"
    }
    
    capas = texto.split(",")
    codigos_nubes = []
    
    for capa in capas[:4]:
        capa = capa.strip()
        if not capa:
            continue
        
        import re
        patron = r'(\d+)\s+([A-Z]{2,4})\s+(\d+)(?:M)?'
        match = re.search(patron, capa)
        
        if match:
            cantidad = int(match.group(1))
            tipo = match.group(2)
            altura = int(match.group(3))
            
            tipo_nube = tipos_nubes.get(tipo, tipo)
            
            if altura <= 3000:
                if altura % 30 != 0:
                    altura = (altura // 30) * 30
                altura_cientos = altura // 30
            else:
                if altura % 1000 != 0:
                    altura = (altura // 1000) * 1000
                altura_cientos = (altura // 1000) * 32
            
            altura_cientos = min(max(altura_cientos, 1), 999)
            
            if cantidad <= 2:
                cod_cant = "FEW"
            elif cantidad <= 4:
                cod_cant = "SCT"
            elif cantidad <= 7:
                cod_cant = "BKN"
            else:
                cod_cant = "OVC"
            
            codigo = f"{cod_cant}{altura_cientos:03d}"
            if tipo_nube in ["CB", "TCU"]:
                codigo += tipo_nube
            if codigo not in codigos_nubes:
                codigos_nubes.append(codigo)
    
    return " ".join(codigos_nubes[:4]) if codigos_nubes else "NSC"

# ============================================
# FUNCIONES DE VALIDACIÓN
# ============================================
def validar_hora(hora_str):
    if not hora_str:
        raise ValueError("Hora requerida - Formato HHMM")
    if len(hora_str) != 4 or not hora_str.isdigit():
        raise ValueError("Hora debe ser HHMM (4 dígitos)")
    h = int(hora_str[:2])
    m = int(hora_str[2:])
    if h > 23 or m > 59:
        raise ValueError("Hora inválida")
    return hora_str

def validar_intensidad_viento(intensidad_str):
    intensidad_str = str(intensidad_str).strip().upper()
    if not intensidad_str:
        raise ValueError("Intensidad de viento requerida")
    
    intensidad_str = intensidad_str.replace(' G ', 'G').replace(' G', 'G').replace('G ', 'G')
    
    if 'G' in intensidad_str:
        partes = intensidad_str.split('G')
        if len(partes) != 2:
            raise ValueError("Formato de ráfagas inválido. Use: 15G25")
        base = int(partes[0])
        rafaga = int(partes[1])
        if base < 0 or base > 100:
            raise ValueError("Intensidad base fuera de rango")
        if rafaga < base:
            raise ValueError("Ráfaga debe ser mayor o igual a intensidad base")
        return intensidad_str
    else:
        intensidad = int(intensidad_str)
        if intensidad < 0 or intensidad > 100:
            raise ValueError("Intensidad fuera de rango")
        return intensidad_str

def validar_numero(valor, min_val, max_val, nombre):
    if not valor:
        raise ValueError(f"{nombre} es obligatorio")
    try:
        num = float(valor)
        if not (min_val <= num <= max_val):
            raise ValueError(f"{nombre} fuera de rango")
        return num
    except:
        raise ValueError(f"{nombre} inválido")

def validar_temp_rocio(temp, rocio):
    if float(rocio) > float(temp):
        raise ValueError(f"Rocío ({rocio}°C) no puede ser > Temperatura ({temp}°C)")
    return True

# ============================================
# FUNCIÓN PRINCIPAL DE GENERACIÓN
# ============================================
def generar_metar(datos):
    try:
        if not datos['dir_viento'] or not datos['int_viento']:
            raise ValueError("Dirección e intensidad del viento son obligatorias")
        if not datos['vis']:
            raise ValueError("Visibilidad es obligatoria")
        if not datos['temp'] or not datos['rocio'] or not datos['qnh']:
            raise ValueError("Temperatura, Rocío y QNH son obligatorios")
        
        hora = validar_hora(datos['hora'])
        valido, msg = validar_hora_auditoria(hora)
        if not valido:
            raise ValueError(msg)
        
        int_viento = validar_intensidad_viento(datos['int_viento'])
        viento = procesar_viento(datos['dir_viento'], int_viento, datos['var_viento'])
        vis_m = convertir_visibilidad(datos['vis'])
        
        vis_min_codigo = ""
        if datos['vis_min']:
            vis_min_codigo, vis_min_error = procesar_visibilidad_minima(datos['vis_min'], vis_m)
            if vis_min_error:
                raise ValueError(vis_min_error)
        
        rvr_codigo = procesar_rvr(datos['rvr'])
        fenomeno = codificar_fenomenos(datos['fenomeno'], vis_m)
        nubes = interpretar_nubes(datos['nubes'], vis_m, fenomeno)
        
        temp = validar_numero(datos['temp'], -10, 40, "Temperatura")
        rocio = validar_numero(datos['rocio'], -10, 40, "Punto de rocío")
        validar_temp_rocio(temp, rocio)
        qnh = validar_numero(datos['qnh'], 850, 1100, "QNH")
        
        temp_metar = redondear_metar(temp)
        rocio_metar = redondear_metar(rocio)
        qnh_metar = int(qnh)
        
        metar_parts = [f"{datos['tipo']} SPJC {datos['dia']}{hora}Z {viento}"]
        
        if nubes == "CAVOK":
            metar_parts.append("CAVOK")
        else:
            metar_parts.append(f"{vis_m:04d}")
            if vis_min_codigo:
                metar_parts.append(vis_min_codigo)
            if rvr_codigo:
                metar_parts.append(rvr_codigo)
            if fenomeno:
                metar_parts.append(fenomeno)
            metar_parts.append(nubes)
        
        metar_parts.append(f"{temp_metar:02d}/{rocio_metar:02d} Q{qnh_metar}")
        
        if datos['suplementaria']:
            metar_parts.append(datos['suplementaria'].upper())
        
        metar_completo = " ".join(metar_parts) + "="
        
        hr_calculada = calcular_hr_automatica(temp, rocio)
        
        registro = {
            'Día': str(datos['dia']).zfill(2),
            'Hora': hora,
            'Tipo': datos['tipo'],
            'Dirección_Viento': datos['dir_viento'],
            'Intensidad_Viento': datos['int_viento'],
            'Variación_Viento': datos['var_viento'],
            'Visibilidad_Original': datos['vis'],
            'Visibilidad_Metros': vis_m,
            'Visibilidad_Mínima': vis_min_codigo,
            'RVR': rvr_codigo,
            'Fenómeno_Texto': datos['fenomeno'],
            'Fenómeno_Código': fenomeno,
            'Nubes_Texto': datos['nubes'],
            'Nubes_Código': "CAVOK" if nubes == "CAVOK" else nubes,
            'Temperatura': temp,
            'Punto_Rocío': rocio,
            'Humedad_Relativa_%': hr_calculada if hr_calculada else "",
            'QNH': qnh,
            'Presión_Estación': datos['presion'],
            'Info_Suplementaria': datos['suplementaria'],
            'METAR_Completo': metar_completo
        }
        
        return {'success': True, 'metar': metar_completo, 'registro': registro}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}

# ============================================
# FUNCIÓN PARA ACTUALIZAR O INSERTAR REGISTRO
# ============================================
def actualizar_o_insertar_registro(registros, nuevo_registro):
    """Actualiza o inserta un registro y guarda automáticamente"""
    dia_nuevo = str(nuevo_registro.get('Día', '')).zfill(2)
    hora_nueva = str(nuevo_registro.get('Hora', '')).zfill(4)
    
    dia_hora_clave = f"{dia_nuevo}_{hora_nueva}"
    accion = "insertado"
    
    for i, registro in enumerate(registros):
        dia_existente = str(registro.get('Día', '')).zfill(2)
        hora_existente = str(registro.get('Hora', '')).zfill(4)
        
        clave_existente = f"{dia_existente}_{hora_existente}"
        
        if clave_existente == dia_hora_clave:
            registros[i] = nuevo_registro
            accion = "actualizado"
            break
    else:
        registros.insert(0, nuevo_registro)
    
    guardar_registros_mes(registros)
    
    return accion

# ============================================
# FUNCIÓN PARA EXPORTAR EXCEL
# ============================================
def exportar_a_excel(registros):
    """Genera un archivo Excel con formato profesional"""
    if not registros:
        return None, "No hay registros para exportar"
    
    try:
        datos_exportar = []
        for r in registros:
            dia = str(r.get('Día', '')).zfill(2)
            hora = str(r.get('Hora', '')).zfill(4)
            
            datos_exportar.append({
                'DIA': dia,
                'HORA': hora,
                'TIPO': r.get('Tipo', ''),
                'DIR VIENTO': r.get('Dirección_Viento', ''),
                'INTENSIDAD': r.get('Intensidad_Viento', ''),
                'VARIACION': r.get('Variación_Viento', ''),
                'VIS (ORIGINAL)': r.get('Visibilidad_Original', ''),
                'VIS (CODIGO)': r.get('Visibilidad_Metros', ''),
                'VIS MIN': r.get('Visibilidad_Mínima', ''),
                'RVR': r.get('RVR', ''),
                'FENOMENO': r.get('Fenómeno_Texto', ''),
                'WX': r.get('Fenómeno_Código', ''),
                'NUBOSIDAD': r.get('Nubes_Texto', ''),
                'CLD': r.get('Nubes_Código', ''),
                'TEMP °C': r.get('Temperatura', ''),
                'ROCÍO °C': r.get('Punto_Rocío', ''),
                'HR %': r.get('Humedad_Relativa_%', ''),
                'QNH': r.get('QNH', ''),
                'PRESION': r.get('Presión_Estación', ''),
                'RMK': r.get('Info_Suplementaria', ''),
                'METAR': r.get('METAR_Completo', '')
            })
        
        df = pd.DataFrame(datos_exportar)
        df = df.sort_values(['DIA', 'HORA'], ascending=[True, True])
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='METAR SPJC', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['METAR SPJC']
            
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            for col in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col)
                col_name = df.columns[col-1]
                
                max_length = len(str(col_name))
                for row in range(2, min(len(df) + 2, 100)):
                    cell_value = df.iloc[row-2, col-1]
                    if cell_value:
                        cell_length = len(str(cell_value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                if col_name == 'METAR':
                    adjusted_width = min(max_length + 5, 120)
                elif col_name in ['TEMP °C', 'ROCÍO °C', 'HR %', 'QNH']:
                    adjusted_width = 12
                else:
                    adjusted_width = min(max_length + 2, 50)
                
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 8)
            
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='0B3D91', end_color='0B3D91', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            worksheet.row_dimensions[1].height = 30
            
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            for row in range(2, len(df) + 2):
                tipo_reporte = worksheet.cell(row=row, column=3).value
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    
                    if tipo_reporte == 'SPECI':
                        cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                        cell.font = Font(name='Calibri', size=10, bold=True)
                    else:
                        cell.font = Font(name='Calibri', size=10)
                    
                    col_name = df.columns[col-1]
                    if col_name == 'METAR':
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            worksheet.freeze_panes = 'A2'
        
        output.seek(0)
        return output, f"✅ {len(registros)} registros exportados"
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return None, f"Error al exportar: {str(e)}"

# ============================================
# FUNCIÓN PARA VERIFICAR Y MOSTRAR TN/TX
# ============================================
def verificar_y_mostrar_tn_tx():
    """
    Muestra campos de TN/TX cuando corresponde:
    - TN a las 12Z (Temperatura Mínima)
    - TX a las 22Z (Temperatura Máxima)
    """
    hora_actual = datetime.now(timezone.utc).strftime("%H%M")
    hora_int = int(hora_actual)
    
    mostrar_tn_tx = False
    mensaje = ""
    tipo_reporte = ""
    
    if 1200 <= hora_int < 1300:
        mostrar_tn_tx = True
        mensaje = "📉 REPORTE TN (Temperatura Mínima) - 12Z"
        tipo_reporte = "TN"
    elif 2200 <= hora_int < 2300:
        mostrar_tn_tx = True
        mensaje = "📈 REPORTE TX (Temperatura Máxima) - 22Z"
        tipo_reporte = "TX"
    
    if mostrar_tn_tx:
        st.markdown(f"### {mensaje}")
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        with col1:
            tn_tx_valor = st.text_input(
                "Valor (°C)",
                key="tn_tx_valor",
                placeholder="Ej: 18.5"
            )
        with col2:
            tn_tx_hora = st.text_input(
                "Hora de ocurrencia",
                key="tn_tx_hora",
                placeholder="HHMM",
                help="Hora en que ocurrió la temperatura (UTC)"
            )
        
        if tn_tx_valor and tn_tx_hora:
            st.session_state.tn_tx_completo = f"{tipo_reporte}{tn_tx_valor}/{tn_tx_hora}Z"
            st.success(f"✅ {st.session_state.tn_tx_completo}")
        else:
            st.warning(f"⚠️ Complete {tipo_reporte} antes de continuar")
            st.session_state.tn_tx_completo = None
    
    return mostrar_tn_tx

# ============================================
# FUNCIÓN PARA LIMPIAR CAMPOS
# ============================================
def limpiar_campos():
    """Limpia todos los campos del formulario"""
    st.session_state.dia = datetime.now(timezone.utc).strftime("%d")
    st.session_state.hora = datetime.now(timezone.utc).strftime("%H%M")
    st.session_state.tipo = TipoReporte.METAR.value
    st.session_state.tipo_selector = TipoReporte.METAR.value
    st.session_state.dir_viento = ""
    st.session_state.int_viento = ""
    st.session_state.var_desde = ""
    st.session_state.var_hasta = ""
    st.session_state.vis = ""
    st.session_state.vis_min = ""
    st.session_state.rvr = ""
    st.session_state.fenomenos_seleccionados = []
    st.session_state.capas_nubes = []
    st.session_state.vv_activo = False
    st.session_state.vv_valor = ""
    st.session_state.temp = ""
    st.session_state.rocio = ""
    st.session_state.qnh = ""
    st.session_state.presion = ""
    st.session_state.suplementaria = ""
    st.session_state.pp_select = ""
    st.session_state.tn_tx_valor = ""
    st.session_state.tn_tx_hora = ""
    st.session_state.hora_bloqueada = True
    st.session_state.campos_inicializados = True

# ============================================
# INICIALIZAR ESTADO DE SESIÓN
# ============================================
if 'registros' not in st.session_state:
    st.session_state.registros = cargar_registros_mes()

if 'historial' not in st.session_state:
    st.session_state.historial = []

if 'contador' not in st.session_state:
    st.session_state.contador = len(st.session_state.registros)

if 'campos_inicializados' not in st.session_state:
    st.session_state.campos_inicializados = False

if 'tipo' not in st.session_state:
    st.session_state.tipo = TipoReporte.METAR.value

if not st.session_state.campos_inicializados:
    st.session_state.dia = datetime.now(timezone.utc).strftime("%d")
    st.session_state.hora = datetime.now(timezone.utc).strftime("%H%M")
    st.session_state.tipo = TipoReporte.METAR.value
    st.session_state.tipo_selector = TipoReporte.METAR.value
    st.session_state.dir_viento = ""
    st.session_state.int_viento = ""
    st.session_state.var_desde = ""
    st.session_state.var_hasta = ""
    st.session_state.vis = ""
    st.session_state.vis_min = ""
    st.session_state.rvr = ""
    st.session_state.fenomenos_seleccionados = []
    st.session_state.capas_nubes = []
    st.session_state.vv_activo = False
    st.session_state.vv_valor = ""
    st.session_state.temp = ""
    st.session_state.rocio = ""
    st.session_state.qnh = ""
    st.session_state.presion = ""
    st.session_state.suplementaria = ""
    st.session_state.pp_select = ""
    st.session_state.tn_tx_valor = ""
    st.session_state.tn_tx_hora = ""
    st.session_state.hora_bloqueada = True
    st.session_state.campos_inicializados = True

# ============================================
# SIDEBAR - TABLAS DE REFERENCIA
# ============================================
with st.sidebar:
    st.markdown("### 📋 TABLAS DE REFERENCIA")
    
    with st.expander("🌧️ Precipitación (PPTRZ)"):
        st.markdown("""
        | Código | Significado |
        |--------|-------------|
        | **PPTRZ** | **Trazas (< 0.1 mm)** |
        | PP001 | 0.1 mm |
        | PP002 | 0.2 mm |
        | PP003 | 0.3 mm |
        | PP004 | 0.4 mm |
        | PP005 | 0.5 mm |
        | PP006 | 0.6 mm |
        | PP007 | 0.7 mm |
        | PP008 | 0.8 mm |
        | PP009 | 0.9 mm |
        | PP010 | 1.0 mm |
        """)
    
    with st.expander("🌡️ TN/TX"):
        st.markdown("""
        | Reporte | Hora | Significado |
        |---------|------|-------------|
        | TN | 12Z | Temperatura Mínima |
        | TX | 22Z | Temperatura Máxima |
        """)
    
    with st.expander("🌫️ Visibilidad Mínima"):
        st.markdown("""
        | Código | Significado |
        |--------|-------------|
        | 0800SW | 800m al Suroeste |
        | 1200NE | 1200m al Noreste |
        | 1500N | 1500m al Norte |
        """)
    
    with st.expander("🌀 RVR"):
        st.markdown("""
        | Código | Significado |
        |--------|-------------|
        | R32/0400 | Pista 32, 400m |
        | R12R/M0050 | Pista 12R, < 50m |
        | R14L/P2000 | Pista 14L, > 2000m |
        """)
    
    with st.expander("☁️ Códigos de Nubes"):
        st.markdown("""
        | Código | Significado |
        |--------|-------------|
        | FEW | 1-2 octas |
        | SCT | 3-4 octas |
        | BKN | 5-7 octas |
        | OVC | 8 octas |
        | VV | Visibilidad Vertical |
        | NSC | Sin nubes significativas |
        | CAVOK | Techo y visibilidad OK |
        """)

# ============================================
# HEADER PRINCIPAL
# ============================================
col_header1, col_header2 = st.columns([3, 1])

with col_header1:
    st.markdown("<h1 style='color: #0b3d91;'>REGISTRO DE OBSERVACIONES ORDINARIAS Y ESPECIALES</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color: #666;'>Aeropuerto Internacional Jorge Chávez - CORPAC Perú</p>", unsafe_allow_html=True)

with col_header2:
    hoy = datetime.now(timezone.utc).strftime('%d/%m/%Y')
    st.markdown(f"""
    <div style='text-align: right;'>
        <p style='color: #666; font-size: 14px;'>{hoy}</p>
        <p style='color: #0b3d91; font-size: 12px; font-weight: bold;'>{obtener_nombre_archivo_mensual()}</p>
        <p style='color: #999; font-size: 11px;'>CORPAC PERU</p>
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")

# ============================================
# INTERFAZ PRINCIPAL CON FRAGMENTS
# ============================================
col_izq, col_der = st.columns([2, 1])

with col_izq:
    # SELECTOR DE TIPO (con callback)
    with st.container():
        st.markdown("<div class='section-title'>TIPO DE REPORTE</div>", unsafe_allow_html=True)
        
        info_col1, info_col2 = st.columns([3, 1])
        with info_col1:
            tipo = st.selectbox(
                "Seleccione tipo de reporte",
                [t.value for t in TipoReporte],
                key='tipo_selector',
                on_change=on_tipo_change,
                help="METAR: hora bloqueada y automática | SPECI: hora modificable"
            )
        with info_col2:
            hora_actual = datetime.now(timezone.utc).strftime("%H%M")
            st.markdown(f"""
            <div style='background: #e7f3ff; padding: 10px; border-radius: 5px; text-align: center;'>
                <small>Hora UTC</small><br>
                <strong>{hora_actual}</strong>Z
            </div>
            """, unsafe_allow_html=True)
        
        if st.session_state.get('hora_bloqueada', False):
            st.info("🔒 **METAR**: Hora automática bloqueada")
        else:
            st.info("🔓 **SPECI**: Hora modificable (máx 15 min de diferencia)")
    
    st.markdown("---")
    
    # DATOS BÁSICOS
    col1, col2 = st.columns(2)
    with col1:
        dia = st.text_input(
            "Día", 
            key='dia', 
            placeholder="01-31",
            value=datetime.now(timezone.utc).strftime("%d"),
            help="Día del mes (01-31)"
        )
    with col2:
        hora = st.text_input(
            "Hora UTC", 
            key='hora',
            disabled=st.session_state.get('hora_bloqueada', False),
            help="Formato HHMM (ej: 1230)"
        )
        
        if not st.session_state.get('hora_bloqueada', False) and hora:
            try:
                valido, msg = validar_hora_auditoria(hora)
                if valido:
                    st.caption(f"✅ {msg}")
                else:
                    st.caption(f"❌ {msg}")
            except:
                pass
    
    st.markdown("---")
    
    # VIENTO
    st.markdown("<div class='section-title'>VIENTO</div>", unsafe_allow_html=True)
    col1, col2, col3, col4 = st.columns([2, 2, 1, 3])
    with col1:
        dir_viento = st.text_input("Dirección", key='dir_viento', placeholder="360")
    with col2:
        int_viento = st.text_input("Intensidad", key='int_viento', placeholder="15")
    with col3:
        st.markdown("<br><h3 style='text-align: center;'>-</h3>", unsafe_allow_html=True)
    with col4:
        col4_1, col4_2 = st.columns(2)
        with col4_1:
            var_desde = st.text_input("Var Desde", key='var_desde', placeholder="340")
        with col4_2:
            var_hasta = st.text_input("Var Hasta", key='var_hasta', placeholder="080")
    
    if var_desde and var_hasta:
        var_viento = f"{var_desde}V{var_hasta}"
    else:
        var_viento = ""
    
    st.markdown("---")
    
    # VISIBILIDAD
    st.markdown("<div class='section-title'>VISIBILIDAD</div>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        vis = st.text_input("Visibilidad", key='vis', placeholder="10km, 5000m, 9999")
    with col2:
        vis_min = st.text_input("Visibilidad Mínima", key='vis_min', placeholder="1200SW")
    with col3:
        rvr = st.text_input("RVR", key='rvr', placeholder="R32/0400")
    
    st.markdown("---")
    
    # FENÓMENOS - FRAGMENT
    st.markdown("<div class='section-title'>FENÓMENOS</div>", unsafe_allow_html=True)
    fragment_fenomenos()
    
    # Obtener el texto de fenómenos del session_state
    fenomeno = " ".join([f.split(" - ")[0] for f in st.session_state.get('fenomenos_seleccionados', [])])
    
    st.markdown("---")
    
    # NUBES - FRAGMENT
    st.markdown("<div class='section-title'>NUBOSIDAD</div>", unsafe_allow_html=True)
    fragment_nubes()
    
    # Obtener el código de nubes
    nubes = convertir_nubes_completo_a_metar()
    
    st.markdown("---")
    
    # TEMPERATURA Y PRESIÓN
    st.markdown("<div class='section-title'>TEMPERATURA Y PRESIÓN</div>", unsafe_allow_html=True)
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        temp = st.text_input("Temp °C", key='temp', placeholder="-10 a 40")
    with col2:
        rocio = st.text_input("Rocío °C", key='rocio', placeholder="≤ Temp")
    with col3:
        qnh = st.text_input("QNH hPa", key='qnh', placeholder="850-1100")
    with col4:
        presion = st.text_input("Presión Est.", key='presion', placeholder="Opcional")
    
    if temp and rocio:
        try:
            hr_calculada = calcular_hr_automatica(float(temp), float(rocio))
            if hr_calculada:
                st.caption(f"💧 HR Calculada: {hr_calculada}% (referencia)")
        except:
            pass
    
    st.markdown("---")
    
    # PRECIPITACIÓN
    st.markdown("<div class='section-title'>PRECIPITACIÓN (OBLIGATORIO)</div>", unsafe_allow_html=True)
    pp_valor = st.selectbox(
        "Cantidad de precipitación",
        options=[""] + list(OPCIONES_PP.keys()),
        format_func=lambda x: f"{x} - {OPCIONES_PP.get(x, 'Seleccione...')}" if x else "Seleccione cantidad...",
        key="pp_select",
        help="PPTRZ = Trazas (<0.1mm), PP001 = 0.1mm, ... PP010 = 1.0mm"
    )
    
    st.markdown("---")
    
    # TN/TX
    mostrar_tn_tx = verificar_y_mostrar_tn_tx()
    
    if mostrar_tn_tx:
        st.markdown("---")
    
    # INFORMACIÓN SUPLEMENTARIA
    st.markdown("<div class='section-title'>INFORMACIÓN SUPLEMENTARIA</div>", unsafe_allow_html=True)
    suplementaria = st.text_input("", key='suplementaria', placeholder="NOSIG RMK CB AL NE")
    
    st.markdown("---")
    
    # BOTONES
    col1, col2 = st.columns(2)
    with col1:
        generar = st.button("GENERAR METAR", use_container_width=True, type="primary")
    with col2:
        limpiar = st.button("LIMPIAR CAMPOS", use_container_width=True)
    
    if limpiar:
        limpiar_campos()
        st.rerun()
    
    if generar:
        # Validaciones finales
        if not pp_valor:
            st.error("❌ El campo de precipitación (PP) es OBLIGATORIO")
        elif mostrar_tn_tx and not st.session_state.get('tn_tx_completo'):
            st.error("❌ Complete TN/TX antes de continuar")
        else:
            # Construir RMK completa
            rmk_completa = suplementaria
            if pp_valor:
                rmk_completa = f"{pp_valor} {rmk_completa}".strip()
            if mostrar_tn_tx and st.session_state.get('tn_tx_completo'):
                rmk_completa = f"{st.session_state.tn_tx_completo} {rmk_completa}".strip()
            
            datos = {
                'tipo': st.session_state.tipo, 'dia': dia, 'hora': hora,
                'dir_viento': dir_viento, 'int_viento': int_viento, 'var_viento': var_viento,
                'vis': vis, 'vis_min': vis_min, 'rvr': rvr,
                'fenomeno': fenomeno, 'nubes': nubes,
                'temp': temp, 'rocio': rocio, 'hr': None,
                'qnh': qnh, 'presion': presion, 'suplementaria': rmk_completa
            }
            
            resultado = generar_metar(datos)
            
            if resultado['success']:
                accion = actualizar_o_insertar_registro(st.session_state.registros, resultado['registro'])
                
                dia_hora_clave = f"{str(resultado['registro']['Día']).zfill(2)}_{resultado['registro']['Hora']}"
                nuevo_historial = []
                
                for metar in st.session_state.historial:
                    match = re.search(r'SPJC (\d{2})(\d{4})Z', metar)
                    if match:
                        dia_hist = match.group(1)
                        hora_hist = match.group(2)
                        if f"{dia_hist}_{hora_hist}" != dia_hora_clave:
                            nuevo_historial.append(metar)
                    else:
                        nuevo_historial.append(metar)
                
                nuevo_historial.insert(0, resultado['metar'])
                st.session_state.historial = nuevo_historial[:20]
                st.session_state.contador = len(st.session_state.registros)
                
                hora_reporte = resultado['registro']['Hora']
                if accion == "actualizado":
                    st.warning(f"🔄 Reporte de las {hora_reporte}Z ACTUALIZADO (reemplazó uno existente)")
                else:
                    st.success(f"✅ Reporte de las {hora_reporte}Z agregado correctamente")
                
                st.session_state.ultimo_metar = resultado['metar']
                st.session_state.ultimo_tipo = st.session_state.tipo
                st.session_state.ultimo_registro = resultado['registro']
            else:
                st.error(f"❌ ERROR: {resultado['error']}")

with col_der:
    st.markdown("<div class='section-title'>📋 ÚLTIMO REPORTE</div>", unsafe_allow_html=True)
    if 'ultimo_metar' in st.session_state:
        tipo_ultimo = st.session_state.get('ultimo_tipo', 'METAR')
        if tipo_ultimo == "SPECI":
            st.markdown(f"<div style='background: #FFE699; padding: 15px; border-radius: 5px; font-family: monospace; border-left: 5px solid #FFC000;'>{st.session_state.ultimo_metar}</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div class='metar-box'>{st.session_state.ultimo_metar}</div>", unsafe_allow_html=True)
    else:
        st.info("---")
    
    st.markdown("---")
    st.markdown("<div class='section-title'>EXPORTAR</div>", unsafe_allow_html=True)
    
    if st.button("📥 Exportar METAR", use_container_width=True, type="primary"):
        if st.session_state.registros:
            excel_file, mensaje = exportar_a_excel(st.session_state.registros)
            if excel_file:
                nombre_archivo = obtener_nombre_archivo_mensual()
                st.download_button(
                    label="✅ Confirmar Descarga",
                    data=excel_file,
                    file_name=nombre_archivo,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.success(mensaje)
            else:
                st.warning(mensaje)
        else:
            st.warning("No hay registros")
    
    if st.button("🗑️ Limpiar Memoria", use_container_width=True):
        st.session_state.registros = []
        st.session_state.historial = []
        st.session_state.contador = 0
        if 'ultimo_metar' in st.session_state:
            del st.session_state.ultimo_metar
        if 'ultimo_tipo' in st.session_state:
            del st.session_state.ultimo_tipo
        if 'ultimo_registro' in st.session_state:
            del st.session_state.ultimo_registro
        st.success("Memoria limpiada")
        st.rerun()
    
    st.markdown("---")
    st.metric("REGISTROS EN MEMORIA", st.session_state.contador)
    
    st.markdown("---")
    st.markdown("<div class='section-title'>HISTORIAL</div>", unsafe_allow_html=True)
    if st.session_state.historial:
        for metar in st.session_state.historial[:10]:
            if "SPECI" in metar:
                st.markdown(f"<div style='background: #FFE699; padding: 8px; margin-bottom: 5px; border-radius: 3px; font-family: monospace; font-size: 12px; border-left: 3px solid #FFC000;'>{metar}</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='background: #f0f0f0; padding: 8px; margin-bottom: 5px; border-radius: 3px; font-family: monospace; font-size: 12px; border-left: 3px solid #0b3d91;'>{metar}</div>", unsafe_allow_html=True)
    else:
        st.info("No hay METARs en el historial")

# ============================================
# FOOTER
# ============================================
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666; padding: 20px;'>
    <p>METAR Digital - CORPAC Peru | Aeropuerto Internacional Jorge Chavez (SPJC)</p>
    <p style='font-size: 0.8rem;'>Sistema de Registro de Observaciones - Versión 14.0 con Fragments</p>
</div>
""", unsafe_allow_html=True)