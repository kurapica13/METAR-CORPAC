"""
METAR DIGITAL WEB - VERSIÓN PROFESIONAL CORPAC PERÚ
Características:
✅ Viento con reglas circulares (340V080)
✅ Visibilidad mínima con cuadrantes (N, NE, E, SE, S, SW, W, NW)
✅ RVR (Runway Visual Range)
✅ Estándar oficial nubes CORPAC (30m/1000m)
✅ Fenómenos especiales (PRFG, VCFG, BCFG, MIFG)
✅ Excel con formato profesional
✅ SIN DUPLICADOS: Reemplaza reportes con misma fecha/hora
"""

import streamlit as st
from datetime import datetime, timezone
import pandas as pd
from pathlib import Path
import re
import time
import os
import base64
from io import BytesIO

# ============================================
# CONFIGURACIÓN DE PÁGINA
# ============================================
st.set_page_config(
    page_title="METAR Digital - CORPAC Perú",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ============================================
# ESTILOS CSS PERSONALIZADOS
# ============================================
st.markdown("""
<style>
    .stApp {
        background-color: #f0f0f0;
    }
    
    .metar-header {
        background: linear-gradient(90deg, #0b3d91 0%, #1a4fa0 100%);
        padding: 1.5rem;
        border-radius: 10px;
        color: white;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    
    .panel {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        margin-bottom: 1rem;
    }
    
    .section-title {
        color: #0b3d91;
        font-weight: 600;
        margin-bottom: 1rem;
        border-bottom: 2px solid #e8eef7;
        padding-bottom: 0.5rem;
    }
    
    .metar-box {
        background: #1e1e1e;
        color: #00ff00;
        padding: 1rem;
        border-radius: 5px;
        font-family: 'Courier New', monospace;
        font-size: 1.1rem;
        border-left: 5px solid #0b3d91;
    }
    
    .historial-item {
        background: #f8f9fa;
        padding: 0.8rem;
        border-radius: 5px;
        margin-bottom: 0.5rem;
        font-family: 'Courier New', monospace;
        border-left: 3px solid #0b3d91;
    }
    
    .historial-item-speci {
        background: #FFE699;
        border-left: 3px solid #FFC000;
    }
    
    .badge-speci {
        background: #FFE699;
        color: #000;
        padding: 2px 8px;
        border-radius: 12px;
        font-size: 0.8rem;
        font-weight: bold;
        margin-left: 10px;
    }
    
    .stButton button {
        width: 100%;
        border-radius: 5px;
        font-weight: 600;
    }
    
    .stTextInput input, .stSelectbox select {
        border-radius: 5px;
        border: 1px solid #ddd;
    }
</style>
""", unsafe_allow_html=True)

# ============================================
# INICIALIZAR ESTADO DE SESIÓN
# ============================================
if 'registros' not in st.session_state:
    st.session_state.registros = []
if 'historial' not in st.session_state:
    st.session_state.historial = []
if 'contador' not in st.session_state:
    st.session_state.contador = 0
if 'campos_inicializados' not in st.session_state:
    st.session_state.campos_inicializados = False

# ============================================
# FUNCIÓN PARA LIMPIAR CAMPOS
# ============================================
def limpiar_campos():
    """Limpia todos los campos del formulario"""
    st.session_state.dia = datetime.now(timezone.utc).strftime("%d")
    st.session_state.hora = datetime.now(timezone.utc).strftime("%H%M")
    st.session_state.tipo = "METAR"
    st.session_state.dir_viento = ""
    st.session_state.int_viento = ""
    st.session_state.var_viento = ""
    st.session_state.vis = ""
    st.session_state.vis_min = ""
    st.session_state.rvr = ""
    st.session_state.fenomeno = ""
    st.session_state.nubes = ""
    st.session_state.temp = ""
    st.session_state.rocio = ""
    st.session_state.hr = ""
    st.session_state.qnh = ""
    st.session_state.presion = ""
    st.session_state.suplementaria = ""
    st.session_state.campos_inicializados = True

# ============================================
# INICIALIZAR CAMPOS VACÍOS
# ============================================
if not st.session_state.campos_inicializados:
    limpiar_campos()

# ============================================
# FUNCIÓN PARA ACTUALIZAR O INSERTAR REGISTRO (EVITAR DUPLICADOS)
# ============================================
def actualizar_o_insertar_registro(registros, nuevo_registro):
    """
    Busca un registro existente con el mismo día y hora.
    Si existe, lo reemplaza. Si no existe, lo agrega al inicio.
    """
    dia_hora_clave = f"{nuevo_registro['Día']}_{nuevo_registro['Hora']}"
    
    # Buscar índice del registro con mismo día y hora
    indice_existente = -1
    for i, registro in enumerate(registros):
        clave_existente = f"{registro.get('Día', '')}_{registro.get('Hora', '')}"
        if clave_existente == dia_hora_clave:
            indice_existente = i
            break
    
    # Si existe, reemplazarlo; si no, insertar al inicio
    if indice_existente != -1:
        registros[indice_existente] = nuevo_registro
        return "actualizado"
    else:
        registros.insert(0, nuevo_registro)
        return "insertado"

# ============================================
# FUNCIONES DE PROCESAMIENTO - VIENTO
# ============================================

def procesar_viento(direccion, intensidad, variacion):
    """
    PROCESAMIENTO DE VIENTO - REGLAS CORPAC PERÚ
    Caso 1: Variación ≥60° y <180° con viento <3kt → VRBxxKT
    Caso 2: Variación ≥60° y <180° con viento ≥3kt → dddffKT bbbVnnn
    Caso 3: Variación ≥180° → VRBxxKT (sin importar velocidad)
    """
    dir_int = int(direccion)
    intensidad_str = str(intensidad).upper().strip()
    
    # Procesar ráfagas
    if 'G' in intensidad_str:
        if 'G' in intensidad_str and not ' ' in intensidad_str.replace('G', ''):
            base_int, gust_int = intensidad_str.split('G')
            int_base = int(base_int)
            int_gust = int(gust_int)
            intensidad_metar = f"{int_base:02d}G{int_gust:02d}"
        else:
            parts = intensidad_str.replace('G', ' ').split()
            int_base = int(parts[0])
            int_gust = int(parts[1])
            intensidad_metar = f"{int_base:02d}G{int_gust:02d}"
    else:
        int_base = int(intensidad_str)
        intensidad_metar = f"{int_base:02d}"
    
    # Si NO hay variación
    if not variacion:
        return f"{dir_int:03d}{intensidad_metar}KT"
    
    try:
        # Extraer valores de variación (formato: bbbVnnn)
        variacion = variacion.upper().replace(' ', '')
        if 'V' not in variacion:
            return f"{dir_int:03d}{intensidad_metar}KT"
        
        desde, hasta = map(int, variacion.split('V'))
        
        # Calcular diferencia CIRCULAR (importante para 340V080)
        diff1 = abs(hasta - desde)
        diff2 = 360 - diff1
        diferencia = min(diff1, diff2)
        
        # CASO 3: Variación ≥ 180°
        if diferencia >= 180:
            return f"VRB{intensidad_metar}KT"
        
        # CASO 1 y 2: Variación ≥ 60° y < 180°
        if diferencia >= 60:
            if int_base < 3:
                # CASO 1: Viento < 3kt
                return f"VRB{intensidad_metar}KT"
            else:
                # CASO 2: Viento ≥ 3kt
                # Formatear para mostrar el rango correcto (siempre de menor a mayor)
                if diff1 <= 180:
                    return f"{dir_int:03d}{intensidad_metar}KT {desde:03d}V{hasta:03d}"
                else:
                    # Para casos como 340V080, mostrar como 080V340 (menor a mayor)
                    return f"{dir_int:03d}{intensidad_metar}KT {hasta:03d}V{desde:03d}"
        
        # Variación < 60°
        return f"{dir_int:03d}{intensidad_metar}KT {variacion}"
        
    except Exception as e:
        return f"{dir_int:03d}{intensidad_metar}KT"

# ============================================
# FUNCIONES DE PROCESAMIENTO - VISIBILIDAD
# ============================================

def convertir_visibilidad(vis_texto):
    """Convierte visibilidad a metros"""
    vis_texto = vis_texto.strip().upper()
    if not vis_texto:
        raise ValueError("Visibilidad es obligatoria")
    
    try:
        if vis_texto.endswith("KM"):
            km = float(vis_texto[:-2])
            return 9999 if km >= 10 else int(km * 1000)
        elif vis_texto.endswith("M"):
            return int(vis_texto[:-1])
        else:
            metros = int(vis_texto)
            return 9999 if metros >= 10000 else metros
    except:
        raise ValueError("Formato de visibilidad inválido")

def procesar_visibilidad_minima(vis_min_texto, vis_m):
    """
    Procesa visibilidad mínima con cuadrantes
    Cuadrantes: N, NE, E, SE, S, SW, W, NW
    Reglas: 1) <1500m 2) <50% vis reinante y <5000m
    """
    if not vis_min_texto:
        return "", ""
    
    vis_min_texto = vis_min_texto.strip().upper()
    cuadrantes = ['N', 'NE', 'E', 'SE', 'S', 'SW', 'W', 'NW']
    
    valor = ""
    cuadrante = ""
    
    for c in cuadrantes:
        if vis_min_texto.endswith(c):
            valor = vis_min_texto[:-len(c)]
            cuadrante = c
            break
    
    if not cuadrante:
        valor = vis_min_texto
    
    try:
        if valor.endswith("KM"):
            km = float(valor[:-2])
            vis_min_m = 9999 if km >= 10 else int(km * 1000)
        elif valor.endswith("M"):
            vis_min_m = int(valor[:-1])
        else:
            vis_min_m = int(valor)
            vis_min_m = 9999 if vis_min_m >= 10000 else vis_min_m
        
        es_valida = False
        if vis_min_m < 1500:
            es_valida = True
        if vis_min_m < (vis_m * 0.5) and vis_min_m < 5000:
            es_valida = True
        
        if not es_valida:
            return "", "⚠️ No cumple reglas de visibilidad mínima"
        
        if cuadrante:
            return f"{vis_min_m:04d}{cuadrante}", ""
        else:
            return f"{vis_min_m:04d}", ""
        
    except:
        return "", "❌ Formato inválido"

def procesar_rvr(rvr_texto):
    """Procesa RVR (Runway Visual Range)"""
    if not rvr_texto:
        return ""
    
    rvr_texto = rvr_texto.strip().upper().replace('M', '').replace('RVR', '')
    
    try:
        rvr_valor = int(rvr_texto)
        if rvr_valor < 50 or rvr_valor > 2000:
            return ""
        return f"RVR{rvr_valor:04d}"
    except:
        return ""

# ============================================
# FUNCIONES DE PROCESAMIENTO - FENÓMENOS
# ============================================

def codificar_fenomenos(texto):
    """CODIFICADOR COMPLETO DE FENÓMENOS - CORPAC"""
    if not texto:
        return ""
    
    texto_lower = texto.lower().strip()
    
    if any(x in texto_lower for x in ["niebla parcial", "prfg", "pr fg", "parcial"]):
        return "PRFG"
    if any(x in texto_lower for x in ["niebla en la vecindad", "vcfg", "vc fg", "vecindad"]):
        return "VCFG"
    if any(x in texto_lower for x in ["niebla en bancos", "bcfg", "bc fg", "bancos"]):
        return "BCFG"
    if any(x in texto_lower for x in ["niebla baja", "mifg", "mi fg", "baja"]):
        return "MIFG"
    
    intensidades = {"ligera": "-", "ligero": "-", "leve": "-", "débil": "-",
                   "fuerte": "+", "intensa": "+", "intenso": "+", "severa": "+"}
    
    descriptores = {"sh": "SH", "chubasco": "SH", "ts": "TS", "tormenta": "TS",
                   "fz": "FZ", "helada": "FZ", "congelante": "FZ"}
    
    fenomenos = {
        "lluvia": "RA", "llovizna": "DZ", "niebla": "FG", "neblina": "BR",
        "nieve": "SN", "granizo": "GR", "cellisca": "GS", "tormenta": "TS",
        "polvo": "DU", "arena": "SA", "humo": "FU", "ceniza": "VA", "calima": "HZ"
    }
    
    partes = [p.strip() for p in texto.split(",")][:3]
    resultados = []
    
    for parte in partes:
        if not parte:
            continue
        parte_lower = parte.lower()
        codigo = ""
        descriptor = ""
        intensidad = ""
        
        for d_texto, d_codigo in descriptores.items():
            if d_texto in parte_lower:
                descriptor = d_codigo
                parte_lower = parte_lower.replace(d_texto, "").strip()
                break
        
        for i_texto, i_codigo in intensidades.items():
            if i_texto in parte_lower:
                intensidad = i_codigo
                parte_lower = parte_lower.replace(i_texto, "").strip()
                break
        
        for f_texto, f_codigo in fenomenos.items():
            if f_texto in parte_lower:
                codigo = f_codigo
                break
        
        if codigo:
            if descriptor:
                codigo = descriptor + codigo
            if intensidad:
                codigo = intensidad + codigo
            resultados.append(codigo)
    
    return " ".join(resultados) if resultados else ""

# ============================================
# FUNCIONES DE PROCESAMIENTO - NUBES
# ============================================

def interpretar_nubes(texto, vis_m, fenomeno):
    """CODIFICADOR DE NUBES - ESTÁNDAR CORPAC PERÚ"""
    texto = texto.strip().upper()
    
    if texto in ["DESPEJADO", "SKC", "CLR", "", "NSC", "SIN NUBES", "NO NUBES"]:
        return "NSC"
    
    if vis_m >= 9999 and not fenomeno.strip() and texto in ["NSC", "SKC", "CLR", "DESPEJADO"]:
        return "CAVOK"
    
    tipos_nubes = {
        "CU": "CU", "SC": "SC", "ST": "ST", "CB": "CB", "TCU": "TCU",
        "AC": "AC", "AS": "AS", "NS": "NS", "CI": "CI"
    }
    
    capas = texto.split(",")
    codigos_nubes = []
    
    for capa in capas[:4]:
        capa = capa.strip()
        if not capa:
            continue
        
        patron = r'(\d+)\s+([A-Z]{2,4})\s+(\d+)(?:M)?'
        match = re.search(patron, capa)
        
        if match:
            cantidad = int(match.group(1))
            tipo = match.group(2)
            altura = int(match.group(3))
            
            tipo_nube = tipos_nubes.get(tipo, tipo)
            
            if altura <= 3000:
                if altura % 30 != 0:
                    altura = (altura // 30) * 30
                altura_cientos = altura // 30
            else:
                if altura % 1000 != 0:
                    altura = (altura // 1000) * 1000
                altura_cientos = (altura // 1000) * 32
            
            altura_cientos = min(max(altura_cientos, 1), 999)
            
            if cantidad <= 2:
                cod_cant = "FEW"
            elif cantidad <= 4:
                cod_cant = "SCT"
            elif cantidad <= 7:
                cod_cant = "BKN"
            else:
                cod_cant = "OVC"
            
            codigo = f"{cod_cant}{altura_cientos:03d}"
            if tipo_nube in ["CB", "TCU"]:
                codigo += tipo_nube
            codigos_nubes.append(codigo)
        
        else:
            cb_match = re.search(r'CB\s+(\d+)(?:M)?', capa)
            if cb_match:
                altura = int(cb_match.group(1))
                if altura <= 3000:
                    altura = (altura // 30) * 30
                    codigos_nubes.append(f"BKN{(altura//30):03d}CB")
                else:
                    altura = (altura // 1000) * 1000
                    codigos_nubes.append(f"BKN{(altura//1000)*32:03d}CB")
            
            tcu_match = re.search(r'TCU\s+(\d+)(?:M)?', capa)
            if tcu_match:
                altura = int(tcu_match.group(1))
                if altura <= 3000:
                    altura = (altura // 30) * 30
                    codigos_nubes.append(f"BKN{(altura//30):03d}TCU")
                else:
                    altura = (altura // 1000) * 1000
                    codigos_nubes.append(f"BKN{(altura//1000)*32:03d}TCU")
    
    return " ".join(codigos_nubes[:4]) if codigos_nubes else "NSC"

def verificar_cavok(vis_m, fenomeno, nubes):
    """Verifica condiciones para CAVOK"""
    return (vis_m >= 9999 and not fenomeno.strip() and nubes in ["NSC", "SKC", "CLR"])

# ============================================
# FUNCIONES DE VALIDACIÓN
# ============================================

def validar_hora(hora_str):
    """Valida formato de hora HHMM"""
    if len(hora_str) != 4 or not hora_str.isdigit():
        raise ValueError("Hora debe ser HHMM (4 dígitos)")
    h = int(hora_str[:2])
    m = int(hora_str[2:])
    if h > 23 or m > 59:
        raise ValueError("Hora inválida")
    return hora_str

def validar_intensidad_viento(intensidad_str):
    """Valida formato de intensidad de viento"""
    intensidad_str = str(intensidad_str).strip().upper()
    if not intensidad_str:
        raise ValueError("Intensidad de viento requerida")
    
    intensidad_str = intensidad_str.replace(' G ', 'G').replace(' G', 'G').replace('G ', 'G')
    
    if 'G' in intensidad_str:
        partes = intensidad_str.split('G')
        if len(partes) != 2:
            raise ValueError("Formato de ráfagas inválido. Use: 15G25")
        base = int(partes[0])
        rafaga = int(partes[1])
        if base < 0 or base > 100:
            raise ValueError("Intensidad base fuera de rango (0-100)")
        if rafaga < base:
            raise ValueError("Ráfaga debe ser mayor o igual a intensidad base")
        if rafaga > 150:
            raise ValueError("Ráfaga excede límite (150 KT)")
        diferencia = rafaga - base
        if diferencia < 10:
            raise ValueError(f"Ráfaga requiere diferencia ≥10 KT (actual: {diferencia} KT)")
        return intensidad_str
    else:
        intensidad = int(intensidad_str)
        if intensidad < 0 or intensidad > 100:
            raise ValueError("Intensidad fuera de rango (0-100)")
        return intensidad_str

def validar_numero(valor, min_val, max_val, nombre):
    """Valida un número dentro de un rango"""
    if not valor:
        raise ValueError(f"{nombre} es obligatorio")
    try:
        num = float(valor)
        if not (min_val <= num <= max_val):
            raise ValueError(f"{nombre} fuera de rango ({min_val}-{max_val})")
        return num
    except ValueError as e:
        raise ValueError(f"{nombre} inválido: {str(e)}")

def validar_temp_rocio(temp, rocio):
    """Valida que rocío ≤ temperatura"""
    if float(rocio) > float(temp):
        raise ValueError(f"Rocío ({rocio}°C) no puede ser > Temperatura ({temp}°C)")
    return True

def validar_humedad(hr):
    """Valida humedad relativa"""
    if not hr:
        return ""
    try:
        num = float(hr)
        if num < 0 or num > 100:
            raise ValueError("Humedad fuera de rango (0-100%)")
        return num
    except:
        raise ValueError("Humedad inválida")

# ============================================
# FUNCIÓN PRINCIPAL DE GENERACIÓN
# ============================================

def generar_metar(datos):
    """Genera código METAR desde los datos del formulario"""
    try:
        # Validar campos obligatorios
        if not datos['dir_viento'] or not datos['int_viento']:
            raise ValueError("Dirección e intensidad del viento son obligatorias")
        if not datos['vis']:
            raise ValueError("Visibilidad es obligatoria")
        if not datos['temp'] or not datos['rocio'] or not datos['qnh']:
            raise ValueError("Temperatura, Rocío y QNH son obligatorios")
        
        # Validar hora
        hora = validar_hora(datos['hora'])
        
        # Procesar viento
        int_viento = validar_intensidad_viento(datos['int_viento'])
        viento = procesar_viento(datos['dir_viento'], int_viento, datos['var_viento'])
        
        # Procesar visibilidad
        vis_m = convertir_visibilidad(datos['vis'])
        
        # Procesar visibilidad mínima con cuadrante
        vis_min_codigo = ""
        if datos['vis_min']:
            vis_min_codigo, vis_min_error = procesar_visibilidad_minima(datos['vis_min'], vis_m)
            if vis_min_error:
                raise ValueError(vis_min_error)
        
        # Procesar RVR
        rvr_codigo = procesar_rvr(datos['rvr'])
        
        # Procesar fenómenos y nubes
        fenomeno = codificar_fenomenos(datos['fenomeno'])
        nubes = interpretar_nubes(datos['nubes'], vis_m, fenomeno)
        
        # Validar temperaturas
        temp = validar_numero(datos['temp'], -10, 40, "Temperatura")
        rocio = validar_numero(datos['rocio'], -10, 40, "Punto de rocío")
        validar_temp_rocio(temp, rocio)
        
        # Validar humedad
        hr = validar_humedad(datos['hr'])
        
        # Validar QNH
        qnh = validar_numero(datos['qnh'], 850, 1100, "QNH")
        
        # Verificar CAVOK
        es_cavok = verificar_cavok(vis_m, fenomeno, nubes)
        
        # Construir METAR
        metar_parts = [f"{datos['tipo']} SPJC {datos['dia']}{hora}Z {viento}"]
        
        if es_cavok:
            metar_parts.append("CAVOK")
        else:
            metar_parts.append(f"{vis_m:04d}")
            if vis_min_codigo:
                metar_parts.append(vis_min_codigo)
            if rvr_codigo:
                metar_parts.append(rvr_codigo)
            if fenomeno:
                metar_parts.append(fenomeno)
            metar_parts.append(nubes)
        
        metar_parts.append(f"{round(temp):02d}/{round(rocio):02d} Q{round(qnh)}")
        
        if datos['suplementaria']:
            metar_parts.append(datos['suplementaria'].upper())
        
        metar_completo = " ".join(metar_parts) + "="
        
        # Crear registro completo
        registro = {
            'Día': datos['dia'],
            'Hora': hora,
            'Tipo': datos['tipo'],
            'Dirección_Viento': datos['dir_viento'],
            'Intensidad_Viento': datos['int_viento'],
            'Variación_Viento': datos['var_viento'],
            'Visibilidad_Original': datos['vis'],
            'Visibilidad_Metros': vis_m,
            'Visibilidad_Mínima': vis_min_codigo,
            'RVR': rvr_codigo,
            'Fenómeno_Texto': datos['fenomeno'],
            'Fenómeno_Código': fenomeno,
            'Nubes_Texto': datos['nubes'],
            'Nubes_Código': "CAVOK" if es_cavok else nubes,
            'Temperatura': temp,
            'Punto_Rocío': rocio,
            'Humedad_Relativa_%': hr if hr else "",
            'QNH': qnh,
            'Presión_Estación': datos['presion'],
            'Info_Suplementaria': datos['suplementaria'],
            'METAR_Completo': metar_completo
        }
        
        return {
            'success': True,
            'metar': metar_completo,
            'registro': registro
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

# ============================================
# FUNCIÓN PARA EXPORTAR EXCEL
# ============================================

def exportar_a_excel(registros):
    """Exporta registros a Excel con formato profesional"""
    if not registros:
        return None, "No hay registros para exportar"
    
    try:
        df = pd.DataFrame(registros)
        
        columnas = [
            'Día', 'Hora', 'Tipo', 'Dirección_Viento', 'Intensidad_Viento',
            'Variación_Viento', 'Visibilidad_Original', 'Visibilidad_Metros',
            'Visibilidad_Mínima', 'RVR', 'Fenómeno_Texto', 'Fenómeno_Código',
            'Nubes_Texto', 'Nubes_Código', 'Temperatura', 'Punto_Rocío',
            'Humedad_Relativa_%', 'QNH', 'Presión_Estación',
            'Info_Suplementaria', 'METAR_Completo'
        ]
        
        for col in columnas:
            if col not in df.columns:
                df[col] = ""
        
        df = df[columnas]
        df['Día'] = df['Día'].astype(str).str.zfill(2)
        df['Hora'] = df['Hora'].astype(str).str.zfill(4)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='METAR', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['METAR']
            
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Ancho automático de columnas
            for col in range(1, len(columnas) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                header_cell = worksheet.cell(row=1, column=col)
                max_length = len(str(header_cell.value)) if header_cell.value else 0
                
                for row in range(2, min(len(df) + 2, 102)):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = min(max_length + 2, 70)
                adjusted_width = max(adjusted_width, 8)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Formato profesional
            COLOR_HEADER = "0B3D91"
            COLOR_HEADER_TEXTO = "FFFFFF"
            COLOR_FILA_IMPAR = "E8EEF7"
            COLOR_BORDE = "CCCCCC"
            COLOR_SPECI = "FFE699"
            
            header_font = Font(name='Calibri', size=11, bold=True, color=COLOR_HEADER_TEXTO)
            header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_border = Border(
                left=Side(style='medium', color=COLOR_HEADER),
                right=Side(style='medium', color=COLOR_HEADER),
                top=Side(style='medium', color=COLOR_HEADER),
                bottom=Side(style='medium', color=COLOR_HEADER)
            )
            
            for col in range(1, len(columnas) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = header_border
            
            worksheet.row_dimensions[1].height = 30
            
            border_datos = Border(
                left=Side(style='thin', color=COLOR_BORDE),
                right=Side(style='thin', color=COLOR_BORDE),
                top=Side(style='thin', color=COLOR_BORDE),
                bottom=Side(style='thin', color=COLOR_BORDE)
            )
            
            fill_impar = PatternFill(start_color=COLOR_FILA_IMPAR, end_color=COLOR_FILA_IMPAR, fill_type="solid")
            font_datos = Font(name='Calibri', size=10)
            alineacion_centrada = Alignment(horizontal='center', vertical='center')
            alineacion_izquierda = Alignment(horizontal='left', vertical='center')
            
            speci_fill = PatternFill(start_color=COLOR_SPECI, end_color=COLOR_SPECI, fill_type="solid")
            speci_font = Font(name='Calibri', size=10, bold=True)
            
            for row in range(2, len(df) + 2):
                es_impar = (row % 2 == 1)
                tipo_reporte = worksheet.cell(row=row, column=3).value
                
                for col in range(1, len(columnas) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    if tipo_reporte == "SPECI":
                        cell.fill = speci_fill
                        cell.font = speci_font
                    else:
                        cell.border = border_datos
                        cell.font = font_datos
                        if es_impar:
                            cell.fill = fill_impar
                    
                    col_letter = get_column_letter(col)
                    if col_letter in ['A', 'B', 'C', 'D', 'E', 'H', 'I', 'J', 'N', 'O', 'P', 'Q', 'R']:
                        cell.alignment = alineacion_centrada
                    else:
                        cell.alignment = alineacion_izquierda
            
            worksheet.freeze_panes = 'A2'
        
        output.seek(0)
        
        speci_count = len([r for r in registros if r.get('Tipo') == 'SPECI'])
        mensaje = f"✅ {len(registros)} registros exportados\n"
        mensaje += f"   📊 METAR: {len(registros) - speci_count}\n"
        mensaje += f"   🟨 SPECI: {speci_count} (resaltados en amarillo)"
        
        return output, mensaje
        
    except Exception as e:
        return None, f"Error al exportar: {str(e)}"

# ============================================
# INTERFAZ DE USUARIO
# ============================================

# Header
col_header1, col_header2 = st.columns([3, 1])
with col_header1:
    st.markdown("<h1 style='color: #0b3d91;'>✈️ METAR DIGITAL - SPJC</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color: #666;'>Aeropuerto Internacional Jorge Chávez - CORPAC Perú</p>", unsafe_allow_html=True)

with col_header2:
    ahora = datetime.now(timezone.utc).strftime("%H:%M:%S")
    st.markdown(f"<h3 style='color: #0b3d91; text-align: right;'>UTC {ahora}</h3>", unsafe_allow_html=True)
    st.markdown(f"<p style='color: #666; text-align: right;'>{datetime.now().strftime('%d/%m/%Y')}</p>", unsafe_allow_html=True)

st.markdown("---")

# Columnas principales
col_izq, col_der = st.columns([2, 1])

with col_izq:
    with st.form(key='metar_form'):
        # DATOS DEL REPORTE
        st.subheader("📋 DATOS DEL REPORTE")
        col1, col2, col3 = st.columns(3)
        with col1:
            tipo = st.selectbox("Tipo", ["METAR", "SPECI"], key='tipo')
        with col2:
            dia = st.text_input("Día", key='dia')
        with col3:
            hora = st.text_input("Hora UTC", key='hora', help="Formato HHMM")
        
        st.markdown("---")
        
        # VIENTO
        st.subheader("💨 VIENTO")
        col1, col2, col3 = st.columns(3)
        with col1:
            dir_viento = st.text_input("Dirección", key='dir_viento', help="Grados (0-360)")
        with col2:
            int_viento = st.text_input("Intensidad (KT)", key='int_viento', help="Nudos. Ráfagas: 15G25")
        with col3:
            var_viento = st.text_input("Variación", key='var_viento', help="Formato: 340V080")
        
        st.markdown("---")
        
        # VISIBILIDAD
        st.subheader("👁️ VISIBILIDAD")
        col1, col2, col3 = st.columns(3)
        with col1:
            vis = st.text_input("Visibilidad", key='vis', help="Ej: 10km, 5000m, 9999")
        with col2:
            vis_min = st.text_input("Visibilidad Mínima", key='vis_min', help="Ej: 1200SW, 0800NE, 3000N")
        with col3:
            rvr = st.text_input("RVR (m)", key='rvr', help="Runway Visual Range - Ej: 0600, 1200")
        
        st.markdown("---")
        
        # FENÓMENOS Y NUBES
        st.subheader("☁️ FENÓMENOS Y NUBES")
        fenomeno = st.text_input("Fenómeno", key='fenomeno', help="Ej: niebla parcial (PRFG), lluvia ligera (-RA)")
        nubes = st.text_input("Nubes", key='nubes', help="Ej: 8 ST 300M, 5 AC 5000M, CB 1500M")
        
        st.markdown("---")
        
        # TEMPERATURA, HUMEDAD Y PRESIÓN
        st.subheader("🌡️ TEMPERATURA, HUMEDAD Y PRESIÓN")
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            temp = st.text_input("Temp °C", key='temp', help="-10 a 40°C")
        with col2:
            rocio = st.text_input("Rocío °C", key='rocio', help="≤ Temperatura")
        with col3:
            hr = st.text_input("HR %", key='hr', help="Humedad Relativa (0-100%)")
        with col4:
            qnh = st.text_input("QNH hPa", key='qnh', help="850-1100 hPa")
        with col5:
            presion = st.text_input("Presión Est.", key='presion', help="Opcional")
        
        st.markdown("---")
        
        # INFORMACIÓN SUPLEMENTARIA
        st.subheader("📝 INFORMACIÓN SUPLEMENTARIA")
        suplementaria = st.text_input("Suplementaria", key='suplementaria', help="Opcional - Ej: RMK CB AL NE")
        
        st.markdown("---")
        
        # BOTONES
        col1, col2 = st.columns(2)
        with col1:
            generar = st.form_submit_button("🛫 GENERAR METAR", use_container_width=True)
        with col2:
            limpiar = st.form_submit_button("🧹 LIMPIAR CAMPOS", use_container_width=True)
        
        if limpiar:
            limpiar_campos()
            st.rerun()
        
        if generar:
            datos = {
                'tipo': tipo,
                'dia': dia,
                'hora': hora,
                'dir_viento': dir_viento,
                'int_viento': int_viento,
                'var_viento': var_viento,
                'vis': vis,
                'vis_min': vis_min,
                'rvr': rvr,
                'fenomeno': fenomeno,
                'nubes': nubes,
                'temp': temp,
                'rocio': rocio,
                'hr': hr,
                'qnh': qnh,
                'presion': presion,
                'suplementaria': suplementaria
            }
            
            resultado = generar_metar(datos)
            
            if resultado['success']:
                # ACTUALIZAR O INSERTAR REGISTRO (EVITAR DUPLICADOS)
                accion = actualizar_o_insertar_registro(st.session_state.registros, resultado['registro'])
                
                # ACTUALIZAR HISTORIAL - ELIMINAR DUPLICADOS Y AGREGAR NUEVO
                dia_hora_clave = f"{resultado['registro']['Día']}_{resultado['registro']['Hora']}"
                
                # Filtrar historial para eliminar cualquier METAR con la misma fecha/hora
                nuevo_historial = []
                for metar in st.session_state.historial:
                    # Verificar si este metar en el historial tiene la misma clave
                    # Extraer día y hora del METAR (formato: METAR SPJC DDhhmmZ)
                    match = re.search(r'SPJC (\d{2})(\d{4})Z', metar)
                    if match:
                        dia_hist = match.group(1)
                        hora_hist = match.group(2)
                        clave_hist = f"{dia_hist}_{hora_hist}"
                        if clave_hist != dia_hora_clave:
                            nuevo_historial.append(metar)
                    else:
                        # Si no se puede extraer, mantenerlo
                        nuevo_historial.append(metar)
                
                # Insertar el nuevo METAR al inicio
                nuevo_historial.insert(0, resultado['metar'])
                st.session_state.historial = nuevo_historial
                
                # Actualizar contador (número de registros únicos)
                st.session_state.contador = len(st.session_state.registros)
                
                if accion == "actualizado":
                    st.warning("🔄 METAR ACTUALIZADO - Reemplazó reporte existente con la misma fecha/hora")
                else:
                    st.success("✅ METAR generado correctamente")
                
                st.session_state.ultimo_metar = resultado['metar']
                st.session_state.ultimo_tipo = tipo
            else:
                st.error(f"❌ {resultado['error']}")

with col_der:
    # METAR GENERADO
    st.subheader("📊 METAR GENERADO")
    if 'ultimo_metar' in st.session_state:
        tipo_ultimo = st.session_state.get('ultimo_tipo', 'METAR')
        if tipo_ultimo == "SPECI":
            st.markdown(f"<div style='background: #FFE699; padding: 15px; border-radius: 5px; font-family: monospace; border-left: 5px solid #FFC000;'><b>⚠️ SPECI</b><br>{st.session_state.ultimo_metar}</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div style='background: #1e1e1e; color: #00ff00; padding: 15px; border-radius: 5px; font-family: monospace; border-left: 5px solid #0b3d91;'>{st.session_state.ultimo_metar}</div>", unsafe_allow_html=True)
    else:
        st.info("---")
    
    st.markdown("---")
    
    # BOTONES DE ACCIÓN
    st.subheader("💾 EXPORTAR")
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("📥 Exportar Excel", use_container_width=True):
            if st.session_state.registros:
                excel_file, mensaje = exportar_a_excel(st.session_state.registros)
                if excel_file:
                    st.download_button(
                        label="📥 Descargar Excel",
                        data=excel_file,
                        file_name=f"METAR_SPJC_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    st.success(mensaje)
                else:
                    st.warning(mensaje)
            else:
                st.warning("⚠️ No hay registros para exportar")
    
    with col2:
        if st.button("🗑️ Guardar y Limpiar", use_container_width=True):
            if st.session_state.registros:
                excel_file, _ = exportar_a_excel(st.session_state.registros)
                if excel_file:
                    st.session_state.registros = []
                    st.session_state.historial = []
                    st.session_state.contador = 0
                    st.success("✅ Memoria limpiada")
                else:
                    st.warning("⚠️ No hay registros para guardar")
            else:
                st.warning("⚠️ No hay registros para limpiar")
    
    st.markdown("---")
    
    # CONTADOR DE REGISTROS ÚNICOS
    st.metric("📋 REGISTROS ÚNICOS EN MEMORIA", st.session_state.contador)
    
    st.markdown("---")
    
    # HISTORIAL (SIN DUPLICADOS)
    st.subheader("📜 HISTORIAL")
    if st.session_state.historial:
        for i, metar in enumerate(st.session_state.historial[:10]):
            if "SPECI" in metar:
                st.markdown(f"<div style='background: #FFE699; padding: 8px; margin-bottom: 5px; border-radius: 3px; font-family: monospace; font-size: 12px; border-left: 3px solid #FFC000;'>{metar}</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='background: #f0f0f0; padding: 8px; margin-bottom: 5px; border-radius: 3px; font-family: monospace; font-size: 12px; border-left: 3px solid #0b3d91;'>{metar}</div>", unsafe_allow_html=True)
    else:
        st.info("No hay METARs en el historial")

# ============================================
# FOOTER
# ============================================
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666; padding: 20px;'>
    <p>✈️ METAR Digital - CORPAC Perú | Aeropuerto Internacional Jorge Chávez (SPJC)</p>
    <p style='font-size: 0.8rem;'>Versión 2.0 - Sistema Profesional de Codificación METAR/SPECI</p>
    <p style='font-size: 0.8rem;'>✓ RVR ✓ Visibilidad Mínima con Cuadrantes ✓ Variación Circular de Viento ✓ SIN DUPLICADOS</p>
</div>
""", unsafe_allow_html=True)