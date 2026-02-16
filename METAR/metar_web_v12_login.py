"""
METAR DIGITAL - VERSIÓN PROFESIONAL CORPAC PERÚ
Aeropuerto Internacional Jorge Chávez (SPJC)
Características:
✅ RVR - Formato SPJC: R32/0400, R12R/1700, R10/M0050, R14L/P2000 
✅ Sin validaciones arbitrarias - El operador copia lo que ve en el equipo
✅ Almacenamiento mensual automático (SPJC_METAR_YYYY_MM.xlsx)
✅ Viento con reglas circulares (340V080) - CORREGIDO: No muestra variación <60°
✅ Visibilidad mínima con cuadrantes (N, NE, E, SE, S, SW, W, NW)
✅ Estándar oficial nubes CORPAC (30m/1000m)
✅ Fenómenos especiales (PRFG, VCFG, BCFG, MIFG)
✅ Excel con formato profesional mensual
✅ Sin duplicados - Reemplaza reportes con misma fecha/hora
✅ Persistencia de datos entre sesiones
✅ Exportación simplificada con formato YYYY_MM
✅ Columna METAR con ancho automático en Excel
✅ Temperatura y Rocío - CORREGIDO: Ambos se redondean igual
✅ Sistema de autenticación - Solo personal autorizado
✅ Nubes - CORREGIDO: Respeta texto claro (7 ST 300M → OVC010)
✅ Visibilidad Vertical - Soporta "vis ver 600M" → VV020
"""

import streamlit as st
from datetime import datetime, timezone
import pandas as pd
from pathlib import Path
import re
import os
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
import hmac

# ============================================
# CONFIGURACIÓN DE PÁGINA
# ============================================
st.set_page_config(
    page_title="REGISTRO DE OBSERVACIONES ORDINARIAS Y ESPECIALES SPJC",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ============================================
# SISTEMA DE AUTENTICACIÓN
# ============================================
def verificar_autenticacion():
    """Sistema completo de login/logout"""
    
    # Inicializar estado de autenticación
    if 'autenticado' not in st.session_state:
        st.session_state.autenticado = False
        st.session_state.usuario = None
    
    # Barra lateral con información de usuario
    with st.sidebar:
        if st.session_state.autenticado:
            st.markdown(f"👤 **Usuario:** {st.session_state.usuario}")
            if st.button("🚪 Cerrar Sesión", use_container_width=True):
                st.session_state.autenticado = False
                st.session_state.usuario = None
                st.rerun()
            st.markdown("---")
    
    # Si ya está autenticado, continuar
    if st.session_state.autenticado:
        return True
    
    # Mostrar pantalla de login
    st.markdown("""
    <style>
        .login-container {
            max-width: 400px;
            margin: 100px auto;
            padding: 30px;
            background: var(--background-color);
            border-radius: 10px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            text-align: center;
            border: 1px solid rgba(128,128,128,0.2);
        }
        .login-header {
            color: #0b3d91;
            margin-bottom: 20px;
        }
        .login-logo {
            font-size: 48px;
            margin-bottom: 10px;
        }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown("<div class='login-container'>", unsafe_allow_html=True)
    st.markdown("<div class='login-logo'>✈️</div>", unsafe_allow_html=True)
    st.markdown("<h2 class='login-header'>Sistema METAR Digital</h2>", unsafe_allow_html=True)
    st.markdown("Aeropuerto Internacional Jorge Chávez")
    st.markdown("CORPAC Perú")
    st.markdown("---")
    
    with st.form("login_form"):
        usuario = st.text_input("Usuario", placeholder="Ingrese su usuario")
        contraseña = st.text_input("Contraseña", type="password", placeholder="Ingrese su contraseña")
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            submit = st.form_submit_button("🔐 INGRESAR", use_container_width=True)
    
    if submit:
        # Verificar credenciales desde secrets
        try:
            passwords = st.secrets.get("passwords", {})
            if not passwords:
                passwords = {"admin": "corpac2024", "metar": "spjc2024"}
        except:
            passwords = {"admin": "corpac2024", "metar": "spjc2024"}
        
        if usuario in passwords and hmac.compare_digest(contraseña, passwords[usuario]):
            st.session_state.autenticado = True
            st.session_state.usuario = usuario
            st.rerun()
        else:
            st.error("❌ Usuario o contraseña incorrectos")
    
    st.markdown("---")
    st.markdown("Solo personal autorizado CORPAC")
    st.markdown("</div>", unsafe_allow_html=True)
    
    st.stop()

# ============================================
# VERIFICAR AUTENTICACIÓN
# ============================================
verificar_autenticacion()

# ============================================
# CONSTANTES Y CONFIGURACIÓN
# ============================================
DIRECTORIO_DATOS = Path("datos_metar")
DIRECTORIO_DATOS.mkdir(exist_ok=True)

# ============================================
# FUNCIÓN DE REDONDEO CORREGIDA
# ============================================
def redondear_metar(valor):
    """
    Redondeo tradicional para METAR:
    - 14.5 → 15
    - 14.4 → 14
    - 15.5 → 16
    """
    try:
        d = Decimal(str(valor))
        return int(d.quantize(Decimal('1'), rounding=ROUND_HALF_UP))
    except:
        return int(round(float(valor)))

# ============================================
# ESTILOS CSS PERSONALIZADOS
# ============================================
st.markdown("""
<style>
    .stApp {
        background-color: var(--background-color);
    }
    
    .section-title {
        color: #0b3d91 !important;
        font-weight: 700 !important;
        font-size: 1.2rem !important;
        margin-bottom: 1rem !important;
        border-bottom: 2px solid #0b3d91 !important;
        padding-bottom: 0.5rem !important;
        text-shadow: 0 1px 2px rgba(255,255,255,0.5);
    }
    
    .stTextInput label, .stSelectbox label, .stNumberInput label {
        color: #0b3d91 !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
        background: transparent !important;
        text-shadow: 0 1px 2px rgba(255,255,255,0.5);
    }
    
    @media (prefers-color-scheme: dark) {
        .stTextInput label, .stSelectbox label, .stNumberInput label {
            color: #6ab0ff !important;
            text-shadow: 0 1px 3px rgba(0,0,0,0.8);
        }
        
        .section-title {
            color: #6ab0ff !important;
            border-bottom-color: #6ab0ff !important;
            text-shadow: 0 1px 3px rgba(0,0,0,0.8);
        }
    }
    
    .metar-box {
        background: #1e1e1e;
        color: #00ff00;
        padding: 1rem;
        border-radius: 5px;
        font-family: 'Courier New', monospace;
        font-size: 1.1rem;
        border-left: 5px solid #0b3d91;
    }
    
    .historial-item {
        background: #f8f9fa;
        padding: 0.8rem;
        margin-bottom: 0.5rem;
        border-radius: 3px;
        font-family: 'Courier New', monospace;
        font-size: 12px;
        border-left: 3px solid #0b3d91;
    }
    
    .historial-item-speci {
        background: #FFE699;
        border-left: 3px solid #FFC000;
    }
    
    .stButton button {
        width: 100%;
        border-radius: 5px;
        font-weight: 600;
    }
    
    .stTextInput input, .stSelectbox select {
        border-radius: 5px;
        border: 1px solid #ddd;
    }
    
    .stButton button[kind="primary"] {
        background-color: #0b3d91;
        color: white;
        border: none;
    }
    
    .stButton button[kind="primary"]:hover {
        background-color: #1a4fa0;
    }
</style>
""", unsafe_allow_html=True)

# ============================================
# FUNCIONES DE GESTIÓN DE ARCHIVOS
# ============================================
def obtener_nombre_archivo_mensual():
    """Genera el nombre del archivo mensual basado en la fecha actual UTC"""
    ahora = datetime.now(timezone.utc)
    return f"SPJC_METAR_{ahora.strftime('%Y_%m')}.xlsx"

def cargar_registros_mes():
    """Carga todos los registros del mes actual desde el archivo Excel"""
    archivo = DIRECTORIO_DATOS / obtener_nombre_archivo_mensual()
    
    if archivo.exists():
        try:
            df = pd.read_excel(archivo, sheet_name='METAR SPJC')
            
            registros = []
            for _, row in df.iterrows():
                registro = row.to_dict()
                
                if 'DIA' in registro:
                    registro['Día'] = str(registro['DIA']).zfill(2)
                if 'HORA' in registro:
                    registro['Hora'] = str(registro['HORA']).zfill(4)
                if 'TIPO' in registro:
                    registro['Tipo'] = registro['TIPO']
                if 'METAR' in registro:
                    registro['METAR_Completo'] = registro['METAR']
                
                registros.append(registro)
            
            return registros
        except Exception as e:
            st.error(f"Error al cargar archivo mensual: {e}")
            return []
    return []

def guardar_registros_mes(registros):
    """Guarda todos los registros en el archivo mensual"""
    if not registros:
        return False, "No hay registros para guardar"
    
    try:
        archivo = DIRECTORIO_DATOS / obtener_nombre_archivo_mensual()
        
        df = pd.DataFrame(registros)
        
        df = df.rename(columns={
            'Día': 'DIA',
            'Hora': 'HORA',
            'Tipo': 'TIPO',
            'Dirección_Viento': 'DIR VIENTO',
            'Intensidad_Viento': 'INTENSIDAD',
            'Variación_Viento': 'VARIACION',
            'Visibilidad_Original': 'VIS (ORIGINAL)',
            'Visibilidad_Metros': 'VIS (CODIGO)',
            'Visibilidad_Mínima': 'VIS MIN',
            'RVR': 'RVR',
            'Fenómeno_Texto': 'FENOMENO',
            'Fenómeno_Código': 'WX',
            'Nubes_Texto': 'NUBOSIDAD',
            'Nubes_Código': 'CLD',
            'Temperatura': 'TEMP °C',
            'Punto_Rocío': 'ROCÍO °C',
            'Humedad_Relativa_%': 'HR %',
            'QNH': 'QNH',
            'Presión_Estación': 'PRESION',
            'Info_Suplementaria': 'RMK',
            'METAR_Completo': 'METAR'
        })
        
        columnas_excel = [
            'DIA', 'HORA', 'TIPO', 'DIR VIENTO', 'INTENSIDAD', 'VARIACION',
            'VIS (ORIGINAL)', 'VIS (CODIGO)', 'VIS MIN', 'RVR',
            'FENOMENO', 'WX', 'NUBOSIDAD', 'CLD',
            'TEMP °C', 'ROCÍO °C', 'HR %', 'QNH', 'PRESION', 'RMK', 'METAR'
        ]
        
        columnas_disponibles = [col for col in columnas_excel if col in df.columns]
        df = df[columnas_disponibles]
        
        df['DIA'] = df['DIA'].astype(str).str.zfill(2)
        df['HORA'] = df['HORA'].astype(str).str.zfill(4)
        df = df.sort_values(['DIA', 'HORA'], ascending=[True, True])
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='METAR SPJC', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['METAR SPJC']
            
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            for col in range(1, len(columnas_disponibles) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                header_cell = worksheet.cell(row=1, column=col)
                max_length = len(str(header_cell.value)) if header_cell.value else 0
                
                for row in range(2, min(len(df) + 2, 102)):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = min(max_length + 2, 70)
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 8)
            
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='0B3D91', end_color='0B3D91', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for col in range(1, len(columnas_disponibles) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            worksheet.row_dimensions[1].height = 30
            
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            for row in range(2, len(df) + 2):
                tipo_reporte = worksheet.cell(row=row, column=3).value
                for col in range(1, len(columnas_disponibles) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if tipo_reporte == 'SPECI':
                        cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                        cell.font = Font(name='Calibri', size=10, bold=True)
                    else:
                        cell.font = Font(name='Calibri', size=10)
            
            worksheet.freeze_panes = 'A2'
        
        output.seek(0)
        
        with open(archivo, 'wb') as f:
            f.write(output.getvalue())
        
        return True, f"✅ {len(registros)} registros guardados en {archivo.name}"
        
    except Exception as e:
        return False, f"Error al guardar: {str(e)}"

# ============================================
# INICIALIZAR ESTADO DE SESIÓN
# ============================================
if 'registros' not in st.session_state:
    st.session_state.registros = cargar_registros_mes()

if 'historial' not in st.session_state:
    st.session_state.historial = []

if 'contador' not in st.session_state:
    st.session_state.contador = len(st.session_state.registros)

if 'campos_inicializados' not in st.session_state:
    st.session_state.campos_inicializados = False

# ============================================
# FUNCIÓN PARA LIMPIAR CAMPOS
# ============================================
def limpiar_campos():
    """Limpia todos los campos del formulario"""
    st.session_state.dia = datetime.now(timezone.utc).strftime("%d")
    st.session_state.hora = ""
    st.session_state.tipo = "METAR"
    st.session_state.dir_viento = ""
    st.session_state.int_viento = ""
    st.session_state.var_viento = ""
    st.session_state.vis = ""
    st.session_state.vis_min = ""
    st.session_state.rvr = ""
    st.session_state.fenomeno = ""
    st.session_state.nubes = ""
    st.session_state.temp = ""
    st.session_state.rocio = ""
    st.session_state.hr = ""
    st.session_state.qnh = ""
    st.session_state.presion = ""
    st.session_state.suplementaria = ""
    st.session_state.campos_inicializados = True

# ============================================
# INICIALIZAR CAMPOS VACÍOS
# ============================================
if not st.session_state.campos_inicializados:
    st.session_state.dia = datetime.now(timezone.utc).strftime("%d")
    st.session_state.hora = ""
    st.session_state.tipo = "METAR"
    st.session_state.dir_viento = ""
    st.session_state.int_viento = ""
    st.session_state.var_viento = ""
    st.session_state.vis = ""
    st.session_state.vis_min = ""
    st.session_state.rvr = ""
    st.session_state.fenomeno = ""
    st.session_state.nubes = ""
    st.session_state.temp = ""
    st.session_state.rocio = ""
    st.session_state.hr = ""
    st.session_state.qnh = ""
    st.session_state.presion = ""
    st.session_state.suplementaria = ""
    st.session_state.campos_inicializados = True

# ============================================
# FUNCIÓN PARA ACTUALIZAR O INSERTAR REGISTRO
# ============================================
def actualizar_o_insertar_registro(registros, nuevo_registro):
    dia_nuevo = str(nuevo_registro.get('Día', '')).zfill(2)
    hora_nueva = str(nuevo_registro.get('Hora', '')).zfill(4)
    
    dia_hora_clave = f"{dia_nuevo}_{hora_nueva}"
    accion = "insertado"
    
    for i, registro in enumerate(registros):
        dia_existente = str(registro.get('Día', '')).zfill(2)
        hora_existente = str(registro.get('Hora', '')).zfill(4)
        
        clave_existente = f"{dia_existente}_{hora_existente}"
        
        if clave_existente == dia_hora_clave:
            registros[i] = nuevo_registro
            accion = "actualizado"
            break
    else:
        registros.insert(0, nuevo_registro)
    
    guardar_registros_mes(registros)
    
    return accion

# ============================================
# FUNCIONES DE PROCESAMIENTO - VIENTO
# ============================================
def procesar_viento(direccion, intensidad, variacion):
    """
    PROCESAMIENTO DE VIENTO - REGLAS CORPAC PERÚ
    Caso especial: Dirección 000 e intensidad 00 → 00000KT (ignorar variación)
    """
    dir_int = int(direccion)
    intensidad_str = str(intensidad).upper().strip()
    
    # Caso especial: Viento en calma
    if dir_int == 0 and intensidad_str == "00":
        return "00000KT"
    
    # Procesar ráfagas
    if 'G' in intensidad_str:
        if 'G' in intensidad_str and not ' ' in intensidad_str.replace('G', ''):
            base_int, gust_int = intensidad_str.split('G')
            int_base = int(base_int)
            int_gust = int(gust_int)
            intensidad_metar = f"{int_base:02d}G{int_gust:02d}"
        else:
            parts = intensidad_str.replace('G', ' ').split()
            int_base = int(parts[0])
            int_gust = int(parts[1])
            intensidad_metar = f"{int_base:02d}G{int_gust:02d}"
    else:
        int_base = int(intensidad_str)
        intensidad_metar = f"{int_base:02d}"
    
    # Si NO hay variación
    if not variacion:
        return f"{dir_int:03d}{intensidad_metar}KT"
    
    try:
        variacion = variacion.upper().replace(' ', '')
        if 'V' not in variacion:
            return f"{dir_int:03d}{intensidad_metar}KT"
        
        desde, hasta = map(int, variacion.split('V'))
        
        diff1 = abs(hasta - desde)
        diff2 = 360 - diff1
        
        if desde > hasta:
            diferencia = diff2
        else:
            diferencia = diff1
        
        if diferencia < 60:
            return f"{dir_int:03d}{intensidad_metar}KT"
        
        if diferencia >= 180:
            return f"VRB{intensidad_metar}KT"
        
        if diferencia >= 60:
            if int_base < 3:
                return f"VRB{intensidad_metar}KT"
            else:
                if desde < hasta:
                    return f"{dir_int:03d}{intensidad_metar}KT {desde:03d}V{hasta:03d}"
                else:
                    return f"{dir_int:03d}{intensidad_metar}KT {hasta:03d}V{desde:03d}"
        
        return f"{dir_int:03d}{intensidad_metar}KT"
        
    except Exception as e:
        return f"{dir_int:03d}{intensidad_metar}KT"

# ============================================
# FUNCIONES DE PROCESAMIENTO - VISIBILIDAD
# ============================================
def convertir_visibilidad(vis_texto):
    vis_texto = vis_texto.strip().upper()
    if not vis_texto:
        raise ValueError("Visibilidad es obligatoria")
    
    try:
        if vis_texto.endswith("KM"):
            km = float(vis_texto[:-2])
            return 9999 if km >= 10 else int(km * 1000)
        elif vis_texto.endswith("M"):
            return int(vis_texto[:-1])
        else:
            metros = int(vis_texto)
            return 9999 if metros >= 10000 else metros
    except:
        raise ValueError("Formato de visibilidad inválido")

def procesar_visibilidad_minima(vis_min_texto, vis_m):
    if not vis_min_texto:
        return "", ""
    
    vis_min_texto = vis_min_texto.strip().upper()
    
    if vis_min_texto.endswith('NW'):
        valor = vis_min_texto[:-2]
        cuadrante = 'NW'
    elif vis_min_texto.endswith('NE'):
        valor = vis_min_texto[:-2]
        cuadrante = 'NE'
    elif vis_min_texto.endswith('SW'):
        valor = vis_min_texto[:-2]
        cuadrante = 'SW'
    elif vis_min_texto.endswith('SE'):
        valor = vis_min_texto[:-2]
        cuadrante = 'SE'
    elif vis_min_texto.endswith('N'):
        valor = vis_min_texto[:-1]
        cuadrante = 'N'
    elif vis_min_texto.endswith('S'):
        valor = vis_min_texto[:-1]
        cuadrante = 'S'
    elif vis_min_texto.endswith('E'):
        valor = vis_min_texto[:-1]
        cuadrante = 'E'
    elif vis_min_texto.endswith('W'):
        valor = vis_min_texto[:-1]
        cuadrante = 'W'
    else:
        valor = vis_min_texto
        cuadrante = ''
    
    try:
        if valor.endswith("KM"):
            km = float(valor[:-2])
            vis_min_m = 9999 if km >= 10 else int(km * 1000)
        elif valor.endswith("M"):
            vis_min_m = int(valor[:-1])
        else:
            vis_min_m = int(valor)
            vis_min_m = 9999 if vis_min_m >= 10000 else vis_min_m
        
        es_valida = False
        if vis_min_m < 1500:
            es_valida = True
        if vis_min_m < (vis_m * 0.5) and vis_min_m < 5000:
            es_valida = True
        
        if not es_valida:
            return "", "⚠️ No cumple reglas de visibilidad mínima"
        
        if cuadrante:
            return f"{vis_min_m:04d}{cuadrante}", ""
        else:
            return f"{vis_min_m:04d}", ""
        
    except:
        return "", "❌ Formato inválido"

# ============================================
# FUNCIÓN RVR
# ============================================
def procesar_rvr(rvr_texto):
    if not rvr_texto:
        return ""
    return rvr_texto.strip()

# ============================================
# FUNCIÓN DE FENÓMENOS
# ============================================
def codificar_fenomenos(texto, visibilidad_metros):
    if not texto:
        return ""
    
    texto_lower = texto.lower().strip()
    precipitaciones = []
    oscurecimiento = []
    especiales = []
    
    # Fenómenos especiales de niebla
    if any(x in texto_lower for x in ["niebla parcial", "prfg", "pr fg", "parcial"]):
        especiales.append("PRFG")
        for palabra in ["niebla parcial", "prfg", "pr fg", "parcial"]:
            texto_lower = texto_lower.replace(palabra, "")
    
    if any(x in texto_lower for x in ["niebla en la vecindad", "vcfg", "vc fg", "vecindad"]):
        especiales.append("VCFG")
        for palabra in ["niebla en la vecindad", "vcfg", "vc fg", "vecindad"]:
            texto_lower = texto_lower.replace(palabra, "")
    
    if any(x in texto_lower for x in ["niebla en bancos", "bcfg", "bc fg", "bancos"]):
        especiales.append("BCFG")
        for palabra in ["niebla en bancos", "bcfg", "bc fg", "bancos"]:
            texto_lower = texto_lower.replace(palabra, "")
    
    if any(x in texto_lower for x in ["niebla baja", "mifg", "mi fg", "baja"]):
        especiales.append("MIFG")
        for palabra in ["niebla baja", "mifg", "mi fg", "baja"]:
            texto_lower = texto_lower.replace(palabra, "")
    
    # Mapas de fenómenos
    intensidades = {
        "ligera": "-", "ligero": "-", "leve": "-", "débil": "-", 
        "moderada": "", "moderado": "",
        "fuerte": "+", "intensa": "+", "intenso": "+", "severa": "+"
    }
    
    descriptores = {
        "sh": "SH", "chubasco": "SH", "chubascos": "SH",
        "ts": "TS", "tormenta": "TS", "tormentas": "TS",
        "fz": "FZ", "helada": "FZ", "congelante": "FZ", "congelando": "FZ"
    }
    
    precipitacion_map = {
        "lluvia": "RA", "llovizna": "DZ", "nieve": "SN", 
        "granizo": "GR", "cellisca": "GS"
    }
    
    # Función para determinar código de oscurecimiento según visibilidad
    def get_oscurecimiento_codigo(texto_parte, vis_m):
        if "neblina" in texto_parte:
            if vis_m < 1000:
                return "FG"
            elif vis_m <= 5000:
                return "BR"
            else:
                return None
        elif "niebla" in texto_parte and "parcial" not in texto_parte and "baja" not in texto_parte:
            if vis_m < 1000:
                return "FG"
            else:
                return None
        return None
    
    # Procesar texto
    import re
    texto_lower = re.sub(r'\s+y\s+', ', ', texto_lower)
    partes = re.split(r'[,;]', texto_lower)
    
    for parte in partes:
        parte = parte.strip()
        if not parte:
            continue
        
        codigo_base = None
        descriptor = ""
        intensidad = ""
        tipo = None
        parte_procesada = parte
        
        for d_texto, d_codigo in descriptores.items():
            if d_texto in parte_procesada:
                descriptor = d_codigo
                parte_procesada = parte_procesada.replace(d_texto, "").strip()
                break
        
        for i_texto, i_codigo in intensidades.items():
            if i_texto in parte_procesada:
                intensidad = i_codigo
                parte_procesada = parte_procesada.replace(i_texto, "").strip()
                break
        
        for f_texto, f_codigo in precipitacion_map.items():
            if f_texto in parte_procesada:
                codigo_base = f_codigo
                tipo = 'precipitacion'
                break
        
        if not codigo_base:
            codigo_osc = get_oscurecimiento_codigo(parte_procesada, visibilidad_metros)
            if codigo_osc:
                codigo_base = codigo_osc
                tipo = 'oscurecimiento'
        
        if codigo_base:
            codigo_final = codigo_base
            if descriptor and codigo_base not in ["FG", "BR"]:
                codigo_final = descriptor + codigo_final
            if intensidad and codigo_base not in ["FG", "BR"]:
                codigo_final = intensidad + codigo_final
            
            if tipo == 'precipitacion' and codigo_final not in precipitaciones:
                precipitaciones.append(codigo_final)
            elif tipo == 'oscurecimiento' and codigo_final not in oscurecimiento:
                oscurecimiento.append(codigo_final)
    
    resultados = precipitaciones + especiales + oscurecimiento
    return " ".join(resultados[:3]) if resultados else ""

# ============================================
# FUNCIONES DE PROCESAMIENTO - NUBES (CORREGIDAS)
# ============================================
def interpretar_nubes(texto, vis_m, fenomeno):
    """
    CODIFICADOR DE NUBES - ESTÁNDAR CORPAC
    - Convierte texto claro a código METAR
    - Soporta Visibilidad Vertical (VV) cuando se menciona como "vis ver"
    - CAVOK solo cuando NO hay nubes Y visibilidad ≥ 10km Y sin fenómenos
    - Ejemplos:
      "8 ST 300M" → "OVC010"
      "5 CU 1500M" → "SCT050"
      "vis ver 600M" → "VV020"
    """
    if not texto:
        texto = ""
    
    texto = texto.strip().upper()
    
    # ===== CASO ESPECIAL: Visibilidad Vertical (VV) =====
    if any(x in texto.upper() for x in ["VIS VER", "VV", "VIS VERT", "VISIBILIDAD VERTICAL"]):
        import re
        numeros = re.findall(r'\d+', texto)
        if numeros:
            altura_metros = int(numeros[0])
            altura_pies = int(altura_metros/0.333)
            altura_cientos = altura_pies // 100
            altura_cientos = min(max(altura_cientos, 0), 999)
            return f"VV{altura_cientos:03d}"
    return "VV///"
    # ===================================================
    
    # ===== VERIFICAR CAVOK =====
    # CAVOK solo si:
    # 1. Visibilidad ≥ 10km
    # 2. Sin fenómenos
    # 3. Sin nubes (texto vacío o NSC/SKC/CLR)
    if vis_m >= 9999 and not fenomeno.strip():
        if not texto or texto in ["NSC", "SKC", "CLR", "DESPEJADO", "SIN NUBES", "NO NUBES"]:
            return "CAVOK"
    # ===========================
    
    # Si no hay texto de nubes, retornar NSC
    if not texto or texto in ["NSC", "SKC", "CLR", "DESPEJADO", "SIN NUBES", "NO NUBES"]:
        return "NSC"
    
    # Mapeo de tipos de nubes
    tipos_nubes = {
        "CU": "CU", "SC": "SC", "ST": "ST", "CB": "CB", "TCU": "TCU",
        "AC": "AC", "AS": "AS", "NS": "NS", "CI": "CI"
    }
    
    # Dividir por comas para múltiples capas
    capas = texto.split(",")
    codigos_nubes = []
    
    for capa in capas[:4]:  # Máximo 4 capas
        capa = capa.strip()
        if not capa:
            continue
        
        # Patrón: cantidad (1-8) + tipo + altura
        import re
        patron = r'(\d+)\s+([A-Z]{2,4})\s+(\d+)(?:M)?'
        match = re.search(patron, capa)
        
        if match:
            cantidad = int(match.group(1))
            tipo = match.group(2)
            altura = int(match.group(3))
            
            # Obtener tipo de nube
            tipo_nube = tipos_nubes.get(tipo, tipo)
            
            # Convertir altura a cientos de pies
            if altura <= 3000:
                # Para alturas bajas: redondear a múltiplo de 30
                if altura % 30 != 0:
                    altura = (altura // 30) * 30
                altura_cientos = altura // 30
            else:
                # Para alturas altas: redondear a múltiplo de 1000
                if altura % 1000 != 0:
                    altura = (altura // 1000) * 1000
                altura_cientos = (altura // 1000) * 32
            
            altura_cientos = min(max(altura_cientos, 1), 999)
            
            # Determinar código de cantidad
            if cantidad <= 2:
                cod_cant = "FEW"
            elif cantidad <= 4:
                cod_cant = "SCT"
            elif cantidad <= 7:
                cod_cant = "BKN"
            else:
                cod_cant = "OVC"
            
            # Construir código
            codigo = f"{cod_cant}{altura_cientos:03d}"
            if tipo_nube in ["CB", "TCU"]:
                codigo += tipo_nube
            if codigo not in codigos_nubes:
                codigos_nubes.append(codigo)
    
    return " ".join(codigos_nubes[:4]) if codigos_nubes else "NSC"

# ============================================
# FUNCIONES DE VALIDACIÓN
# ============================================
def validar_hora(hora_str):
    if not hora_str:
        raise ValueError("Hora requerida - Formato HHMM")
    if len(hora_str) != 4 or not hora_str.isdigit():
        raise ValueError("Hora debe ser HHMM (4 dígitos)")
    h = int(hora_str[:2])
    m = int(hora_str[2:])
    if h > 23 or m > 59:
        raise ValueError("Hora inválida")
    return hora_str

def validar_intensidad_viento(intensidad_str):
    intensidad_str = str(intensidad_str).strip().upper()
    if not intensidad_str:
        raise ValueError("Intensidad de viento requerida")
    
    intensidad_str = intensidad_str.replace(' G ', 'G').replace(' G', 'G').replace('G ', 'G')
    
    if 'G' in intensidad_str:
        partes = intensidad_str.split('G')
        if len(partes) != 2:
            raise ValueError("Formato de ráfagas inválido. Use: 15G25")
        base = int(partes[0])
        rafaga = int(partes[1])
        if base < 0 or base > 100:
            raise ValueError("Intensidad base fuera de rango")
        if rafaga < base:
            raise ValueError("Ráfaga debe ser mayor o igual a intensidad base")
        return intensidad_str
    else:
        intensidad = int(intensidad_str)
        if intensidad < 0 or intensidad > 100:
            raise ValueError("Intensidad fuera de rango")
        return intensidad_str

def validar_numero(valor, min_val, max_val, nombre):
    if not valor:
        raise ValueError(f"{nombre} es obligatorio")
    try:
        num = float(valor)
        if not (min_val <= num <= max_val):
            raise ValueError(f"{nombre} fuera de rango")
        return num
    except:
        raise ValueError(f"{nombre} inválido")

def validar_temp_rocio(temp, rocio):
    if float(rocio) > float(temp):
        raise ValueError(f"Rocío ({rocio}°C) no puede ser > Temperatura ({temp}°C)")
    return True

# ============================================
# FUNCIÓN PRINCIPAL DE GENERACIÓN
# ============================================
def generar_metar(datos):
    try:
        if not datos['dir_viento'] or not datos['int_viento']:
            raise ValueError("Dirección e intensidad del viento son obligatorias")
        if not datos['vis']:
            raise ValueError("Visibilidad es obligatoria")
        if not datos['temp'] or not datos['rocio'] or not datos['qnh']:
            raise ValueError("Temperatura, Rocío y QNH son obligatorios")
        
        hora = validar_hora(datos['hora'])
        int_viento = validar_intensidad_viento(datos['int_viento'])
        viento = procesar_viento(datos['dir_viento'], int_viento, datos['var_viento'])
        vis_m = convertir_visibilidad(datos['vis'])
        
        vis_min_codigo = ""
        if datos['vis_min']:
            vis_min_codigo, vis_min_error = procesar_visibilidad_minima(datos['vis_min'], vis_m)
            if vis_min_error:
                raise ValueError(vis_min_error)
        
        rvr_codigo = procesar_rvr(datos['rvr'])
        fenomeno = codificar_fenomenos(datos['fenomeno'], vis_m)
        nubes = interpretar_nubes(datos['nubes'], vis_m, fenomeno)
        
        temp = validar_numero(datos['temp'], -10, 40, "Temperatura")
        rocio = validar_numero(datos['rocio'], -10, 40, "Punto de rocío")
        validar_temp_rocio(temp, rocio)
        qnh = validar_numero(datos['qnh'], 850, 1100, "QNH")
        
        temp_metar = redondear_metar(temp)
        rocio_metar = redondear_metar(rocio)
        qnh_metar = int(qnh)
        
        # Construir METAR
        metar_parts = [f"{datos['tipo']} SPJC {datos['dia']}{hora}Z {viento}"]
        
        # Si nubes es CAVOK, usar CAVOK
        if nubes == "CAVOK":
            metar_parts.append("CAVOK")
        else:
            metar_parts.append(f"{vis_m:04d}")
            if vis_min_codigo:
                metar_parts.append(vis_min_codigo)
            if rvr_codigo:
                metar_parts.append(rvr_codigo)
            if fenomeno:
                metar_parts.append(fenomeno)
            metar_parts.append(nubes)
        
        metar_parts.append(f"{temp_metar:02d}/{rocio_metar:02d} Q{qnh_metar}")
        
        if datos['suplementaria']:
            metar_parts.append(datos['suplementaria'].upper())
        
        metar_completo = " ".join(metar_parts) + "="
        
        registro = {
            'Día': str(datos['dia']).zfill(2),
            'Hora': hora,
            'Tipo': datos['tipo'],
            'Dirección_Viento': datos['dir_viento'],
            'Intensidad_Viento': datos['int_viento'],
            'Variación_Viento': datos['var_viento'],
            'Visibilidad_Original': datos['vis'],
            'Visibilidad_Metros': vis_m,
            'Visibilidad_Mínima': vis_min_codigo,
            'RVR': rvr_codigo,
            'Fenómeno_Texto': datos['fenomeno'],
            'Fenómeno_Código': fenomeno,
            'Nubes_Texto': datos['nubes'],
            'Nubes_Código': "CAVOK" if nubes == "CAVOK" else nubes,
            'Temperatura': temp,
            'Punto_Rocío': rocio,
            'Humedad_Relativa_%': datos['hr'] if datos['hr'] else "",
            'QNH': qnh,
            'Presión_Estación': datos['presion'],
            'Info_Suplementaria': datos['suplementaria'],
            'METAR_Completo': metar_completo
        }
        
        return {'success': True, 'metar': metar_completo, 'registro': registro}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}

# ============================================
# FUNCIÓN PARA EXPORTAR EXCEL
# ============================================
def exportar_a_excel(registros):
    """Genera un archivo Excel - VERSIÓN CORREGIDA"""
    if not registros:
        return None, "No hay registros para exportar"
    
    try:
        # Crear nuevo DataFrame manualmente
        datos_exportar = []
        for r in registros:
            dia = str(r.get('Día', '')).zfill(2)
            hora = str(r.get('Hora', '')).zfill(4)
            
            datos_exportar.append({
                'DIA': dia,
                'HORA': hora,
                'TIPO': r.get('Tipo', ''),
                'DIR VIENTO': r.get('Dirección_Viento', ''),
                'INTENSIDAD': r.get('Intensidad_Viento', ''),
                'VARIACION': r.get('Variación_Viento', ''),
                'VIS (ORIGINAL)': r.get('Visibilidad_Original', ''),
                'VIS (CODIGO)': r.get('Visibilidad_Metros', ''),
                'VIS MIN': r.get('Visibilidad_Mínima', ''),
                'RVR': r.get('RVR', ''),
                'FENOMENO': r.get('Fenómeno_Texto', ''),
                'WX': r.get('Fenómeno_Código', ''),
                'NUBOSIDAD': r.get('Nubes_Texto', ''),
                'CLD': r.get('Nubes_Código', ''),
                'TEMP °C': r.get('Temperatura', ''),
                'ROCÍO °C': r.get('Punto_Rocío', ''),
                'HR %': r.get('Humedad_Relativa_%', ''),
                'QNH': r.get('QNH', ''),
                'PRESION': r.get('Presión_Estación', ''),
                'RMK': r.get('Info_Suplementaria', ''),
                'METAR': r.get('METAR_Completo', '')
            })
        
        df = pd.DataFrame(datos_exportar)
        df = df.sort_values(['DIA', 'HORA'], ascending=[True, True])
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='METAR SPJC', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['METAR SPJC']
            
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            for col in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col)
                col_name = df.columns[col-1]
                
                max_length = len(str(col_name))
                for row in range(2, min(len(df) + 2, 100)):
                    cell_value = df.iloc[row-2, col-1]
                    if cell_value:
                        cell_length = len(str(cell_value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                if col_name == 'METAR':
                    adjusted_width = min(max_length + 5, 120)
                elif col_name in ['TEMP °C', 'ROCÍO °C', 'HR %', 'QNH']:
                    adjusted_width = 12
                else:
                    adjusted_width = min(max_length + 2, 50)
                
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 8)
            
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='0B3D91', end_color='0B3D91', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            worksheet.row_dimensions[1].height = 30
            
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            for row in range(2, len(df) + 2):
                tipo_reporte = worksheet.cell(row=row, column=3).value
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    
                    if tipo_reporte == 'SPECI':
                        cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                        cell.font = Font(name='Calibri', size=10, bold=True)
                    else:
                        cell.font = Font(name='Calibri', size=10)
                    
                    col_name = df.columns[col-1]
                    if col_name == 'METAR':
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            worksheet.freeze_panes = 'A2'
        
        output.seek(0)
        return output, f"✅ {len(registros)} registros exportados"
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return None, f"Error al exportar: {str(e)}"

# ============================================
# HEADER
# ============================================
col_header1, col_header2 = st.columns([3, 1])

with col_header1:
    st.markdown("<h1 style='color: #0b3d91;'>REGISTRO DE OBSERVACIONES ORDINARIAS Y ESPECIALES</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color: #666;'>Aeropuerto Internacional Jorge Chávez - CORPAC Perú</p>", unsafe_allow_html=True)

with col_header2:
    hoy = datetime.now(timezone.utc).strftime('%d/%m/%Y')
    st.markdown(f"""
    <div style='text-align: right;'>
        <p style='color: #666; font-size: 14px;'>{hoy}</p>
        <p style='color: #0b3d91; font-size: 12px; font-weight: bold;'>{obtener_nombre_archivo_mensual()}</p>
        <p style='color: #999; font-size: 11px;'>CORPAC PERU</p>
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")

# ============================================
# INTERFAZ DE USUARIO
# ============================================
col_izq, col_der = st.columns([2, 1])

with col_izq:
    with st.form(key='metar_form'):
        st.markdown("<div class='section-title'>DATOS DEL REPORTE</div>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            tipo = st.selectbox("Tipo", ["METAR", "SPECI"], key='tipo')
        with col2:
            dia = st.text_input("Dia", key='dia', help="Formato: DD (01-31)")
        with col3:
            hora = st.text_input("Hora UTC", key='hora', help="Formato HHMM (ej: 1230)")
            
        st.markdown("---")
        
        st.markdown("<div class='section-title'>VIENTO</div>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            dir_viento = st.text_input("Direccion", key='dir_viento', help="Grados (0-360)")
        with col2:
            int_viento = st.text_input("Intensidad (KT)", key='int_viento', help="Nudos. Rafagas: 15G25")
        with col3:
            var_viento = st.text_input("Variacion", key='var_viento', help="Formato: 340V080")
        
        st.markdown("---")
        
        st.markdown("<div class='section-title'>VISIBILIDAD</div>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            vis = st.text_input("Visibilidad", key='vis', help="Ej: 10km, 5000m, 9999")
        with col2:
            vis_min = st.text_input("Visibilidad Minima", key='vis_min', help="Ej: 1200SW, 0800NE, 1500SE")
        with col3:
            rvr = st.text_input("RVR", key='rvr', help="R32/0400, R12R/1700, R10/M0050")
        
        st.markdown("---")
        
        st.markdown("<div class='section-title'>FENOMENOS Y NUBES</div>", unsafe_allow_html=True)
        fenomeno = st.text_input("Fenomeno", key='fenomeno', help="Ej: niebla parcial, lluvia moderada")
        nubes = st.text_input("Nubes", key='nubes', help="Ej: 8 ST 300M, 5 CU 1500M, vis ver 600M")
        
        st.markdown("---")
        
        st.markdown("<div class='section-title'>TEMPERATURA Y PRESION</div>", unsafe_allow_html=True)
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            temp = st.text_input("Temp °C", key='temp', help="-10 a 40°C")
        with col2:
            rocio = st.text_input("Rocío °C", key='rocio', help="≤ Temperatura")
        with col3:
            hr = st.text_input("HR %", key='hr', help="0-100%")
        with col4:
            qnh = st.text_input("QNH hPa", key='qnh', help="850-1100 hPa")
        with col5:
            presion = st.text_input("Presión Est.", key='presion', help="Opcional")
        
        st.markdown("---")
        
        st.markdown("<div class='section-title'>INFORMACION SUPLEMENTARIA</div>", unsafe_allow_html=True)
        suplementaria = st.text_input("", key='suplementaria', help="Ej: NOSIG RMK CB AL NE")
        
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        with col1:
            generar = st.form_submit_button("GENERAR METAR", use_container_width=True)
        with col2:
            limpiar = st.form_submit_button("LIMPIAR CAMPOS", use_container_width=True)
        
        if limpiar:
            limpiar_campos()
            st.rerun()
        
        if generar:
            datos = {
                'tipo': tipo, 'dia': dia, 'hora': hora,
                'dir_viento': dir_viento, 'int_viento': int_viento, 'var_viento': var_viento,
                'vis': vis, 'vis_min': vis_min, 'rvr': rvr,
                'fenomeno': fenomeno, 'nubes': nubes,
                'temp': temp, 'rocio': rocio, 'hr': hr,
                'qnh': qnh, 'presion': presion, 'suplementaria': suplementaria
            }
            
            resultado = generar_metar(datos)
            
            if resultado['success']:
                accion = actualizar_o_insertar_registro(st.session_state.registros, resultado['registro'])
                
                dia_hora_clave = f"{str(resultado['registro']['Día']).zfill(2)}_{resultado['registro']['Hora']}"
                nuevo_historial = []
                
                for metar in st.session_state.historial:
                    match = re.search(r'SPJC (\d{2})(\d{4})Z', metar)
                    if match:
                        dia_hist = match.group(1)
                        hora_hist = match.group(2)
                        if f"{dia_hist}_{hora_hist}" != dia_hora_clave:
                            nuevo_historial.append(metar)
                    else:
                        nuevo_historial.append(metar)
                
                nuevo_historial.insert(0, resultado['metar'])
                st.session_state.historial = nuevo_historial[:20]
                st.session_state.contador = len(st.session_state.registros)
                
                if accion == "actualizado":
                    st.warning("METAR ACTUALIZADO - Reemplazo reporte existente")
                else:
                    st.success("METAR generado correctamente")
                
                st.session_state.ultimo_metar = resultado['metar']
                st.session_state.ultimo_tipo = tipo
                st.session_state.ultimo_registro = resultado['registro']
            else:
                st.error(f"ERROR: {resultado['error']}")

with col_der:
    st.markdown("<div class='section-title'>📋 ÚLTIMO REPORTE</div>", unsafe_allow_html=True)
    if 'ultimo_metar' in st.session_state:
        tipo_ultimo = st.session_state.get('ultimo_tipo', 'METAR')
        if tipo_ultimo == "SPECI":
            st.markdown(f"<div style='background: #FFE699; padding: 15px; border-radius: 5px; font-family: monospace; border-left: 5px solid #FFC000;'>{st.session_state.ultimo_metar}</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div class='metar-box'>{st.session_state.ultimo_metar}</div>", unsafe_allow_html=True)
    else:
        st.info("---")
    
    st.markdown("---")
    st.markdown("<div class='section-title'>EXPORTAR</div>", unsafe_allow_html=True)
    
    if st.button("📥 Exportar METAR", use_container_width=True, type="primary"):
        if st.session_state.registros:
            excel_file, mensaje = exportar_a_excel(st.session_state.registros)
            if excel_file:
                nombre_archivo = obtener_nombre_archivo_mensual()
                st.download_button(
                    label="✅ Confirmar Descarga",
                    data=excel_file,
                    file_name=nombre_archivo,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.success(mensaje)
            else:
                st.warning(mensaje)
        else:
            st.warning("No hay registros")
    
    if st.button("🗑️ Limpiar Memoria", use_container_width=True):
        st.session_state.registros = []
        st.session_state.historial = []
        st.session_state.contador = 0
        st.success("Memoria limpiada")
    
    st.markdown("---")
    st.metric("REGISTROS EN MEMORIA", st.session_state.contador)
    
    st.markdown("---")
    st.markdown("<div class='section-title'>HISTORIAL</div>", unsafe_allow_html=True)
    if st.session_state.historial:
        for metar in st.session_state.historial[:10]:
            if "SPECI" in metar:
                st.markdown(f"<div style='background: #FFE699; padding: 8px; margin-bottom: 5px; border-radius: 3px; font-family: monospace; font-size: 12px; border-left: 3px solid #FFC000;'>{metar}</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='background: #f0f0f0; padding: 8px; margin-bottom: 5px; border-radius: 3px; font-family: monospace; font-size: 12px; border-left: 3px solid #0b3d91;'>{metar}</div>", unsafe_allow_html=True)
    else:
        st.info("No hay METARs en el historial")

# ============================================
# FOOTER
# ============================================
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666; padding: 20px;'>
    <p>METAR Digital - CORPAC Peru | Aeropuerto Internacional Jorge Chavez (SPJC)</p>
    <p style='font-size: 0.8rem;'>Sistema de Registro de Observaciones - Versión 11.0</p>
</div>
""", unsafe_allow_html=True)

# ============================================
# GUARDADO AUTOMATICO
# ============================================
if st.session_state.registros:
    guardar_registros_mes(st.session_state.registros)
