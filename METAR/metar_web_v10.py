"""
METAR DIGITAL - VERSIÓN PROFESIONAL CORPAC PERÚ
Aeropuerto Internacional Jorge Chávez (SPJC)
Características:
✅ RVR - Formato SPJC: R32/0400, R12R/1700, R10/M0050, R14L/P2000 
✅ Sin validaciones arbitrarias - El operador copia lo que ve en el equipo
✅ Almacenamiento mensual automático (SPJC_METAR_YYYY_MM.xlsx)
✅ Viento con reglas circulares (340V080)
✅ Visibilidad mínima con cuadrantes (N, NE, E, SE, S, SW, W, NW)
✅ Estándar oficial nubes CORPAC (30m/1000m)
✅ Fenómenos especiales (PRFG, VCFG, BCFG, MIFG)
✅ Excel con formato profesional mensual
✅ Sin duplicados - Reemplaza reportes con misma fecha/hora
✅ Persistencia de datos entre sesiones
✅ Exportación simplificada con formato YYYY_MM
✅ Columna METAR con ancho automático en Excel
"""

import streamlit as st
from datetime import datetime, timezone
import pandas as pd
from pathlib import Path
import re
import os
from io import BytesIO

# ============================================
# CONFIGURACIÓN DE PÁGINA
# ============================================
st.set_page_config(
    page_title="REGISTRO DE OBSERVACIONES ORDINARIAS Y ESPECIALES SPJC",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ============================================
# CONSTANTES Y CONFIGURACIÓN
# ============================================
DIRECTORIO_DATOS = Path("datos_metar")
DIRECTORIO_DATOS.mkdir(exist_ok=True)

# ============================================
# ESTILOS CSS PERSONALIZADOS
# ============================================
st.markdown("""
<style>
    .stApp {
        background-color: #f0f8ff;
    }
    
    .metar-header {
        background: linear-gradient(90deg, #0b3d91 0%, #1a4fa0 100%);
        padding: 1.5rem;
        border-radius: 10px;
        color: white;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    
    .panel {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        margin-bottom: 1rem;
    }
    
    .section-title {
        color: #0b3d91;
        font-weight: 600;
        margin-bottom: 1rem;
        border-bottom: 2px solid #e8eef7;
        padding-bottom: 0.5rem;
    }
    
    .metar-box {
        background: #1e1e1e;
        color: #00ff00;
        padding: 1rem;
        border-radius: 5px;
        font-family: 'Courier New', monospace;
        font-size: 1.1rem;
        border-left: 5px solid #0b3d91;
    }
    
    .historial-item {
        background: #f8f9fa;
        padding: 0.8rem;
        margin-bottom: 0.5rem;
        border-radius: 3px;
        font-family: 'Courier New', monospace;
        font-size: 12px;
        border-left: 3px solid #0b3d91;
    }
    
    .historial-item-speci {
        background: #FFE699;
        border-left: 3px solid #FFC000;
    }
    
    .stButton button {
        width: 100%;
        border-radius: 5px;
        font-weight: 600;
    }
    
    .stTextInput input, .stSelectbox select {
        border-radius: 5px;
        border: 1px solid #ddd;
    }
    
    .hora-requerida {
        color: #ff0000;
        font-size: 12px;
        font-weight: bold;
    }
    
    /* Estilo para botón primario */
    .stButton button[kind="primary"] {
        background-color: #0b3d91;
        color: white;
        border: none;
    }
    
    .stButton button[kind="primary"]:hover {
        background-color: #1a4fa0;
    }
</style>
""", unsafe_allow_html=True)

# ============================================
# FUNCIONES DE GESTIÓN DE ARCHIVOS
# ============================================
def obtener_nombre_archivo_mensual():
    """Genera el nombre del archivo mensual basado en la fecha actual UTC"""
    ahora = datetime.now(timezone.utc)
    return f"SPJC_METAR_{ahora.strftime('%Y_%m')}.xlsx"

def cargar_registros_mes():
    """Carga todos los registros del mes actual desde el archivo Excel"""
    archivo = DIRECTORIO_DATOS / obtener_nombre_archivo_mensual()
    
    if archivo.exists():
        try:
            df = pd.read_excel(archivo, sheet_name='METAR SPJC')
            
            registros = []
            for _, row in df.iterrows():
                registro = row.to_dict()
                
                if 'DIA' in registro:
                    registro['Día'] = str(registro['DIA']).zfill(2)
                if 'HORA' in registro:
                    registro['Hora'] = str(registro['HORA']).zfill(4)
                if 'TIPO' in registro:
                    registro['Tipo'] = registro['TIPO']
                if 'METAR' in registro:
                    registro['METAR_Completo'] = registro['METAR']
                
                registros.append(registro)
            
            return registros
        except Exception as e:
            st.error(f"Error al cargar archivo mensual: {e}")
            return []
    return []

def guardar_registros_mes(registros):
    """Guarda todos los registros en el archivo mensual"""
    if not registros:
        return False, "No hay registros para guardar"
    
    try:
        archivo = DIRECTORIO_DATOS / obtener_nombre_archivo_mensual()
        
        df = pd.DataFrame(registros)
        
        df = df.rename(columns={
            'Día': 'DIA',
            'Hora': 'HORA',
            'Tipo': 'TIPO',
            'Dirección_Viento': 'DIR VIENTO',
            'Intensidad_Viento': 'INTENSIDAD',
            'Variación_Viento': 'VARIACION',
            'Visibilidad_Original': 'VIS (ORIGINAL)',
            'Visibilidad_Metros': 'VIS (CODIGO)',
            'Visibilidad_Mínima': 'VIS MIN',
            'RVR': 'RVR',
            'Fenómeno_Texto': 'FENOMENO',
            'Fenómeno_Código': 'WX',
            'Nubes_Texto': 'NUBOSIDAD',
            'Nubes_Código': 'CLD',
            'Temperatura': 'TEMP °C',
            'Punto_Rocío': 'ROCÍO °C',
            'Humedad_Relativa_%': 'HR %',
            'QNH': 'QNH',
            'Presión_Estación': 'PRESION',
            'Info_Suplementaria': 'RMK',
            'METAR_Completo': 'METAR'
        })
        
        columnas_excel = [
            'DIA', 'HORA', 'TIPO', 'DIR VIENTO', 'INTENSIDAD', 'VARIACION',
            'VIS (ORIGINAL)', 'VIS (CODIGO)', 'VIS MIN', 'RVR',
            'FENOMENO', 'WX', 'NUBOSIDAD', 'CLD',
            'TEMP °C', 'ROCÍO °C', 'HR %', 'QNH', 'PRESION', 'RMK', 'METAR'
        ]
        
        columnas_disponibles = [col for col in columnas_excel if col in df.columns]
        df = df[columnas_disponibles]
        
        df['DIA'] = df['DIA'].astype(str).str.zfill(2)
        df['HORA'] = df['HORA'].astype(str).str.zfill(4)
        df = df.sort_values(['DIA', 'HORA'], ascending=[True, True])
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='METAR SPJC', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['METAR SPJC']
            
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            for col in range(1, len(columnas_disponibles) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                header_cell = worksheet.cell(row=1, column=col)
                max_length = len(str(header_cell.value)) if header_cell.value else 0
                
                for row in range(2, min(len(df) + 2, 102)):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = min(max_length + 2, 70)
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 8)
            
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='0B3D91', end_color='0B3D91', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for col in range(1, len(columnas_disponibles) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            worksheet.row_dimensions[1].height = 30
            
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            for row in range(2, len(df) + 2):
                tipo_reporte = worksheet.cell(row=row, column=3).value
                for col in range(1, len(columnas_disponibles) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if tipo_reporte == 'SPECI':
                        cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                        cell.font = Font(name='Calibri', size=10, bold=True)
                    else:
                        cell.font = Font(name='Calibri', size=10)
            
            worksheet.freeze_panes = 'A2'
        
        output.seek(0)
        
        with open(archivo, 'wb') as f:
            f.write(output.getvalue())
        
        return True, f"✅ {len(registros)} registros guardados en {archivo.name}"
        
    except Exception as e:
        return False, f"Error al guardar: {str(e)}"

# ============================================
# INICIALIZAR ESTADO DE SESIÓN
# ============================================
if 'registros' not in st.session_state:
    st.session_state.registros = cargar_registros_mes()

if 'historial' not in st.session_state:
    st.session_state.historial = []

if 'contador' not in st.session_state:
    st.session_state.contador = len(st.session_state.registros)

if 'campos_inicializados' not in st.session_state:
    st.session_state.campos_inicializados = False

# ============================================
# FUNCIÓN PARA LIMPIAR CAMPOS - HORA VACÍA
# ============================================
def limpiar_campos():
    """Limpia todos los campos del formulario - Hora queda VACÍA"""
    st.session_state.dia = datetime.now(timezone.utc).strftime("%d")
    st.session_state.hora = ""
    st.session_state.tipo = "METAR"
    st.session_state.dir_viento = ""
    st.session_state.int_viento = ""
    st.session_state.var_viento = ""
    st.session_state.vis = ""
    st.session_state.vis_min = ""
    st.session_state.rvr = ""
    st.session_state.fenomeno = ""
    st.session_state.nubes = ""
    st.session_state.temp = ""
    st.session_state.rocio = ""
    st.session_state.hr = ""
    st.session_state.qnh = ""
    st.session_state.presion = ""
    st.session_state.suplementaria = ""
    st.session_state.campos_inicializados = True

# ============================================
# INICIALIZAR CAMPOS VACÍOS - HORA VACÍA
# ============================================
if not st.session_state.campos_inicializados:
    st.session_state.dia = datetime.now(timezone.utc).strftime("%d")
    st.session_state.hora = ""
    st.session_state.tipo = "METAR"
    st.session_state.dir_viento = ""
    st.session_state.int_viento = ""
    st.session_state.var_viento = ""
    st.session_state.vis = ""
    st.session_state.vis_min = ""
    st.session_state.rvr = ""
    st.session_state.fenomeno = ""
    st.session_state.nubes = ""
    st.session_state.temp = ""
    st.session_state.rocio = ""
    st.session_state.hr = ""
    st.session_state.qnh = ""
    st.session_state.presion = ""
    st.session_state.suplementaria = ""
    st.session_state.campos_inicializados = True

# ============================================
# FUNCIÓN PARA ACTUALIZAR O INSERTAR REGISTRO
# ============================================
def actualizar_o_insertar_registro(registros, nuevo_registro):
    """
    Busca un registro existente con el mismo día y hora.
    Si existe, lo reemplaza y guarda automáticamente en el archivo mensual.
    """
    dia_nuevo = str(nuevo_registro.get('Día', '')).zfill(2)
    hora_nueva = str(nuevo_registro.get('Hora', '')).zfill(4)
    
    dia_hora_clave = f"{dia_nuevo}_{hora_nueva}"
    accion = "insertado"
    
    for i, registro in enumerate(registros):
        dia_existente = str(registro.get('Día', '')).zfill(2)
        hora_existente = str(registro.get('Hora', '')).zfill(4)
        
        clave_existente = f"{dia_existente}_{hora_existente}"
        
        if clave_existente == dia_hora_clave:
            registros[i] = nuevo_registro
            accion = "actualizado"
            break
    else:
        registros.insert(0, nuevo_registro)
    
    guardar_registros_mes(registros)
    
    return accion

# ============================================
# FUNCIONES DE PROCESAMIENTO - VIENTO
# ============================================
def procesar_viento(direccion, intensidad, variacion):
    """
    PROCESAMIENTO DE VIENTO - REGLAS CORPAC PERÚ
    Caso 1: Variación ≥60° y <180° con viento <3kt → VRBxxKT
    Caso 2: Variación ≥60° y <180° con viento ≥3kt → dddffKT bbbVnnn
    Caso 3: Variación ≥180° → VRBxxKT
    Caso 4: Variación <60° → NO se incluye en el METAR final
    """
    dir_int = int(direccion)
    intensidad_str = str(intensidad).upper().strip()
    
    # Procesar ráfagas
    if 'G' in intensidad_str:
        if 'G' in intensidad_str and not ' ' in intensidad_str.replace('G', ''):
            base_int, gust_int = intensidad_str.split('G')
            int_base = int(base_int)
            int_gust = int(gust_int)
            intensidad_metar = f"{int_base:02d}G{int_gust:02d}"
        else:
            parts = intensidad_str.replace('G', ' ').split()
            int_base = int(parts[0])
            int_gust = int(parts[1])
            intensidad_metar = f"{int_base:02d}G{int_gust:02d}"
    else:
        int_base = int(intensidad_str)
        intensidad_metar = f"{int_base:02d}"
    
    # Si NO hay variación
    if not variacion:
        return f"{dir_int:03d}{intensidad_metar}KT"
    
    try:
        # Extraer valores de variación (formato: bbbVnnn)
        variacion = variacion.upper().replace(' ', '')
        if 'V' not in variacion:
            return f"{dir_int:03d}{intensidad_metar}KT"
        
        desde, hasta = map(int, variacion.split('V'))
        
        # Calcular diferencia CIRCULAR
        diff1 = abs(hasta - desde)
        diff2 = 360 - diff1
        diferencia = min(diff1, diff2)
        
        # CASO 4: Variación < 60° - NO incluir en METAR
        if diferencia < 60:
            return f"{dir_int:03d}{intensidad_metar}KT"
        
        # CASO 3: Variación ≥ 180°
        if diferencia >= 180:
            return f"VRB{intensidad_metar}KT"
        
        # CASO 1 y 2: Variación ≥ 60° y < 180°
        if diferencia >= 60:
            if int_base < 3:
                # CASO 1: Viento < 3kt
                return f"VRB{intensidad_metar}KT"
            else:
                # CASO 2: Viento ≥ 3kt
                if diff1 <= 180:
                    return f"{dir_int:03d}{intensidad_metar}KT {desde:03d}V{hasta:03d}"
                else:
                    # Para casos como 340V080, mostrar como 080V340
                    return f"{dir_int:03d}{intensidad_metar}KT {hasta:03d}V{desde:03d}"
        
        return f"{dir_int:03d}{intensidad_metar}KT"
        
    except Exception as e:
        return f"{dir_int:03d}{intensidad_metar}KT"
        
# ============================================
# FUNCIONES DE PROCESAMIENTO - VISIBILIDAD
# ============================================
def convertir_visibilidad(vis_texto):
    """Convierte visibilidad a metros"""
    vis_texto = vis_texto.strip().upper()
    if not vis_texto:
        raise ValueError("Visibilidad es obligatoria")
    
    try:
        if vis_texto.endswith("KM"):
            km = float(vis_texto[:-2])
            return 9999 if km >= 10 else int(km * 1000)
        elif vis_texto.endswith("M"):
            return int(vis_texto[:-1])
        else:
            metros = int(vis_texto)
            return 9999 if metros >= 10000 else metros
    except:
        raise ValueError("Formato de visibilidad inválido")

def procesar_visibilidad_minima(vis_min_texto, vis_m):
    """
    Procesa visibilidad mínima con cuadrantes
    Cuadrantes: N, NE, E, SE, S, SW, W, NW
    """
    if not vis_min_texto:
        return "", ""
    
    vis_min_texto = vis_min_texto.strip().upper()
    
    if vis_min_texto.endswith('NW'):
        valor = vis_min_texto[:-2]
        cuadrante = 'NW'
    elif vis_min_texto.endswith('NE'):
        valor = vis_min_texto[:-2]
        cuadrante = 'NE'
    elif vis_min_texto.endswith('SW'):
        valor = vis_min_texto[:-2]
        cuadrante = 'SW'
    elif vis_min_texto.endswith('SE'):
        valor = vis_min_texto[:-2]
        cuadrante = 'SE'
    elif vis_min_texto.endswith('N'):
        valor = vis_min_texto[:-1]
        cuadrante = 'N'
    elif vis_min_texto.endswith('S'):
        valor = vis_min_texto[:-1]
        cuadrante = 'S'
    elif vis_min_texto.endswith('E'):
        valor = vis_min_texto[:-1]
        cuadrante = 'E'
    elif vis_min_texto.endswith('W'):
        valor = vis_min_texto[:-1]
        cuadrante = 'W'
    else:
        valor = vis_min_texto
        cuadrante = ''
    
    try:
        if valor.endswith("KM"):
            km = float(valor[:-2])
            vis_min_m = 9999 if km >= 10 else int(km * 1000)
        elif valor.endswith("M"):
            vis_min_m = int(valor[:-1])
        else:
            vis_min_m = int(valor)
            vis_min_m = 9999 if vis_min_m >= 10000 else vis_min_m
        
        es_valida = False
        if vis_min_m < 1500:
            es_valida = True
        if vis_min_m < (vis_m * 0.5) and vis_min_m < 5000:
            es_valida = True
        
        if not es_valida:
            return "", "⚠️ No cumple reglas de visibilidad mínima"
        
        if cuadrante:
            return f"{vis_min_m:04d}{cuadrante}", ""
        else:
            return f"{vis_min_m:04d}", ""
        
    except:
        return "", "❌ Formato inválido"

# ============================================
# FUNCIÓN CORREGIDA: RVR - EXACTAMENTE COMO EN AWOS
# ============================================
def procesar_rvr(rvr_texto):
    """
    Procesa RVR - Formato SPJC: R{designador}/{valor}
    El operador COPIA EXACTAMENTE lo que muestra el AWOS
    NO se aplican validaciones, formateos ni límites
    
    Ejemplos válidos:
    - R32/0400      (pista 32, 400m)
    - R12R/1700     (pista 12 derecha, 1700m)  
    - R10/M0050     (pista 10, menor a 50m)
    - R14L/P2000    (pista 14 izquierda, mayor a 2000m)
    - R15/0600U     (pista 15, 600m tendencia up)
    - R13/0800D     (pista 13, 800m tendencia down)
    - R11/0450N     (pista 11, 450m sin cambio)
    """
    if not rvr_texto:
        return ""
    
    # Solo limpiar espacios, mantener todo lo demás exactamente igual
    rvr_texto = rvr_texto.strip()
    
    # Devolver exactamente lo que ingresó el operador
    # Sin modificar absolutamente nada
    return rvr_texto

# ============================================
# FUNCIONES DE PROCESAMIENTO - FENÓMENOS
# ============================================
def codificar_fenomenos(texto):
    """CODIFICADOR DE FENÓMENOS"""
    if not texto:
        return ""
    
    texto_lower = texto.lower().strip()
    
    if any(x in texto_lower for x in ["niebla parcial", "prfg", "pr fg", "parcial"]):
        return "PRFG"
    if any(x in texto_lower for x in ["niebla en la vecindad", "vcfg", "vc fg", "vecindad"]):
        return "VCFG"
    if any(x in texto_lower for x in ["niebla en bancos", "bcfg", "bc fg", "bancos"]):
        return "BCFG"
    if any(x in texto_lower for x in ["niebla baja", "mifg", "mi fg", "baja"]):
        return "MIFG"
    
    intensidades = {"ligera": "-", "ligero": "-", "leve": "-", "débil": "-",
                   "fuerte": "+", "intensa": "+", "intenso": "+", "severa": "+"}
    
    descriptores = {"sh": "SH", "chubasco": "SH", "ts": "TS", "tormenta": "TS",
                   "fz": "FZ", "helada": "FZ", "congelante": "FZ"}
    
    fenomenos = {
        "lluvia": "RA", "llovizna": "DZ", "niebla": "FG", "neblina": "BR",
        "nieve": "SN", "granizo": "GR", "cellisca": "GS", "tormenta": "TS",
        "polvo": "DU", "arena": "SA", "humo": "FU", "ceniza": "VA", "calima": "HZ"
    }
    
    partes = [p.strip() for p in texto.split(",")][:3]
    resultados = []
    
    for parte in partes:
        if not parte:
            continue
        parte_lower = parte.lower()
        codigo = ""
        descriptor = ""
        intensidad = ""
        
        for d_texto, d_codigo in descriptores.items():
            if d_texto in parte_lower:
                descriptor = d_codigo
                parte_lower = parte_lower.replace(d_texto, "").strip()
                break
        
        for i_texto, i_codigo in intensidades.items():
            if i_texto in parte_lower:
                intensidad = i_codigo
                parte_lower = parte_lower.replace(i_texto, "").strip()
                break
        
        for f_texto, f_codigo in fenomenos.items():
            if f_texto in parte_lower:
                codigo = f_codigo
                break
        
        if codigo:
            if descriptor:
                codigo = descriptor + codigo
            if intensidad:
                codigo = intensidad + codigo
            resultados.append(codigo)
    
    return " ".join(resultados) if resultados else ""

# ============================================
# FUNCIONES DE PROCESAMIENTO - NUBES
# ============================================
def interpretar_nubes(texto, vis_m, fenomeno):
    """CODIFICADOR DE NUBES - ESTÁNDAR CORPAC"""
    texto = texto.strip().upper()
    
    if texto in ["DESPEJADO", "SKC", "CLR", "", "NSC", "SIN NUBES", "NO NUBES"]:
        return "NSC"
    
    if vis_m >= 9999 and not fenomeno.strip() and texto in ["NSC", "SKC", "CLR", "DESPEJADO"]:
        return "CAVOK"
    
    tipos_nubes = {
        "CU": "CU", "SC": "SC", "ST": "ST", "CB": "CB", "TCU": "TCU",
        "AC": "AC", "AS": "AS", "NS": "NS", "CI": "CI"
    }
    
    capas = texto.split(",")
    codigos_nubes = []
    
    for capa in capas[:4]:
        capa = capa.strip()
        if not capa:
            continue
        
        patron = r'(\d+)\s+([A-Z]{2,4})\s+(\d+)(?:M)?'
        match = re.search(patron, capa)
        
        if match:
            cantidad = int(match.group(1))
            tipo = match.group(2)
            altura = int(match.group(3))
            tipo_nube = tipos_nubes.get(tipo, tipo)
            
            if altura <= 3000:
                if altura % 30 != 0:
                    altura = (altura // 30) * 30
                altura_cientos = altura // 30
            else:
                if altura % 1000 != 0:
                    altura = (altura // 1000) * 1000
                altura_cientos = (altura // 1000) * 32
            
            altura_cientos = min(max(altura_cientos, 1), 999)
            
            if cantidad <= 2:
                cod_cant = "FEW"
            elif cantidad <= 4:
                cod_cant = "SCT"
            elif cantidad <= 7:
                cod_cant = "BKN"
            else:
                cod_cant = "OVC"
            
            codigo = f"{cod_cant}{altura_cientos:03d}"
            if tipo_nube in ["CB", "TCU"]:
                codigo += tipo_nube
            codigos_nubes.append(codigo)
    
    return " ".join(codigos_nubes[:4]) if codigos_nubes else "NSC"

def verificar_cavok(vis_m, fenomeno, nubes):
    """Verifica condiciones para CAVOK"""
    return (vis_m >= 9999 and not fenomeno.strip() and nubes in ["NSC", "SKC", "CLR"])

# ============================================
# FUNCIONES DE VALIDACIÓN
# ============================================
def validar_hora(hora_str):
    if not hora_str:
        raise ValueError("Formato HHMM (ej: 1230)")
    if len(hora_str) != 4 or not hora_str.isdigit():
        raise ValueError("Hora debe ser HHMM (4 dígitos)")
    h = int(hora_str[:2])
    m = int(hora_str[2:])
    if h > 23 or m > 59:
        raise ValueError("Hora inválida")
    return hora_str

def validar_intensidad_viento(intensidad_str):
    intensidad_str = str(intensidad_str).strip().upper()
    if not intensidad_str:
        raise ValueError("Intensidad de viento requerida")
    
    intensidad_str = intensidad_str.replace(' G ', 'G').replace(' G', 'G').replace('G ', 'G')
    
    if 'G' in intensidad_str:
        partes = intensidad_str.split('G')
        if len(partes) != 2:
            raise ValueError("Formato de ráfagas inválido. Use: 15G25")
        base = int(partes[0])
        rafaga = int(partes[1])
        if base < 0 or base > 100:
            raise ValueError("Intensidad base fuera de rango (0-100)")
        if rafaga < base:
            raise ValueError("Ráfaga debe ser mayor o igual a intensidad base")
        if rafaga > 150:
            raise ValueError("Ráfaga excede límite (150 KT)")
        diferencia = rafaga - base
        if diferencia < 10:
            raise ValueError(f"Ráfaga requiere diferencia ≥10 KT (actual: {diferencia} KT)")
        return intensidad_str
    else:
        intensidad = int(intensidad_str)
        if intensidad < 0 or intensidad > 100:
            raise ValueError("Intensidad fuera de rango (0-100)")
        return intensidad_str

def validar_numero(valor, min_val, max_val, nombre):
    if not valor:
        raise ValueError(f"{nombre} es obligatorio")
    try:
        num = float(valor)
        if not (min_val <= num <= max_val):
            raise ValueError(f"{nombre} fuera de rango ({min_val}-{max_val})")
        return num
    except:
        raise ValueError(f"{nombre} inválido")

def validar_temp_rocio(temp, rocio):
    if float(rocio) > float(temp):
        raise ValueError(f"Rocío ({rocio}°C) no puede ser > Temperatura ({temp}°C)")
    return True

# ============================================
# FUNCIÓN PRINCIPAL DE GENERACIÓN
# ============================================
def generar_metar(datos):
    try:
        if not datos['dir_viento'] or not datos['int_viento']:
            raise ValueError("Dirección e intensidad del viento son obligatorias")
        if not datos['vis']:
            raise ValueError("Visibilidad es obligatoria")
        if not datos['temp'] or not datos['rocio'] or not datos['qnh']:
            raise ValueError("Temperatura, Rocío y QNH son obligatorios")
        
        hora = validar_hora(datos['hora'])
        int_viento = validar_intensidad_viento(datos['int_viento'])
        viento = procesar_viento(datos['dir_viento'], int_viento, datos['var_viento'])
        vis_m = convertir_visibilidad(datos['vis'])
        
        vis_min_codigo = ""
        if datos['vis_min']:
            vis_min_codigo, vis_min_error = procesar_visibilidad_minima(datos['vis_min'], vis_m)
            if vis_min_error:
                raise ValueError(vis_min_error)
        
        # RVR - AHORA DEVUELVE EXACTAMENTE LO QUE INGRESÓ EL OPERADOR
        rvr_codigo = procesar_rvr(datos['rvr'])
        
        fenomeno = codificar_fenomenos(datos['fenomeno'])
        nubes = interpretar_nubes(datos['nubes'], vis_m, fenomeno)
        
        temp = validar_numero(datos['temp'], -10, 40, "Temperatura")
        rocio = validar_numero(datos['rocio'], -10, 40, "Punto de rocío")
        validar_temp_rocio(temp, rocio)
        qnh = validar_numero(datos['qnh'], 850, 1100, "QNH")
        
        es_cavok = verificar_cavok(vis_m, fenomeno, nubes)
        
        metar_parts = [f"{datos['tipo']} SPJC {datos['dia']}{hora}Z {viento}"]
        
        if es_cavok:
            metar_parts.append("CAVOK")
        else:
            metar_parts.append(f"{vis_m:04d}")
            if vis_min_codigo:
                metar_parts.append(vis_min_codigo)
            if rvr_codigo:
                metar_parts.append(rvr_codigo)  # Se agrega exactamente como se ingresó
            if fenomeno:
                metar_parts.append(fenomeno)
            metar_parts.append(nubes)
        
        metar_parts.append(f"{int(round(temp)):02d}/{int(round(rocio)):02d} Q{int(round(qnh))}")
        
        if datos['suplementaria']:
            metar_parts.append(datos['suplementaria'].upper())
        
        metar_completo = " ".join(metar_parts) + "="
        
        registro = {
            'Día': str(datos['dia']).zfill(2),
            'Hora': hora,
            'Tipo': datos['tipo'],
            'Dirección_Viento': datos['dir_viento'],
            'Intensidad_Viento': datos['int_viento'],
            'Variación_Viento': datos['var_viento'],
            'Visibilidad_Original': datos['vis'],
            'Visibilidad_Metros': vis_m,
            'Visibilidad_Mínima': vis_min_codigo,
            'RVR': rvr_codigo,  # Se guarda exactamente como se ingresó
            'Fenómeno_Texto': datos['fenomeno'],
            'Fenómeno_Código': fenomeno,
            'Nubes_Texto': datos['nubes'],
            'Nubes_Código': "CAVOK" if es_cavok else nubes,
            'Temperatura': temp,
            'Punto_Rocío': rocio,
            'Humedad_Relativa_%': datos['hr'] if datos['hr'] else "",
            'QNH': qnh,
            'Presión_Estación': datos['presion'],
            'Info_Suplementaria': datos['suplementaria'],
            'METAR_Completo': metar_completo
        }
        
        return {
            'success': True,
            'metar': metar_completo,
            'registro': registro
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

# ============================================
# FUNCIÓN PARA EXPORTAR EXCEL (VERSIÓN CORREGIDA)
# ============================================
def exportar_a_excel(registros):
    """Genera un archivo Excel para descargar con columna METAR ajustada"""
    if not registros:
        return None, "No hay registros para exportar"
    
    try:
        df = pd.DataFrame(registros)
        
        df = df.rename(columns={
            'Día': 'DIA',
            'Hora': 'HORA',
            'Tipo': 'TIPO',
            'Dirección_Viento': 'DIR VIENTO',
            'Intensidad_Viento': 'INTENSIDAD',
            'Variación_Viento': 'VARIACION',
            'Visibilidad_Original': 'VIS (ORIGINAL)',
            'Visibilidad_Metros': 'VIS (CODIGO)',
            'Visibilidad_Mínima': 'VIS MIN',
            'RVR': 'RVR',
            'Fenómeno_Texto': 'FENOMENO',
            'Fenómeno_Código': 'WX',
            'Nubes_Texto': 'NUBOSIDAD',
            'Nubes_Código': 'CLD',
            'Temperatura': 'TEMP °C',
            'Punto_Rocío': 'ROCÍO °C',
            'Humedad_Relativa_%': 'HR %',
            'QNH': 'QNH',
            'Presión_Estación': 'PRESION',
            'Info_Suplementaria': 'RMK',
            'METAR_Completo': 'METAR'
        })
        
        columnas_excel = [
            'DIA', 'HORA', 'TIPO', 'DIR VIENTO', 'INTENSIDAD', 'VARIACION',
            'VIS (ORIGINAL)', 'VIS (CODIGO)', 'VIS MIN', 'RVR',
            'FENOMENO', 'WX', 'NUBOSIDAD', 'CLD',
            'TEMP °C', 'ROCÍO °C', 'HR %', 'QNH', 'PRESION', 'RMK', 'METAR'
        ]
        
        columnas_disponibles = [col for col in columnas_excel if col in df.columns]
        df = df[columnas_disponibles]
        
        df['DIA'] = df['DIA'].astype(str).str.zfill(2)
        df['HORA'] = df['HORA'].astype(str).str.zfill(4)
        df = df.sort_values(['DIA', 'HORA'], ascending=[True, True])
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='METAR SPJC', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['METAR SPJC']
            
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # ============================================
            # CORRECCIÓN: Calcular ancho de columnas correctamente
            # ============================================
            for col in range(1, len(columnas_disponibles) + 1):
                column_letter = get_column_letter(col)
                
                # Obtener el nombre de la columna para tratamientos especiales
                col_name = columnas_disponibles[col-1]
                
                # Empezar con el ancho del encabezado
                header_cell = worksheet.cell(row=1, column=col)
                max_length = len(str(header_cell.value)) if header_cell.value else 0
                
                # Revisar todas las filas de datos (hasta 1000 filas para rendimiento)
                for row in range(2, min(len(df) + 2, 1002)):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                # ============================================
                # AJUSTE ESPECIAL PARA COLUMNA METAR
                # ============================================
                if col_name == 'METAR':
                    # La columna METAR necesita ser más ancha
                    adjusted_width = min(max_length + 5, 120)  # Máximo 120 caracteres
                else:
                    # Para las demás columnas, ancho estándar
                    adjusted_width = min(max_length + 2, 50)   # Máximo 50 caracteres
                
                # Ancho mínimo de 8 para columnas muy cortas
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 8)
            
            # Aplicar estilos al encabezado
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='0B3D91', end_color='0B3D91', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for col in range(1, len(columnas_disponibles) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            worksheet.row_dimensions[1].height = 30
            
            # Bordes para todas las celdas
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Aplicar formato a las filas de datos
            for row in range(2, len(df) + 2):
                tipo_reporte = worksheet.cell(row=row, column=3).value  # Columna TIPO
                for col in range(1, len(columnas_disponibles) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    
                    if tipo_reporte == 'SPECI':
                        cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                        cell.font = Font(name='Calibri', size=10, bold=True)
                    else:
                        cell.font = Font(name='Calibri', size=10)
                    
                    # ============================================
                    # ALINEACIÓN ESPECIAL PARA METAR (izquierda)
                    # ============================================
                    col_name = columnas_disponibles[col-1]
                    if col_name == 'METAR':
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Congelar paneles
            worksheet.freeze_panes = 'A2'
        
        output.seek(0)
        return output, f"✅ {len(registros)} registros exportados"
        
    except Exception as e:
        return None, f"Error al exportar: {str(e)}"

# ============================================
# HEADER SIMPLIFICADO - SIN RELOJ
# ============================================
col_header1, col_header2 = st.columns([3, 1])

with col_header1:
    st.markdown("<h1 style='color: #0b3d91;'>REGISTRO DE OBSERVACIONES ORDINARIAS Y ESPECIALES</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color: #666;'>Aeropuerto Internacional Jorge Chávez - CORPAC Perú</p>", unsafe_allow_html=True)

with col_header2:
    hoy = datetime.now(timezone.utc).strftime('%d/%m/%Y')
    st.markdown(f"""
    <div style='text-align: right;'>
        <p style='color: #666; font-size: 14px;'>{hoy}</p>
        <p style='color: #0b3d91; font-size: 12px; font-weight: bold;'>{obtener_nombre_archivo_mensual()}</p>
        <p style='color: #999; font-size: 11px;'>CORPAC PERU</p>
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")

# ============================================
# INTERFAZ DE USUARIO
# ============================================
col_izq, col_der = st.columns([2, 1])

with col_izq:
    with st.form(key='metar_form'):
        # DATOS DEL REPORTE
        st.markdown("<div class='section-title'>DATOS DEL REPORTE</div>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            tipo = st.selectbox("Tipo", ["METAR", "SPECI"], key='tipo')
        with col2:
            dia = st.text_input("Dia", key='dia', help="Formato: DD (01-31)")
        with col3:
            hora = st.text_input("Hora UTC", key='hora', help="INGRESE HORA MANUALMENTE - Formato HHMM (ej: 1230)")
            
        st.markdown("---")
        
        # VIENTO
        st.markdown("<div class='section-title'>VIENTO</div>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            dir_viento = st.text_input("Direccion", key='dir_viento', help="Grados (0-360)")
        with col2:
            int_viento = st.text_input("Intensidad (KT)", key='int_viento', help="Nudos. Rafagas: 15G25")
        with col3:
            var_viento = st.text_input("Variacion", key='var_viento', help="Formato: 340V080")
        
        st.markdown("---")
        
        # VISIBILIDAD
        st.markdown("<div class='section-title'>VISIBILIDAD</div>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            vis = st.text_input("Visibilidad", key='vis', help="Ej: 10km, 5000m, 9999")
        with col2:
            vis_min = st.text_input("Visibilidad Minima", key='vis_min', help="Ej: 1200SW, 0800NE, 1500SE, 2000NW")
        with col3:
            rvr = st.text_input("RVR", key='rvr', help="R32/0400, R12R/1700, R10/M0050, R14L/P2000")
        
        st.markdown("---")
        
        # FENOMENOS Y NUBES
        st.markdown("<div class='section-title'>FENOMENOS Y NUBES</div>", unsafe_allow_html=True)
        fenomeno = st.text_input("Fenomeno", key='fenomeno', help="Ej: niebla parcial (PRFG), lluvia ligera (-RA)")
        nubes = st.text_input("Nubes", key='nubes', help="Ej: 8 ST 300M, 5 AC 5000M, CB 1500M")
        
        st.markdown("---")
        
        # TEMPERATURA, HUMEDAD Y PRESION
        st.markdown("<div class='section-title'>TEMPERATURA, HUMEDAD Y PRESION</div>", unsafe_allow_html=True)
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            temp = st.text_input("Temperatura del aire", key='temp', help="-10 a 40C")
        with col2:
            rocio = st.text_input("Temperatura de Rocio", key='rocio', help="≤ Temperatura")
        with col3:
            hr = st.text_input("HR %", key='hr', help="Humedad Relativa (0-100%)")
        with col4:
            qnh = st.text_input("QNH hPa", key='qnh', help="850-1100 hPa")
        with col5:
            presion = st.text_input("Presion Estacion", key='presion', help="Para el registro")
        
        st.markdown("---")
        
        # INFORMACION SUPLEMENTARIA
        st.markdown("<div class='section-title'>INFORMACION SUPLEMENTARIA</div>", unsafe_allow_html=True)
        suplementaria = st.text_input("", key='suplementaria', help="Ej: NOSIG RMK PP000")
        
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        with col1:
            generar = st.form_submit_button("GENERAR METAR", use_container_width=True)
        with col2:
            limpiar = st.form_submit_button("LIMPIAR CAMPOS", use_container_width=True)
        
        if limpiar:
            limpiar_campos()
            st.rerun()
        
        if generar:
            datos = {
                'tipo': tipo, 'dia': dia, 'hora': hora,
                'dir_viento': dir_viento, 'int_viento': int_viento, 'var_viento': var_viento,
                'vis': vis, 'vis_min': vis_min, 'rvr': rvr,
                'fenomeno': fenomeno, 'nubes': nubes,
                'temp': temp, 'rocio': rocio, 'hr': hr,
                'qnh': qnh, 'presion': presion, 'suplementaria': suplementaria
            }
            
            resultado = generar_metar(datos)
            
            if resultado['success']:
                accion = actualizar_o_insertar_registro(st.session_state.registros, resultado['registro'])
                
                dia_hora_clave = f"{str(resultado['registro']['Día']).zfill(2)}_{resultado['registro']['Hora']}"
                nuevo_historial = []
                
                for metar in st.session_state.historial:
                    match = re.search(r'SPJC (\d{2})(\d{4})Z', metar)
                    if match:
                        dia_hist = match.group(1)
                        hora_hist = match.group(2)
                        if f"{dia_hist}_{hora_hist}" != dia_hora_clave:
                            nuevo_historial.append(metar)
                    else:
                        nuevo_historial.append(metar)
                
                nuevo_historial.insert(0, resultado['metar'])
                st.session_state.historial = nuevo_historial[:20]
                st.session_state.contador = len(st.session_state.registros)
                
                if accion == "actualizado":
                    st.warning("METAR ACTUALIZADO - Reemplazo reporte existente con la misma fecha/hora")
                else:
                    st.success("METAR generado correctamente")
                
                st.session_state.ultimo_metar = resultado['metar']
                st.session_state.ultimo_tipo = tipo
                st.session_state.ultimo_registro = resultado['registro']  # Guardar para el panel detallado
            else:
                st.error(f"ERROR: {resultado['error']}")

with col_der:
    st.markdown("<div class='section-title'>📋 ÚLTIMO REPORTE GENERADO</div>", unsafe_allow_html=True)
    if 'ultimo_metar' in st.session_state and 'ultimo_registro' in st.session_state:
        reg = st.session_state.ultimo_registro
        tipo_ultimo = st.session_state.get('ultimo_tipo', 'METAR')
        
        # Panel detallado del METAR
        with st.container():
            if tipo_ultimo == "SPECI":
                st.markdown(f"""
                <div style='background: #FFE699; padding: 15px; border-radius: 8px; margin-bottom: 15px;'>
                    <span style='background: #FF9800; color: white; padding: 2px 8px; border-radius: 4px; font-weight: bold;'>SPECI</span>
                </div>
                """, unsafe_allow_html=True)
            
            # Mostrar componentes clave
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"**Tipo:** `{reg.get('Tipo', '')}`")
                st.markdown(f"**Viento:** `{reg.get('Dirección_Viento', '')}{reg.get('Intensidad_Viento', '')}KT {reg.get('Variación_Viento', '')}`")
                st.markdown(f"**Vis:** `{reg.get('Visibilidad_Original', '')}`")
                st.markdown(f"**RVR:** `{reg.get('RVR', '')}`")
            with col2:
                st.markdown(f"**WX:** `{reg.get('Fenómeno_Código', '')}`")
                st.markdown(f"**Nubes:** `{reg.get('Nubes_Código', '')}`")
                st.markdown(f"**Temp/Rocío/QNH:** `{reg.get('Temperatura', '')}/{reg.get('Punto_Rocío', '')} Q{reg.get('QNH', '')}`")
            
            st.markdown("---")
            st.markdown("**Código Completo:**")
            # Caja del METAR completo
            if tipo_ultimo == "SPECI":
                st.markdown(f"<div style='background: #FFE699; padding: 15px; border-radius: 5px; font-family: monospace; border-left: 5px solid #FFC000;'>{st.session_state.ultimo_metar}</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div class='metar-box'>{st.session_state.ultimo_metar}</div>", unsafe_allow_html=True)
    else:
        st.info("---")
    
    st.markdown("---")
    
    # ============================================
    # BOTÓN DE EXPORTACIÓN SIMPLIFICADO
    # ============================================
    st.markdown("<div class='section-title'>EXPORTAR</div>", unsafe_allow_html=True)
    
    if st.button("📥 Exportar METAR", use_container_width=True, type="primary"):
        if st.session_state.registros:
            excel_file, mensaje = exportar_a_excel(st.session_state.registros)
            if excel_file:
                # Usar SOLO formato YYYY_MM
                nombre_archivo = obtener_nombre_archivo_mensual()
                
                st.download_button(
                    label="✅ Confirmar Descarga",
                    data=excel_file,
                    file_name=nombre_archivo,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.success(mensaje)
            else:
                st.warning(mensaje)
        else:
            st.warning("No hay registros para exportar")
    
    if st.button("🗑️ Limpiar Memoria", use_container_width=True):
        st.session_state.registros = []
        st.session_state.historial = []
        st.session_state.contador = 0
        st.success("Memoria limpiada")
    
    st.markdown("---")
    st.metric("REGISTROS EN MEMORIA", st.session_state.contador)
    
    if st.session_state.registros:
        st.markdown("---")
        st.markdown("<div class='section-title'>RESUMEN POR DIA</div>", unsafe_allow_html=True)
        df_resumen = pd.DataFrame(st.session_state.registros)
        if 'Día' in df_resumen.columns:
            resumen_dias = df_resumen.groupby('Día').size().reset_index(name='Cantidad')
            resumen_dias = resumen_dias.sort_values('Día')
            for _, row in resumen_dias.iterrows():
                st.markdown(f"**Dia {str(row['Día']).zfill(2)}:** {row['Cantidad']} reportes")
    
    st.markdown("---")
    st.markdown("<div class='section-title'>HISTORIAL</div>", unsafe_allow_html=True)
    if st.session_state.historial:
        for metar in st.session_state.historial[:10]:
            if "SPECI" in metar:
                st.markdown(f"<div style='background: #FFE699; padding: 8px; margin-bottom: 5px; border-radius: 3px; font-family: monospace; font-size: 12px; border-left: 3px solid #FFC000;'>{metar}</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='background: #f0f0f0; padding: 8px; margin-bottom: 5px; border-radius: 3px; font-family: monospace; font-size: 12px; border-left: 3px solid #0b3d91;'>{metar}</div>", unsafe_allow_html=True)
    else:
        st.info("No hay METARs en el historial")

# ============================================
# FOOTER
# ============================================
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666; padding: 20px;'>
    <p>METAR Digital - CORPAC Peru | Aeropuerto Internacional Jorge Chavez (SPJC)</p>
    <p style='font-size: 0.8rem;'>RVR: Copiar exactamente del AWOS - Formato: R32/0400, R12R/1700, R10/M0050, R14L/P2000</p>
</div>
""", unsafe_allow_html=True)

# ============================================
# GUARDADO AUTOMATICO
# ============================================
if st.session_state.registros:
    guardar_registros_mes(st.session_state.registros)
