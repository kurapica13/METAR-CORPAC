"""
METAR DIGITAL - VERSIÓN FLASK
Aeropuerto Internacional Jorge Chávez (SPJC) - CORPAC Perú
"""

from flask import Flask, render_template, request, session, redirect, url_for, send_file, jsonify
from datetime import datetime, timezone
from pathlib import Path
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
import pandas as pd
import re
import hmac
import os

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = "corpac_spjc_2024_secreto"   # Cambia esto en producción

DIRECTORIO_DATOS = Path("datos_metar")
DIRECTORIO_DATOS.mkdir(exist_ok=True)

# Usuarios y contraseñas (puedes editar aquí)
USUARIOS = {
    "admin":  "corpac2024",
    "metar":  "spjc2024",
}

FENOMENOS_OPCIONES = {
    "Lluvia ligera (-RA)":        "-RA",
    "Lluvia moderada (RA)":       "RA",
    "Lluvia fuerte (+RA)":        "+RA",
    "Llovizna ligera (-DZ)":      "-DZ",
    "Llovizna moderada (DZ)":     "DZ",
    "Lluvia con tormenta (TSRA)": "TSRA",
    "Chubasco de lluvia (SHRA)":  "SHRA",
    "Nieve ligera (-SN)":         "-SN",
    "Nieve moderada (SN)":        "SN",
    "Granizo (GR)":               "GR",
    "Niebla (FG)":                "FG",
    "Neblina (BR)":               "BR",
    "Niebla parcial (PRFG)":      "PRFG",
    "Niebla en bancos (BCFG)":    "BCFG",
    "Niebla baja (MIFG)":         "MIFG",
    "Niebla en vecindad (VCFG)":  "VCFG",
    "Tormenta eléctrica (TS)":    "TS",
    "Polvo en suspensión (DU)":   "DU",
    "Arena (SA)":                 "SA",
    "Humo (FU)":                  "FU",
}

TIPOS_NUBE   = ["ST", "SC", "CU", "TCU", "CB", "AC", "AS", "NS", "CI", "FS"]
OCTAS_LABELS = {
    1: "1 octa (FEW)", 2: "2 octas (FEW)",
    3: "3 octas (SCT)", 4: "4 octas (SCT)",
    5: "5 octas (BKN)", 6: "6 octas (BKN)", 7: "7 octas (BKN)",
    8: "8 octas (OVC)",
}


# ─────────────────────────────────────────────
# HELPERS DE SESIÓN
# ─────────────────────────────────────────────
def sesion_init():
    """Inicializa los datos de sesión si no existen."""
    if "registros" not in session:
        session["registros"] = cargar_registros_mes()
    if "historial" not in session:
        session["historial"] = []
    if "fenomenos_lista" not in session:
        session["fenomenos_lista"] = []
    if "nubes_lista" not in session:
        session["nubes_lista"] = []
    if "ultimo_metar" not in session:
        session["ultimo_metar"] = None
    if "ultimo_tipo" not in session:
        session["ultimo_tipo"] = None


# ─────────────────────────────────────────────
# ARCHIVOS EXCEL
# ─────────────────────────────────────────────
def obtener_nombre_archivo():
    ahora = datetime.now(timezone.utc)
    return f"SPJC_METAR_{ahora.strftime('%Y_%m')}.xlsx"

def cargar_registros_mes():
    archivo = DIRECTORIO_DATOS / obtener_nombre_archivo()
    if archivo.exists():
        try:
            df = pd.read_excel(archivo, sheet_name="METAR SPJC")
            registros = []
            for _, row in df.iterrows():
                r = row.to_dict()
                r["Día"]           = str(r.get("DIA",  "")).zfill(2)
                r["Hora"]          = str(r.get("HORA", "")).zfill(4)
                r["Tipo"]          = r.get("TIPO", "")
                r["METAR_Completo"]= r.get("METAR", "")
                registros.append(r)
            return registros
        except Exception:
            return []
    return []

def guardar_registros_mes(registros):
    if not registros:
        return
    try:
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        archivo = DIRECTORIO_DATOS / obtener_nombre_archivo()
        datos   = []
        for r in registros:
            datos.append({
                "DIA":           str(r.get("Día",  "")).zfill(2),
                "HORA":          str(r.get("Hora", "")).zfill(4),
                "TIPO":          r.get("Tipo", ""),
                "DIR VIENTO":    r.get("Dirección_Viento", ""),
                "INTENSIDAD":    r.get("Intensidad_Viento", ""),
                "VARIACION":     r.get("Variación_Viento", ""),
                "VIS (ORIGINAL)":r.get("Visibilidad_Original", ""),
                "VIS (CODIGO)":  r.get("Visibilidad_Metros", ""),
                "VIS MIN":       r.get("Visibilidad_Mínima", ""),
                "RVR":           r.get("RVR", ""),
                "FENOMENO":      r.get("Fenómeno_Texto", ""),
                "WX":            r.get("Fenómeno_Código", ""),
                "NUBOSIDAD":     r.get("Nubes_Texto", ""),
                "CLD":           r.get("Nubes_Código", ""),
                "TEMP °C":       r.get("Temperatura", ""),
                "ROCÍO °C":      r.get("Punto_Rocío", ""),
                "HR %":          r.get("Humedad_Relativa_%", ""),
                "QNH":           r.get("QNH", ""),
                "PRESION":       r.get("Presión_Estación", ""),
                "RMK":           r.get("Info_Suplementaria", ""),
                "METAR":         r.get("METAR_Completo", ""),
            })
        df = pd.DataFrame(datos).sort_values(["DIA", "HORA"])
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="METAR SPJC", index=False)
            ws = writer.sheets["METAR SPJC"]
            for col in range(1, len(df.columns) + 1):
                letter  = get_column_letter(col)
                max_len = max(
                    (len(str(ws.cell(row=r, column=col).value or ""))
                     for r in range(1, min(len(datos)+2, 102))), default=8)
                ws.column_dimensions[letter].width = min(max_len + 3, 80)
            hf = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            hb = PatternFill(start_color="0B3D91", end_color="0B3D91", fill_type="solid")
            ha = Alignment(horizontal="center", vertical="center")
            for col in range(1, len(df.columns) + 1):
                c = ws.cell(row=1, column=col)
                c.font, c.fill, c.alignment = hf, hb, ha
            ws.row_dimensions[1].height = 30
            border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )
            for row in range(2, len(datos) + 2):
                tipo_r = ws.cell(row=row, column=3).value
                for col in range(1, len(df.columns) + 1):
                    c = ws.cell(row=row, column=col)
                    c.border    = border
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    if tipo_r == "SPECI":
                        c.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                        c.font = Font(name="Calibri", size=10, bold=True)
                    else:
                        c.font = Font(name="Calibri", size=10)
            ws.freeze_panes = "A2"
        output.seek(0)
        with open(archivo, "wb") as f:
            f.write(output.getvalue())
    except Exception as e:
        print(f"Error guardando Excel: {e}")


# ─────────────────────────────────────────────
# LÓGICA METAR (mismas funciones que antes)
# ─────────────────────────────────────────────
def redondear_metar(valor):
    try:
        return int(Decimal(str(valor)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except Exception:
        return int(round(float(valor)))

def procesar_viento(direccion, intensidad, variacion):
    dir_int = int(direccion)
    intensidad_str = str(intensidad).upper().strip()
    if dir_int == 0 and intensidad_str == "00":
        return "00000KT"
    if "G" in intensidad_str:
        partes = intensidad_str.split("G")
        intensidad_metar = f"{int(partes[0]):02d}G{int(partes[1]):02d}"
        int_base = int(partes[0])
    else:
        int_base = int(intensidad_str)
        intensidad_metar = f"{int_base:02d}"
    if not variacion:
        return f"{dir_int:03d}{intensidad_metar}KT"
    try:
        variacion = variacion.upper().replace(" ", "")
        if "V" not in variacion:
            return f"{dir_int:03d}{intensidad_metar}KT"
        desde, hasta = map(int, variacion.split("V"))
        diff1 = abs(hasta - desde)
        diferencia = diff1 if desde < hasta else 360 - diff1
        if diferencia < 60:
            return f"{dir_int:03d}{intensidad_metar}KT"
        if diferencia >= 180 or int_base < 3:
            return f"VRB{intensidad_metar}KT"
        d1, d2 = (desde, hasta) if desde < hasta else (hasta, desde)
        return f"{dir_int:03d}{intensidad_metar}KT {d1:03d}V{d2:03d}"
    except Exception:
        return f"{dir_int:03d}{intensidad_metar}KT"

def convertir_visibilidad(vis_texto):
    vis_texto = vis_texto.strip().upper()
    if not vis_texto:
        raise ValueError("Visibilidad es obligatoria")
    try:
        if vis_texto.endswith("KM"):
            km = float(vis_texto[:-2])
            return 9999 if km >= 10 else int(km * 1000)
        elif vis_texto.endswith("M"):
            return int(vis_texto[:-1])
        else:
            metros = int(vis_texto)
            return 9999 if metros >= 10000 else metros
    except Exception:
        raise ValueError("Formato de visibilidad inválido")

def procesar_visibilidad_minima(vis_min_texto, vis_m):
    if not vis_min_texto:
        return "", ""
    vis_min_texto = vis_min_texto.strip().upper()
    cuadrante = ""
    valor = vis_min_texto
    for cq in ["NW", "NE", "SW", "SE", "N", "S", "E", "W"]:
        if vis_min_texto.endswith(cq):
            cuadrante = cq
            valor = vis_min_texto[:-len(cq)]
            break
    try:
        if valor.endswith("KM"):
            vis_min_m = 9999 if float(valor[:-2]) >= 10 else int(float(valor[:-2]) * 1000)
        elif valor.endswith("M"):
            vis_min_m = int(valor[:-1])
        else:
            vis_min_m = int(valor)
            vis_min_m = 9999 if vis_min_m >= 10000 else vis_min_m
        if not (vis_min_m < 1500 or (vis_min_m < vis_m * 0.5 and vis_min_m < 5000)):
            return "", "No cumple reglas de visibilidad mínima"
        return f"{vis_min_m:04d}{cuadrante}", ""
    except Exception:
        return "", "Formato inválido"

def procesar_rvr(rvr_texto):
    return rvr_texto.strip() if rvr_texto else ""

def interpretar_nubes_lista(nubes_lista, vis_m, fenomeno):
    if not nubes_lista:
        if vis_m >= 9999 and not fenomeno.strip():
            return "CAVOK"
        return "NSC"
    codigos = []
    for capa in nubes_lista[:4]:
        octas     = int(capa.get("octas", 0))
        tipo_nube = capa.get("tipo", "SC").upper()
        altura_m  = int(capa.get("altura_m", 300))
        if octas <= 2:   cod = "FEW"
        elif octas <= 4: cod = "SCT"
        elif octas <= 7: cod = "BKN"
        else:            cod = "OVC"
        altura_ft = max(1, min(round(altura_m / 30), 999))
        codigo = f"{cod}{altura_ft:03d}"
        if tipo_nube in ("CB", "TCU"):
            codigo += tipo_nube
        if codigo not in codigos:
            codigos.append(codigo)
    return " ".join(codigos) if codigos else "NSC"

def validar_info_suplementaria(hora, texto):
    if not texto or not texto.strip():
        return False, "Falta información suplementaria obligatoria: precipitación PPxxx"
    partes = texto.strip().upper().split()
    tiene_precip = any(
        p.startswith("PP") and len(p) >= 4
        and p[2:5].replace("T","").replace("R","").replace("Z","").isdigit()
        for p in partes
    )
    if not tiene_precip:
        return False, "Falta precipitación: debe incluir PPxxx (ej: PP000, PP001, PPTRZ)"
    if hora and hora.isdigit() and len(hora) == 4:
        h = int(hora)
        if h == 1200 and not any(p.startswith("TN") for p in partes):
            return False, "Las 12Z requieren temperatura mínima (TNxxx)"
        if h == 2200 and not any(p.startswith("TX") for p in partes):
            return False, "Las 22Z requieren temperatura máxima (TXxxx)"
    return True, ""

def generar_metar(datos):
    try:
        if not datos["dir_viento"] or not datos["int_viento"]:
            raise ValueError("Dirección e intensidad del viento son obligatorias")
        if not datos["vis"]:
            raise ValueError("Visibilidad es obligatoria")
        if not datos["temp"] or not datos["rocio"] or not datos["qnh"]:
            raise ValueError("Temperatura, Rocío y QNH son obligatorios")

        hora = datos["hora"]
        if not hora or len(hora) != 4 or not hora.isdigit():
            raise ValueError("Hora debe ser HHMM (4 dígitos)")

        viento  = procesar_viento(datos["dir_viento"], datos["int_viento"], datos["var_viento"])
        vis_m   = convertir_visibilidad(datos["vis"])
        vis_min_codigo = ""
        if datos["vis_min"]:
            vis_min_codigo, err = procesar_visibilidad_minima(datos["vis_min"], vis_m)
            if err:
                raise ValueError(err)
        rvr_codigo = procesar_rvr(datos["rvr"])
        fenomeno   = " ".join(datos["fenomenos"][:3]) if datos["fenomenos"] else ""
        nubes      = interpretar_nubes_lista(datos["nubes"], vis_m, fenomeno)

        temp  = float(datos["temp"])
        rocio = float(datos["rocio"])
        qnh   = float(datos["qnh"])
        if rocio > temp:
            raise ValueError(f"Rocío ({rocio}°C) no puede ser > Temperatura ({temp}°C)")
        if not (-10 <= temp <= 40):
            raise ValueError("Temperatura fuera de rango (-10 a 40°C)")
        if not (850 <= qnh <= 1100):
            raise ValueError("QNH fuera de rango (850-1100 hPa)")

        es_valida, err_sup = validar_info_suplementaria(hora, datos["suplementaria"])
        if not es_valida:
            raise ValueError(err_sup)

        t_m = redondear_metar(temp)
        r_m = redondear_metar(rocio)
        q_m = int(qnh)

        partes = [f"{datos['tipo']} SPJC {datos['dia']}{hora}Z {viento}"]
        if nubes == "CAVOK":
            partes.append("CAVOK")
        else:
            partes.append(f"{vis_m:04d}")
            if vis_min_codigo: partes.append(vis_min_codigo)
            if rvr_codigo:     partes.append(rvr_codigo)
            if fenomeno:       partes.append(fenomeno)
            partes.append(nubes)
        partes.append(f"{t_m:02d}/{r_m:02d} Q{q_m}")
        sup = datos["suplementaria"].strip().upper() if datos["suplementaria"] else ""
        if sup:
            partes.append(sup)
        metar = " ".join(partes) + "="

        registro = {
            "Día": str(datos["dia"]).zfill(2),
            "Hora": hora,
            "Tipo": datos["tipo"],
            "Dirección_Viento":    datos["dir_viento"],
            "Intensidad_Viento":   datos["int_viento"],
            "Variación_Viento":    datos["var_viento"],
            "Visibilidad_Original":datos["vis"],
            "Visibilidad_Metros":  vis_m,
            "Visibilidad_Mínima":  vis_min_codigo,
            "RVR":                 rvr_codigo,
            "Fenómeno_Texto":      fenomeno,
            "Fenómeno_Código":     fenomeno,
            "Nubes_Texto":         str(datos["nubes"]),
            "Nubes_Código":        nubes,
            "Temperatura":         temp,
            "Punto_Rocío":         rocio,
            "Humedad_Relativa_%":  datos.get("hr", ""),
            "QNH":                 qnh,
            "Presión_Estación":    datos.get("presion", ""),
            "Info_Suplementaria":  datos.get("suplementaria", ""),
            "METAR_Completo":      metar,
        }
        return {"success": True, "metar": metar, "registro": registro}
    except Exception as e:
        return {"success": False, "error": str(e)}

def actualizar_o_insertar(registros, nuevo):
    clave = f"{str(nuevo['Día']).zfill(2)}_{str(nuevo['Hora']).zfill(4)}"
    for i, r in enumerate(registros):
        if f"{str(r.get('Día','')).zfill(2)}_{str(r.get('Hora','')).zfill(4)}" == clave:
            registros[i] = nuevo
            return "actualizado"
    registros.insert(0, nuevo)
    return "insertado"


# ─────────────────────────────────────────────
# RUTAS
# ─────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        usuario  = request.form.get("usuario", "").strip()
        password = request.form.get("password", "")
        if usuario in USUARIOS and hmac.compare_digest(password, USUARIOS[usuario]):
            session.clear()
            session["usuario"] = usuario
            sesion_init()
            return redirect(url_for("index"))
        error = "Usuario o contraseña incorrectos"
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/", methods=["GET"])
def index():
    if "usuario" not in session:
        return redirect(url_for("login"))
    sesion_init()
    return render_template("index.html",
        usuario          = session["usuario"],
        historial        = session["historial"][:10],
        contador         = len(session["registros"]),
        ultimo_metar     = session.get("ultimo_metar"),
        ultimo_tipo      = session.get("ultimo_tipo"),
        fenomenos_lista  = session["fenomenos_lista"],
        nubes_lista      = session["nubes_lista"],
        fenomenos_opciones = FENOMENOS_OPCIONES,
        tipos_nube       = TIPOS_NUBE,
        octas_labels     = OCTAS_LABELS,
        hoy              = datetime.now(timezone.utc).strftime("%d/%m/%Y"),
        dia_hoy          = datetime.now(timezone.utc).strftime("%d"),
        archivo_mes      = obtener_nombre_archivo(),
        mensaje          = session.pop("mensaje", None),
        tipo_mensaje     = session.pop("tipo_mensaje", None),
        form_data        = session.pop("form_data", {}),
        error_metar      = session.pop("error_metar", None),
    )

@app.route("/generar", methods=["POST"])
def generar():
    if "usuario" not in session:
        return redirect(url_for("login"))
    sesion_init()

    # Guardar datos del form para repoblar si hay error
    form_data = {k: v for k, v in request.form.items()}
    session["form_data"] = form_data

    datos = {
        "tipo":         request.form.get("tipo", "METAR"),
        "dia":          request.form.get("dia", "").strip(),
        "hora":         request.form.get("hora", "").strip(),
        "dir_viento":   request.form.get("dir_viento", "").strip(),
        "int_viento":   request.form.get("int_viento", "").strip(),
        "var_viento":   request.form.get("var_viento", "").strip(),
        "vis":          request.form.get("vis", "").strip(),
        "vis_min":      request.form.get("vis_min", "").strip(),
        "rvr":          request.form.get("rvr", "").strip(),
        "fenomenos":    session["fenomenos_lista"],
        "nubes":        session["nubes_lista"],
        "temp":         request.form.get("temp", "").strip(),
        "rocio":        request.form.get("rocio", "").strip(),
        "hr":           request.form.get("hr", "").strip(),
        "qnh":          request.form.get("qnh", "").strip(),
        "presion":      request.form.get("presion", "").strip(),
        "suplementaria":request.form.get("suplementaria", "").strip(),
    }

    resultado = generar_metar(datos)

    if resultado["success"]:
        accion = actualizar_o_insertar(session["registros"], resultado["registro"])
        guardar_registros_mes(session["registros"])

        # Actualizar historial
        metar     = resultado["metar"]
        clave_new = f"{resultado['registro']['Día']}_{resultado['registro']['Hora']}"
        hist      = [m for m in session["historial"]
                     if not (lambda mo: mo and f"{mo.group(1)}_{mo.group(2)}" == clave_new)
                        (re.search(r"SPJC (\d{2})(\d{4})Z", m))]
        hist.insert(0, metar)
        session["historial"]    = hist[:20]
        session["ultimo_metar"] = metar
        session["ultimo_tipo"]  = datos["tipo"]
        session["fenomenos_lista"] = []
        session["nubes_lista"]     = []
        session["form_data"]       = {}

        msg = "METAR ACTUALIZADO" if accion == "actualizado" else "METAR generado correctamente"
        session["mensaje"]      = msg
        session["tipo_mensaje"] = "warning" if accion == "actualizado" else "success"
    else:
        session["error_metar"] = resultado["error"]

    session.modified = True
    return redirect(url_for("index"))

@app.route("/fenomeno/agregar", methods=["POST"])
def agregar_fenomeno():
    if "usuario" not in session:
        return redirect(url_for("login"))
    sesion_init()
    codigo = request.form.get("fenomeno_codigo", "").strip()
    if codigo and codigo not in session["fenomenos_lista"] and len(session["fenomenos_lista"]) < 3:
        session["fenomenos_lista"].append(codigo)
        session.modified = True
    return redirect(url_for("index"))

@app.route("/fenomeno/eliminar/<int:idx>", methods=["POST"])
def eliminar_fenomeno(idx):
    if "usuario" not in session:
        return redirect(url_for("login"))
    sesion_init()
    if 0 <= idx < len(session["fenomenos_lista"]):
        session["fenomenos_lista"].pop(idx)
        session.modified = True
    return redirect(url_for("index"))

@app.route("/nube/agregar", methods=["POST"])
def agregar_nube():
    if "usuario" not in session:
        return redirect(url_for("login"))
    sesion_init()
    if len(session["nubes_lista"]) < 4:
        session["nubes_lista"].append({
            "octas":    int(request.form.get("octas", 1)),
            "tipo":     request.form.get("tipo_nube", "SC"),
            "altura_m": int(request.form.get("altura_m", 300)),
        })
        session.modified = True
    return redirect(url_for("index"))

@app.route("/nube/eliminar/<int:idx>", methods=["POST"])
def eliminar_nube(idx):
    if "usuario" not in session:
        return redirect(url_for("login"))
    sesion_init()
    if 0 <= idx < len(session["nubes_lista"]):
        session["nubes_lista"].pop(idx)
        session.modified = True
    return redirect(url_for("index"))

@app.route("/exportar")
def exportar():
    if "usuario" not in session:
        return redirect(url_for("login"))
    sesion_init()
    registros = session.get("registros", [])
    if not registros:
        session["mensaje"]      = "No hay registros para exportar"
        session["tipo_mensaje"] = "warning"
        return redirect(url_for("index"))
    guardar_registros_mes(registros)
    archivo = DIRECTORIO_DATOS / obtener_nombre_archivo()
    if archivo.exists():
        return send_file(str(archivo.resolve()),
                         as_attachment=True,
                         download_name=obtener_nombre_archivo())
    session["mensaje"]      = "Error al generar el archivo"
    session["tipo_mensaje"] = "error"
    return redirect(url_for("index"))

@app.route("/limpiar_memoria", methods=["POST"])
def limpiar_memoria():
    if "usuario" not in session:
        return redirect(url_for("login"))
    session["registros"] = []
    session["historial"] = []
    session["mensaje"]      = "Memoria limpiada"
    session["tipo_mensaje"] = "success"
    session.modified = True
    return redirect(url_for("index"))


# ─────────────────────────────────────────────
# ARRANQUE
# ─────────────────────────────────────────────
if __name__ == "__main__":
    app.run(debug=True, port=5000)
